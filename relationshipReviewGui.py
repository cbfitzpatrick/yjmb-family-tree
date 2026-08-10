#!/usr/bin/env python3
"""
relationshipReviewGui.py

Local visual review tool for YJMB RAT/VET relationship cleanup.

Purpose
-------
The tool scans the master worksheet (default: ``People on Tree``) in
``YJMB Trees.xlsx`` and finds relationship references that do not resolve to
exactly one unique person row.  Each unresolved VET/RAT reference can then be
reviewed in a Tkinter GUI:

* compare the reference against similar existing people;
* explicitly choose an existing person or create a new person;
* correct the referenced person's name, RAT year, or section/instrument;
* optionally make the reciprocal VET/RAT field consistent;
* preview a one-hop local tree containing the reviewed person, their VET(s),
  and their RAT(s);
* answer "Does this look correct?" before anything is saved;
* skip an item without changing the workbook.

Safety
------
No fuzzy match is accepted automatically.  The workbook is changed only after
pressing "Yes — save and continue" on the preview screen.  Before the first
write, a timestamped backup is created under ``backups/relationship_review``.

The script edits only the selected master worksheet.  Other section worksheets
are left untouched.

Dependencies
------------
* Python 3.10+
* openpyxl
* Tkinter (included with normal Windows Python installations)

Typical use
-----------
Place this file in the same directory as ``YJMB Trees.xlsx`` and run:

    python .\\relationshipReviewGui.py

Optional scan-only mode (does not open a GUI or modify the workbook):

    python .\\relationshipReviewGui.py --scan-only
"""

from __future__ import annotations

import argparse
import copy
import difflib
import os
import re
import shutil
import sys
import tempfile
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError as exc:  # pragma: no cover - Windows Python normally includes Tk
    raise SystemExit(
        "Tkinter is not available in this Python installation. "
        "On Windows, reinstall Python with the optional Tcl/Tk components enabled."
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "YJMB Trees.xlsx"
DEFAULT_SHEET = "People on Tree"

# The full-band project's existing section palette.
SECTION_COLORS: dict[str, str] = {
    "flute/piccolo": "#f5c1ce",
    "clarinet": "#ECD5E3",
    "sax/saxophone": "#F3B0C3",
    "trumpet": "#fa9078",
    "mellophone": "#fcb17e",
    "trombone": "#7cf7b7",
    "baritone": "#8bf0e6",
    "sousaphone": "#f797d2",
    "front ensemble": "#C6DBDA",
    "battery": "#D4F0F0",
    "guard": "#FF968A",
    "goldrush": "#F6EAC2",
    "golden girl": "#FFFFB5",
    "unknown": "#D9D9D9",
}

SECTION_ALIASES: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("flute/piccolo", re.compile(r"\b(?:flutes?|piccolos?)\b", re.I)),
    ("clarinet", re.compile(r"\bclarinets?\b", re.I)),
    ("sax/saxophone", re.compile(r"\b(?:sax(?:ophone)?s?|alto\s+sax(?:ophone)?s?|tenor\s+sax(?:ophone)?s?|baritone\s+sax(?:ophone)?s?|bari\s+sax(?:ophone)?s?)\b", re.I)),
    ("trumpet", re.compile(r"\btrumpets?\b", re.I)),
    ("mellophone", re.compile(r"\b(?:mellos?|mellophones?)\b", re.I)),
    ("trombone", re.compile(r"\b(?:trombones?|bones?)\b", re.I)),
    ("baritone", re.compile(r"\b(?:baritones?|euphoniums?)\b", re.I)),
    ("sousaphone", re.compile(r"\b(?:sousaphones?|tubas?)\b", re.I)),
    ("front ensemble", re.compile(r"\b(?:front\s+ensemble|pit)\b", re.I)),
    ("battery", re.compile(r"\b(?:battery|drum\s*line|drumline|snare|tenors?|quads?|bass\s+drums?|cymbals?)\b", re.I)),
    ("guard", re.compile(r"\b(?:color\s+guard|guard)\b", re.I)),
    ("goldrush", re.compile(r"\bgold\s*rush\b|\bgoldrush\b", re.I)),
    ("golden girl", re.compile(r"\bgolden\s+girls?\b", re.I)),
)

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "given": ("givenpreferredname", "givenname", "preferredname", "firstname"),
    "nickname": ("nickname", "nick name"),
    "family": ("familymaidenname", "familyname", "maidenname", "lastname", "surname"),
    "married": ("marriedname", "marriedsurname", "spousesurname", "currentlastname"),
    "year": ("ratyear", "year"),
    "instrument": ("instrument", "instruments", "section"),
    "vet": ("vet", "vetsnameratyearandinstruments"),
}

RELATION_RE = re.compile(r"^\s*(.*?)\s*\((.*?)\)\s*\(([^()]*)\)\s*$")
YEAR_RE = re.compile(r"(?<!\d)((?:19|20)\d{2})(?!\d)")

CARD_WIDTH = 150
CARD_HEIGHT = 80
CARD_GAP = 30
TREE_MARGIN_X = 105
TREE_MARGIN_Y = 70
YEAR_BAND_HEIGHT = 105
CONNECTOR_COLOR = "#777777"
CONNECTOR_WIDTH = 4


# ---------------------------------------------------------------------------
# Normalization and data objects
# ---------------------------------------------------------------------------


def normalize_spaces(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u00a0", " ").replace("’", "'").replace("`", "'")
    return re.sub(r"\s+", " ", text).strip()


def normalized_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_spaces(value).casefold())


def loose_name_key(value: object) -> str:
    text = unicodedata.normalize("NFKD", normalize_spaces(value).casefold())
    return "".join(char for char in text if char.isalnum())


def stable_name(given: str, family: str) -> str:
    return normalize_spaces(f"{given} {family}") or given or family


def parse_year(value: object) -> int | None:
    text = normalize_spaces(value)
    match = YEAR_RE.search(text)
    if not match:
        return None
    return int(match.group(1))


def normalize_sections(raw: object) -> list[str]:
    text = normalize_spaces(raw)
    if not text:
        return ["unknown"]
    found: list[tuple[int, str]] = []
    for section, pattern in SECTION_ALIASES:
        for match in pattern.finditer(text):
            found.append((match.start(), section))
    found.sort(key=lambda item: item[0])
    result: list[str] = []
    for _, section in found:
        if section not in result:
            result.append(section)
    return result or ["unknown"]


def section_overlap(left: object, right: object) -> bool:
    return bool(set(normalize_sections(left)) & set(normalize_sections(right)) - {"unknown"})


@dataclass(frozen=True)
class RelationReference:
    raw: str
    name: str
    year: int | None
    instrument: str


@dataclass
class PersonRecord:
    row: int
    given: str
    nickname: str
    family: str
    married: str
    year: int | None
    instrument: str
    vet_raw: str
    rat_raws: list[tuple[str, str]]

    @property
    def name(self) -> str:
        return stable_name(self.given, self.family)

    @property
    def aliases(self) -> set[str]:
        aliases = {self.name}
        if self.married:
            aliases.add(stable_name(self.given, self.married))
        if self.nickname:
            aliases.add(stable_name(self.nickname, self.family))
            if self.married:
                aliases.add(stable_name(self.nickname, self.married))
        return {alias for alias in aliases if alias}

    @property
    def canonical_relation(self) -> str:
        return format_relation(self.name, self.year, self.instrument)


@dataclass
class CandidateMatch:
    person: PersonRecord
    score: float
    explanation: str


@dataclass
class ReviewItem:
    source_row: int
    source_name: str
    column: int
    column_label: str
    role: str  # "VET" or "RAT"
    raw: str
    reference: RelationReference
    exact_candidate_rows: list[int] = field(default_factory=list)

    @property
    def key(self) -> tuple[int, int, str]:
        return (self.source_row, self.column, self.raw)


@dataclass
class ResolutionDraft:
    action: str  # existing | new
    target_row: int | None
    given: str
    family: str
    year: int | None
    instrument: str
    ensure_reciprocal: bool
    replace_conflicting_vet: bool
    update_all_references: bool

    @property
    def name(self) -> str:
        return stable_name(self.given, self.family)

    @property
    def canonical_relation(self) -> str:
        return format_relation(self.name, self.year, self.instrument)


@dataclass
class PreviewPerson:
    key: str
    name: str
    year: int | None
    instrument: str
    unresolved: bool = False
    proposed: bool = False


@dataclass
class LocalPreview:
    center: PreviewPerson
    vets: list[PreviewPerson]
    rats: list[PreviewPerson]
    warnings: list[str]


# ---------------------------------------------------------------------------
# Workbook model
# ---------------------------------------------------------------------------


def parse_relation(raw_value: object) -> RelationReference | None:
    raw = normalize_spaces(raw_value)
    if not raw:
        return None
    match = RELATION_RE.match(raw)
    if match:
        name = normalize_spaces(match.group(1))
        year = parse_year(match.group(2))
        instrument = normalize_spaces(match.group(3))
    else:
        name = normalize_spaces(raw.split(" (", 1)[0])
        year = parse_year(raw)
        instrument = ""
    if not name:
        return None
    return RelationReference(raw=raw, name=name, year=year, instrument=instrument)


def format_relation(name: str, year: int | None, instrument: str) -> str:
    return f"{normalize_spaces(name)} ({year or ''}) ({normalize_spaces(instrument)})"


class WorkbookModel:
    def __init__(self, path: Path, sheet_name: str = DEFAULT_SHEET) -> None:
        self.path = path.resolve()
        self.sheet_name = sheet_name
        self.workbook = None
        self.ws = None
        self.header_row = 1
        self.mapping: dict[str, int] = {}
        self.rat_columns: list[tuple[str, int]] = []
        self.header_labels: dict[int, str] = {}
        self.people: list[PersonRecord] = []
        self.people_by_row: dict[int, PersonRecord] = {}
        self._backup_path: Path | None = None
        self.load()

    def load(self) -> None:
        if self.workbook is not None:
            try:
                self.workbook.close()
            except Exception:
                pass
        self.workbook = load_workbook(self.path, data_only=False, read_only=False)
        if self.sheet_name not in self.workbook.sheetnames:
            possible = []
            for candidate in self.workbook.sheetnames:
                try:
                    self._discover_headers(self.workbook[candidate])
                except ValueError:
                    continue
                possible.append(candidate)
            if not possible:
                raise ValueError(
                    f"No worksheet contains the expected name/year/VET/RAT headers. "
                    f"Available sheets: {', '.join(self.workbook.sheetnames)}"
                )
            self.sheet_name = possible[0]
        self.ws = self.workbook[self.sheet_name]
        self.header_row, self.mapping, self.rat_columns = self._discover_headers(self.ws)
        self.header_labels = {
            col: normalize_spaces(self.ws.cell(self.header_row, col).value) or f"Column {col}"
            for col in range(1, self.ws.max_column + 1)
        }
        self.refresh_people()

    @staticmethod
    def _discover_headers(ws) -> tuple[int, dict[str, int], list[tuple[str, int]]]:
        header_row = None
        headers: dict[int, str] = {}
        for row in range(1, min(ws.max_row, 10) + 1):
            candidate = {
                col: normalized_header(ws.cell(row, col).value)
                for col in range(1, ws.max_column + 1)
            }
            values = set(candidate.values())
            if "givenpreferredname" in values and "familymaidenname" in values:
                header_row = row
                headers = candidate
                break
        if header_row is None:
            raise ValueError("Four-column name headers were not found.")

        mapping: dict[str, int] = {}
        for field_name, aliases in HEADER_ALIASES.items():
            keys = {normalized_header(alias) for alias in aliases}
            for col, header in headers.items():
                if header in keys:
                    mapping[field_name] = col
                    break
        required = ("given", "nickname", "family", "married", "year", "instrument", "vet")
        missing = [field_name for field_name in required if field_name not in mapping]
        if missing:
            raise ValueError("Missing required columns: " + ", ".join(missing))

        rat_columns: list[tuple[str, int]] = []
        for col, header in headers.items():
            match = re.fullmatch(r"rat(\d+)", header)
            if match:
                rat_columns.append((f"RAT {int(match.group(1))}", col))
        rat_columns.sort(key=lambda item: int(item[0].split()[1]))
        if not rat_columns:
            raise ValueError("No RAT relationship columns were found.")
        return header_row, mapping, rat_columns

    def refresh_people(self) -> None:
        assert self.ws is not None
        people: list[PersonRecord] = []
        for row in range(self.header_row + 1, self.ws.max_row + 1):
            given = normalize_spaces(self.ws.cell(row, self.mapping["given"]).value)
            nickname = normalize_spaces(self.ws.cell(row, self.mapping["nickname"]).value)
            family = normalize_spaces(self.ws.cell(row, self.mapping["family"]).value)
            married = normalize_spaces(self.ws.cell(row, self.mapping["married"]).value)
            if not any((given, nickname, family, married)):
                continue
            person = PersonRecord(
                row=row,
                given=given,
                nickname=nickname,
                family=family,
                married=married,
                year=parse_year(self.ws.cell(row, self.mapping["year"]).value),
                instrument=normalize_spaces(self.ws.cell(row, self.mapping["instrument"]).value),
                vet_raw=normalize_spaces(self.ws.cell(row, self.mapping["vet"]).value),
                rat_raws=[
                    (label, normalize_spaces(self.ws.cell(row, col).value))
                    for label, col in self.rat_columns
                    if normalize_spaces(self.ws.cell(row, col).value)
                ],
            )
            people.append(person)
        self.people = people
        self.people_by_row = {person.row: person for person in people}

    def unique_candidates(self, reference: RelationReference) -> list[PersonRecord]:
        """Return the strongest exact-name candidates for a relationship reference.

        A relationship is considered uniquely resolved when exactly one candidate
        survives name + (when available) RAT-year filtering. Instrument is used as
        an additional disambiguator only when it reduces a multi-candidate set.
        """
        key = loose_name_key(reference.name)
        candidates = [
            person
            for person in self.people
            if any(loose_name_key(alias) == key for alias in person.aliases)
        ]
        if reference.year is not None:
            by_year = [person for person in candidates if person.year == reference.year]
            if by_year:
                candidates = by_year
        if len(candidates) > 1 and reference.instrument:
            by_section = [
                person for person in candidates if section_overlap(reference.instrument, person.instrument)
            ]
            if by_section:
                candidates = by_section
        return candidates

    def resolve_unique(self, reference: RelationReference) -> PersonRecord | None:
        candidates = self.unique_candidates(reference)
        return candidates[0] if len(candidates) == 1 else None

    def scan_review_items(self, skipped: set[tuple[int, int, str]] | None = None) -> list[ReviewItem]:
        skipped = skipped or set()
        items: list[ReviewItem] = []
        for person in self.people:
            relation_cells = [("VET", self.mapping["vet"], person.vet_raw)]
            relation_cells.extend(
                (label, col, normalize_spaces(self.ws.cell(person.row, col).value))
                for label, col in self.rat_columns
                if normalize_spaces(self.ws.cell(person.row, col).value)
            )
            for label, col, raw in relation_cells:
                if not raw:
                    continue
                key = (person.row, col, raw)
                if key in skipped:
                    continue
                reference = parse_relation(raw)
                if not reference:
                    items.append(
                        ReviewItem(
                            source_row=person.row,
                            source_name=person.name,
                            column=col,
                            column_label=label,
                            role="VET" if label == "VET" else "RAT",
                            raw=raw,
                            reference=RelationReference(raw=raw, name=raw, year=None, instrument=""),
                            exact_candidate_rows=[],
                        )
                    )
                    continue
                candidates = self.unique_candidates(reference)
                if len(candidates) != 1:
                    items.append(
                        ReviewItem(
                            source_row=person.row,
                            source_name=person.name,
                            column=col,
                            column_label=label,
                            role="VET" if label == "VET" else "RAT",
                            raw=raw,
                            reference=reference,
                            exact_candidate_rows=[candidate.row for candidate in candidates],
                        )
                    )
        items.sort(key=lambda item: (item.source_row, item.column))
        return items

    def fuzzy_candidates(self, reference: RelationReference, limit: int = 10) -> list[CandidateMatch]:
        ref_name = loose_name_key(reference.name)
        scored: list[CandidateMatch] = []
        for person in self.people:
            best_name_ratio = 0.0
            for alias in person.aliases:
                ratio = difflib.SequenceMatcher(None, ref_name, loose_name_key(alias)).ratio()
                best_name_ratio = max(best_name_ratio, ratio)
            score = best_name_ratio * 100.0
            explanations = [f"name {best_name_ratio * 100:.0f}%"]
            if reference.year is not None and person.year is not None:
                delta = abs(reference.year - person.year)
                if delta == 0:
                    score += 18
                    explanations.append("same RAT year")
                elif delta == 1:
                    score += 8
                    explanations.append("RAT year ±1")
                elif delta <= 3:
                    score += 3
            if reference.instrument and person.instrument and section_overlap(reference.instrument, person.instrument):
                score += 9
                explanations.append("same section")
            # Exact candidate rows (for duplicate-name ambiguity) should always be visible.
            if person.row in self.unique_candidates(reference):
                score += 25
                explanations.append("exact-name candidate")
            scored.append(CandidateMatch(person=person, score=score, explanation=", ".join(explanations)))
        scored.sort(key=lambda match: (-match.score, match.person.year or 9999, match.person.family.casefold(), match.person.given.casefold()))
        return scored[:limit]

    def relationship_cells_resolving_to(self, target_row: int) -> list[tuple[int, int]]:
        locations: list[tuple[int, int]] = []
        for person in self.people:
            columns = [self.mapping["vet"], *(col for _, col in self.rat_columns)]
            for col in columns:
                raw = normalize_spaces(self.ws.cell(person.row, col).value)
                reference = parse_relation(raw)
                if not reference:
                    continue
                resolved = self.resolve_unique(reference)
                if resolved and resolved.row == target_row:
                    locations.append((person.row, col))
        return locations

    def build_preview(self, item: ReviewItem, draft: ResolutionDraft) -> LocalPreview:
        source = self.people_by_row[item.source_row]
        if draft.action == "existing" and draft.target_row is not None:
            original = self.people_by_row[draft.target_row]
            center = PreviewPerson(
                key=f"row-{original.row}",
                name=draft.name,
                year=draft.year,
                instrument=draft.instrument,
                proposed=True,
            )
            existing_vet_raw = original.vet_raw
            existing_rat_raws = list(original.rat_raws)
        else:
            center = PreviewPerson(
                key="new-person",
                name=draft.name,
                year=draft.year,
                instrument=draft.instrument,
                proposed=True,
            )
            existing_vet_raw = ""
            existing_rat_raws = []

        warnings: list[str] = []
        vets: list[PreviewPerson] = []
        rats: list[PreviewPerson] = []

        # Existing one-hop relationships for the target person.
        if existing_vet_raw:
            ref = parse_relation(existing_vet_raw)
            if ref:
                resolved = self.resolve_unique(ref)
                vets.append(self._preview_from_ref(ref, resolved, suffix="existing-vet"))

        for index, (_, raw) in enumerate(existing_rat_raws, start=1):
            ref = parse_relation(raw)
            if not ref:
                continue
            resolved = self.resolve_unique(ref)
            rats.append(self._preview_from_ref(ref, resolved, suffix=f"existing-rat-{index}"))

        source_preview = PreviewPerson(
            key=f"source-{source.row}",
            name=source.name,
            year=source.year,
            instrument=source.instrument,
            proposed=True,
        )

        if item.role == "VET":
            # The reviewed target is the source person's VET, so source is a RAT of center.
            if not any(self._same_preview_person(person, source_preview) for person in rats):
                rats.append(source_preview)
        else:
            # The reviewed target is a RAT of source, so source is center's VET.
            existing_different = [
                person for person in vets if not self._same_preview_person(person, source_preview)
            ]
            if existing_different:
                warnings.append(
                    "The selected/new RAT already has a different VET in its own row. "
                    "The proposed VET is shown too. Reciprocal replacement will occur only if you enable it."
                )
            if not any(self._same_preview_person(person, source_preview) for person in vets):
                vets.append(source_preview)

        vets.sort(key=lambda person: (person.year or 9999, person.name.casefold()))
        rats.sort(key=lambda person: (person.year or 9999, person.name.casefold()))
        return LocalPreview(center=center, vets=vets, rats=rats, warnings=warnings)

    def _preview_from_ref(
        self, reference: RelationReference, resolved: PersonRecord | None, *, suffix: str
    ) -> PreviewPerson:
        if resolved:
            return PreviewPerson(
                key=f"row-{resolved.row}-{suffix}",
                name=resolved.name,
                year=resolved.year,
                instrument=resolved.instrument,
                unresolved=False,
            )
        return PreviewPerson(
            key=f"unresolved-{suffix}-{loose_name_key(reference.name)}",
            name=reference.name,
            year=reference.year,
            instrument=reference.instrument,
            unresolved=True,
        )

    @staticmethod
    def _same_preview_person(left: PreviewPerson, right: PreviewPerson) -> bool:
        return (
            loose_name_key(left.name) == loose_name_key(right.name)
            and (left.year is None or right.year is None or left.year == right.year)
        )

    def validate_draft(self, item: ReviewItem, draft: ResolutionDraft) -> list[str]:
        errors: list[str] = []
        if not draft.given.strip():
            errors.append("Given/Preferred Name is required.")
        if not draft.family.strip():
            errors.append("Family/Maiden Name is required.")
        if draft.year is None:
            errors.append("A four-digit RAT year is required for a unique person entry.")
        else:
            current_year = datetime.now().year
            if not (1908 <= draft.year <= current_year + 1):
                errors.append(f"RAT year must be between 1908 and {current_year + 1}.")
        if draft.action == "existing":
            if draft.target_row is None or draft.target_row not in self.people_by_row:
                errors.append("Select an existing person from the suggestion list.")
        elif draft.action == "new":
            conflicts = [
                person
                for person in self.people
                if loose_name_key(person.name) == loose_name_key(draft.name)
                and person.year == draft.year
            ]
            if conflicts:
                errors.append(
                    "A person with the same name and RAT year already exists. "
                    "Choose that existing row or correct the new-person details."
                )
        else:
            errors.append("Choose whether this is an existing person or a new person.")

        if item.role == "RAT" and draft.action == "existing" and draft.target_row:
            target = self.people_by_row[draft.target_row]
            if target.vet_raw:
                ref = parse_relation(target.vet_raw)
                resolved = self.resolve_unique(ref) if ref else None
                if resolved and resolved.row != item.source_row and draft.ensure_reciprocal and not draft.replace_conflicting_vet:
                    # Not a hard error: the source relationship can still be corrected while
                    # leaving the target's existing VET untouched. Warn in GUI instead.
                    pass
        return errors

    def commit_resolution(self, item: ReviewItem, draft: ResolutionDraft) -> tuple[int, list[str]]:
        """Apply a confirmed resolution and save atomically.

        Returns ``(target_row, messages)``.

        The workbook is checked for an *actual* Windows write lock before any
        in-memory edits are made.  Excel's ``~$...xlsx`` owner file is only a
        hint and can be left behind after a crash, so its mere presence must
        not prevent a save.
        """
        assert self.ws is not None
        self._assert_workbook_writable()
        messages: list[str] = []
        source = self.people_by_row[item.source_row]

        # Collect relationship locations before editing an existing person's identity.
        related_locations: list[tuple[int, int]] = []
        if draft.action == "existing" and draft.target_row is not None and draft.update_all_references:
            related_locations = self.relationship_cells_resolving_to(draft.target_row)

        if draft.action == "new":
            target_row = self._append_new_person(draft)
            messages.append(f"Created new person row {target_row}: {draft.name}.")
        else:
            assert draft.target_row is not None
            target_row = draft.target_row
            self._update_person_identity(target_row, draft)
            messages.append(f"Updated existing person row {target_row}: {draft.name}.")

        # Rewrite the currently reviewed source cell to the target's confirmed identity.
        canonical = draft.canonical_relation
        self.ws.cell(item.source_row, item.column).value = canonical
        messages.append(
            f"Updated {item.source_name} {item.column_label} to {canonical}."
        )

        # If an existing person's identity/year/section changed, keep all references
        # that uniquely pointed to that row consistent with the new canonical text.
        if draft.action == "existing" and draft.update_all_references:
            for row, col in related_locations:
                self.ws.cell(row, col).value = canonical
            if related_locations:
                messages.append(
                    f"Updated {len(related_locations)} other relationship reference(s) to the confirmed identity."
                )

        reciprocal_written = False
        if draft.ensure_reciprocal:
            if item.role == "VET":
                reciprocal_written = self._ensure_rat_reference(
                    parent_row=target_row,
                    child_row=item.source_row,
                    messages=messages,
                )
            else:
                reciprocal_written = self._ensure_vet_reference(
                    child_row=target_row,
                    parent_row=item.source_row,
                    replace_conflict=draft.replace_conflicting_vet,
                    messages=messages,
                )

        self._mark_status_if_present(item, reciprocal_written)
        self._backup_once()
        self._save_atomic()
        self.refresh_people()
        return target_row, messages

    def _update_person_identity(self, row: int, draft: ResolutionDraft) -> None:
        self.ws.cell(row, self.mapping["given"]).value = draft.given
        self.ws.cell(row, self.mapping["family"]).value = draft.family
        self.ws.cell(row, self.mapping["year"]).value = draft.year
        self.ws.cell(row, self.mapping["instrument"]).value = draft.instrument

    def _append_new_person(self, draft: ResolutionDraft) -> int:
        assert self.ws is not None
        old_max_row = self.ws.max_row
        row = old_max_row + 1

        # Copy style from the previous populated person row for visual consistency.
        source_style_row = max(self.header_row + 1, old_max_row)
        for col in range(1, self.ws.max_column + 1):
            src = self.ws.cell(source_style_row, col)
            dst = self.ws.cell(row, col)
            if src.has_style:
                dst._style = copy.copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)

        self.ws.cell(row, self.mapping["given"]).value = draft.given
        self.ws.cell(row, self.mapping["nickname"]).value = None
        self.ws.cell(row, self.mapping["family"]).value = draft.family
        self.ws.cell(row, self.mapping["married"]).value = None
        self.ws.cell(row, self.mapping["year"]).value = draft.year
        self.ws.cell(row, self.mapping["instrument"]).value = draft.instrument
        self.ws.cell(row, self.mapping["vet"]).value = None
        for _, col in self.rat_columns:
            self.ws.cell(row, col).value = None

        # Extend Excel tables that cover the master data area.
        for table in self.ws.tables.values():
            try:
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            except ValueError:
                continue
            if min_row <= self.header_row <= max_row and max_row == old_max_row:
                table.ref = (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(max_col)}{row}"
                )
        if self.ws.auto_filter and self.ws.auto_filter.ref:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(self.ws.auto_filter.ref)
                if max_row == old_max_row:
                    self.ws.auto_filter.ref = (
                        f"{get_column_letter(min_col)}{min_row}:"
                        f"{get_column_letter(max_col)}{row}"
                    )
            except ValueError:
                pass
        return row

    def _person_relation_for_row(self, row: int) -> str:
        given = normalize_spaces(self.ws.cell(row, self.mapping["given"]).value)
        family = normalize_spaces(self.ws.cell(row, self.mapping["family"]).value)
        year = parse_year(self.ws.cell(row, self.mapping["year"]).value)
        instrument = normalize_spaces(self.ws.cell(row, self.mapping["instrument"]).value)
        return format_relation(stable_name(given, family), year, instrument)

    def _ensure_rat_reference(self, parent_row: int, child_row: int, messages: list[str]) -> bool:
        child_relation = self._person_relation_for_row(child_row)
        child_ref = parse_relation(child_relation)
        # Existing equivalent reference?
        for label, col in self.rat_columns:
            raw = normalize_spaces(self.ws.cell(parent_row, col).value)
            if not raw:
                continue
            ref = parse_relation(raw)
            if ref and child_ref and loose_name_key(ref.name) == loose_name_key(child_ref.name) and (
                ref.year is None or child_ref.year is None or ref.year == child_ref.year
            ):
                self.ws.cell(parent_row, col).value = child_relation
                messages.append(f"Confirmed reciprocal {label} on row {parent_row}.")
                return True
        # First blank RAT slot.
        for label, col in self.rat_columns:
            if not normalize_spaces(self.ws.cell(parent_row, col).value):
                self.ws.cell(parent_row, col).value = child_relation
                messages.append(f"Added reciprocal {label} on row {parent_row}.")
                return True
        messages.append(
            f"Could not add reciprocal RAT to row {parent_row}: all existing RAT columns are occupied."
        )
        return False

    def _ensure_vet_reference(
        self,
        child_row: int,
        parent_row: int,
        *,
        replace_conflict: bool,
        messages: list[str],
    ) -> bool:
        parent_relation = self._person_relation_for_row(parent_row)
        vet_cell = self.ws.cell(child_row, self.mapping["vet"])
        existing = normalize_spaces(vet_cell.value)
        if not existing:
            vet_cell.value = parent_relation
            messages.append(f"Added reciprocal VET on row {child_row}.")
            return True
        existing_ref = parse_relation(existing)
        parent_ref = parse_relation(parent_relation)
        if existing_ref and parent_ref and loose_name_key(existing_ref.name) == loose_name_key(parent_ref.name) and (
            existing_ref.year is None or parent_ref.year is None or existing_ref.year == parent_ref.year
        ):
            vet_cell.value = parent_relation
            messages.append(f"Confirmed reciprocal VET on row {child_row}.")
            return True
        if replace_conflict:
            vet_cell.value = parent_relation
            messages.append(
                f"Replaced conflicting VET on row {child_row}: {existing!r} -> {parent_relation!r}."
            )
            return True
        messages.append(
            f"Left row {child_row}'s existing VET unchanged because it conflicts with the proposed relationship: {existing!r}."
        )
        return False

    def _mark_status_if_present(self, item: ReviewItem, reciprocal_written: bool) -> None:
        """If v9-style per-slot status columns already exist, update the source status."""
        status_label = (
            "VET Relationship Status"
            if item.role == "VET"
            else f"{item.column_label} Relationship Status"
        )
        status_key = normalized_header(status_label)
        status_col = None
        for col in range(1, self.ws.max_column + 1):
            if normalized_header(self.ws.cell(self.header_row, col).value) == status_key:
                status_col = col
                break
        if status_col:
            self.ws.cell(item.source_row, status_col).value = (
                "Reciprocated — validated on both profiles"
                if reciprocal_written
                else "Unreciprocated — pending validation from related profile"
            )

    def _backup_once(self) -> None:
        if self._backup_path is not None:
            return
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = self.path.parent / "backups" / "relationship_review" / stamp
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / self.path.name
        shutil.copy2(self.path, backup_path)
        self._backup_path = backup_path

    def _assert_workbook_writable(self) -> None:
        """Raise PermissionError only when the workbook itself is really locked.

        Excel normally creates an owner file named ``~$<workbook>.xlsx`` while
        a workbook is open.  Those owner files occasionally remain after Excel
        closes or crashes, so checking only for that filename creates false
        positives.  Instead, try to open the actual workbook for read/write
        access.  On Windows, also attempt a non-blocking one-byte lock.
        """
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")

        handle = None
        byte_locked = False
        try:
            # On Windows this open normally fails immediately when Excel has
            # opened the workbook with write sharing denied.
            handle = open(self.path, "r+b")

            if os.name == "nt":
                try:
                    import msvcrt

                    handle.seek(0)
                    msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
                    byte_locked = True
                except (OSError, PermissionError) as exc:
                    raise PermissionError(
                        f"{self.path.name} is currently locked by another program. "
                        "Close the workbook in Excel (and any File Explorer preview pane "
                        "or other program using it), then try again."
                    ) from exc
                finally:
                    if byte_locked:
                        try:
                            handle.seek(0)
                            msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
                        except OSError:
                            pass
        except PermissionError as exc:
            # Preserve our clearer message above when one was already created.
            if str(exc).startswith(self.path.name):
                raise
            raise PermissionError(
                f"{self.path.name} is currently locked by another program. "
                "Close the workbook in Excel (and any File Explorer preview pane "
                "or other program using it), then try again."
            ) from exc
        finally:
            if handle is not None:
                try:
                    handle.close()
                except OSError:
                    pass

    def _save_atomic(self) -> None:
        # Re-check immediately before replacement in case Excel was opened after
        # the user entered the confirmation screen.  Do NOT block merely because
        # a stale ``~$...xlsx`` owner file exists.
        self._assert_workbook_writable()

        tmp_fd, tmp_name = tempfile.mkstemp(
            prefix=f".{self.path.stem}_relationship_review_",
            suffix=self.path.suffix,
            dir=self.path.parent,
        )
        os.close(tmp_fd)
        tmp_path = Path(tmp_name)
        try:
            self.workbook.save(tmp_path)
            try:
                os.replace(tmp_path, self.path)
            except PermissionError as exc:
                raise PermissionError(
                    f"Windows would not replace {self.path.name}. Another program may "
                    "have opened the workbook while this review was in progress. "
                    "Close it and try the save again."
                ) from exc
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except OSError:
                    pass


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------


class RelationshipReviewApp(tk.Tk):
    def __init__(self, model: WorkbookModel) -> None:
        super().__init__()
        self.model = model
        self.title("YJMB VET/RAT Relationship Reviewer")
        self.geometry("1320x860")
        self.minsize(1080, 720)

        self.session_skips: set[tuple[int, int, str]] = set()
        self.review_items: list[ReviewItem] = []
        self.current_item: ReviewItem | None = None
        self.current_candidates: list[CandidateMatch] = []
        self.preview_draft: ResolutionDraft | None = None

        self._configure_style()
        self._build_ui()
        self.rescan(select_first=True)

    def _configure_style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass
        style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"))
        style.configure("Heading.TLabel", font=("Segoe UI", 11, "bold"))
        style.configure("Muted.TLabel", foreground="#555555")
        style.configure("Danger.TLabel", foreground="#9c2f2f")
        style.configure("Success.TLabel", foreground="#246b37")

    def _build_ui(self) -> None:
        top = ttk.Frame(self, padding=(12, 10))
        top.pack(fill="x")
        ttk.Label(top, text="YJMB Relationship Reviewer", style="Title.TLabel").pack(side="left")
        self.status_var = tk.StringVar()
        ttk.Label(top, textvariable=self.status_var, style="Muted.TLabel").pack(side="left", padx=18)
        ttk.Button(top, text="Rescan workbook", command=lambda: self.rescan(select_first=False)).pack(side="right")
        ttk.Button(top, text="Reset skipped", command=self.reset_skips).pack(side="right", padx=(0, 8))

        path_text = f"Workbook: {self.model.path}    Sheet: {self.model.sheet_name}"
        ttk.Label(self, text=path_text, style="Muted.TLabel", padding=(12, 0, 12, 8)).pack(fill="x")

        paned = ttk.Panedwindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        left = ttk.Frame(paned, padding=8)
        right = ttk.Frame(paned, padding=8)
        paned.add(left, weight=1)
        paned.add(right, weight=3)

        ttk.Label(left, text="Needs review", style="Heading.TLabel").pack(anchor="w")
        ttk.Label(
            left,
            text="References that do not resolve to exactly one person row.",
            style="Muted.TLabel",
            wraplength=290,
        ).pack(anchor="w", pady=(0, 8))

        queue_frame = ttk.Frame(left)
        queue_frame.pack(fill="both", expand=True)
        self.queue = ttk.Treeview(
            queue_frame,
            columns=("source", "role", "reference"),
            show="headings",
            selectmode="browse",
            height=24,
        )
        self.queue.heading("source", text="Source")
        self.queue.heading("role", text="Field")
        self.queue.heading("reference", text="Unresolved reference")
        self.queue.column("source", width=120, stretch=False)
        self.queue.column("role", width=55, stretch=False)
        self.queue.column("reference", width=190, stretch=True)
        qscroll = ttk.Scrollbar(queue_frame, orient="vertical", command=self.queue.yview)
        self.queue.configure(yscrollcommand=qscroll.set)
        self.queue.pack(side="left", fill="both", expand=True)
        qscroll.pack(side="right", fill="y")
        self.queue.bind("<<TreeviewSelect>>", self._queue_selected)

        self.notebook = ttk.Notebook(right)
        self.notebook.pack(fill="both", expand=True)
        self.resolve_tab = ttk.Frame(self.notebook, padding=12)
        self.preview_tab = ttk.Frame(self.notebook, padding=12)
        self.notebook.add(self.resolve_tab, text="1. Resolve / Correct")
        self.notebook.add(self.preview_tab, text="2. Local Tree Preview")
        self._build_resolve_tab()
        self._build_preview_tab()

    def _build_resolve_tab(self) -> None:
        self.item_title_var = tk.StringVar(value="Select a review item")
        ttk.Label(self.resolve_tab, textvariable=self.item_title_var, style="Heading.TLabel").pack(anchor="w")
        self.item_raw_var = tk.StringVar()
        ttk.Label(self.resolve_tab, textvariable=self.item_raw_var, style="Muted.TLabel", wraplength=820).pack(anchor="w", pady=(2, 12))

        ref_box = ttk.LabelFrame(self.resolve_tab, text="What does the reference appear to say?", padding=10)
        ref_box.pack(fill="x", pady=(0, 10))
        ttk.Label(
            ref_box,
            text="If the original cell itself contains a typo or wrong year/section, correct it here before choosing a person.",
            style="Muted.TLabel",
            wraplength=820,
        ).grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 8))
        self.ref_name_var = tk.StringVar()
        self.ref_year_var = tk.StringVar()
        self.ref_instrument_var = tk.StringVar()
        ttk.Label(ref_box, text="Referenced name").grid(row=1, column=0, sticky="w")
        ttk.Entry(ref_box, textvariable=self.ref_name_var, width=32).grid(row=2, column=0, sticky="ew", padx=(0, 8))
        ttk.Label(ref_box, text="RAT year").grid(row=1, column=1, sticky="w")
        ttk.Entry(ref_box, textvariable=self.ref_year_var, width=12).grid(row=2, column=1, sticky="ew", padx=(0, 8))
        ttk.Label(ref_box, text="Section / instrument").grid(row=1, column=2, sticky="w")
        ttk.Entry(ref_box, textvariable=self.ref_instrument_var, width=28).grid(row=2, column=2, sticky="ew", padx=(0, 8))
        ttk.Button(ref_box, text="Refresh similar names", command=self.refresh_candidates_from_reference).grid(row=2, column=3, sticky="w")
        ref_box.columnconfigure(0, weight=2)
        ref_box.columnconfigure(2, weight=2)

        candidates_box = ttk.LabelFrame(self.resolve_tab, text="Is this one of these existing people?", padding=10)
        candidates_box.pack(fill="both", expand=True, pady=(0, 10))
        ttk.Label(
            candidates_box,
            text="Select a row only if you are confident it is the same person. Similarity is a suggestion, not an automatic correction.",
            style="Muted.TLabel",
            wraplength=820,
        ).pack(anchor="w", pady=(0, 6))
        cand_frame = ttk.Frame(candidates_box)
        cand_frame.pack(fill="both", expand=True)
        self.candidates = ttk.Treeview(
            cand_frame,
            columns=("score", "name", "year", "section", "row", "why"),
            show="headings",
            selectmode="browse",
            height=8,
        )
        for col, title, width in (
            ("score", "Match", 62),
            ("name", "Existing person", 190),
            ("year", "RAT year", 70),
            ("section", "Section", 145),
            ("row", "Row", 50),
            ("why", "Why suggested", 220),
        ):
            self.candidates.heading(col, text=title)
            self.candidates.column(col, width=width, stretch=col in {"name", "section", "why"})
        cscroll = ttk.Scrollbar(cand_frame, orient="vertical", command=self.candidates.yview)
        self.candidates.configure(yscrollcommand=cscroll.set)
        self.candidates.pack(side="left", fill="both", expand=True)
        cscroll.pack(side="right", fill="y")
        self.candidates.bind("<<TreeviewSelect>>", self._candidate_selected)
        self.candidates.bind("<Double-1>", self._candidate_selected)

        target_box = ttk.LabelFrame(self.resolve_tab, text="Resolution", padding=10)
        target_box.pack(fill="x")
        self.action_var = tk.StringVar(value="existing")
        ttk.Radiobutton(
            target_box,
            text="This is an existing person selected above",
            variable=self.action_var,
            value="existing",
            command=self._action_changed,
        ).grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Radiobutton(
            target_box,
            text="This is a new person who needs their own unique row",
            variable=self.action_var,
            value="new",
            command=self._action_changed,
        ).grid(row=0, column=2, columnspan=2, sticky="w", padx=(18, 0))

        ttk.Label(
            target_box,
            text="Does anything need to be corrected? These are the values that will identify this person's row and relationship references.",
            style="Muted.TLabel",
            wraplength=820,
        ).grid(row=1, column=0, columnspan=6, sticky="w", pady=(8, 6))

        self.target_given_var = tk.StringVar()
        self.target_family_var = tk.StringVar()
        self.target_year_var = tk.StringVar()
        self.target_instrument_var = tk.StringVar()
        labels_vars = (
            ("First / Preferred Name", self.target_given_var, 0, 22),
            ("Family / Maiden Name", self.target_family_var, 1, 22),
            ("RAT Year", self.target_year_var, 2, 10),
            ("Section / Instrument", self.target_instrument_var, 3, 26),
        )
        for label, variable, column, width in labels_vars:
            ttk.Label(target_box, text=label).grid(row=2, column=column, sticky="w", padx=(0, 8))
            ttk.Entry(target_box, textvariable=variable, width=width).grid(row=3, column=column, sticky="ew", padx=(0, 8))

        self.ensure_reciprocal_var = tk.BooleanVar(value=True)
        self.update_all_refs_var = tk.BooleanVar(value=True)
        self.replace_conflicting_vet_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            target_box,
            text="Ensure the reciprocal VET/RAT field is present when safe",
            variable=self.ensure_reciprocal_var,
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(10, 0))
        ttk.Checkbutton(
            target_box,
            text="If an existing person's identity changes, update other references that uniquely point to that row",
            variable=self.update_all_refs_var,
        ).grid(row=4, column=2, columnspan=2, sticky="w", pady=(10, 0), padx=(18, 0))
        self.replace_conflict_check = ttk.Checkbutton(
            target_box,
            text="Replace a conflicting existing VET on the selected RAT's row",
            variable=self.replace_conflicting_vet_var,
        )
        self.replace_conflict_check.grid(row=5, column=0, columnspan=4, sticky="w", pady=(4, 0))

        for col in range(4):
            target_box.columnconfigure(col, weight=1)

        self.resolve_message_var = tk.StringVar()
        ttk.Label(self.resolve_tab, textvariable=self.resolve_message_var, style="Danger.TLabel", wraplength=850).pack(anchor="w", pady=(8, 0))

        buttons = ttk.Frame(self.resolve_tab)
        buttons.pack(fill="x", pady=(10, 0))
        ttk.Button(buttons, text="Skip this reference", command=self.skip_current).pack(side="left")
        ttk.Button(buttons, text="Preview local tree →", command=self.preview_current).pack(side="right")

    def _build_preview_tab(self) -> None:
        heading = ttk.Frame(self.preview_tab)
        heading.pack(fill="x")
        ttk.Label(heading, text="Localized relationship preview", style="Heading.TLabel").pack(side="left")
        ttk.Label(
            heading,
            text="Only the reviewed person, their VET(s), and their RAT(s) are shown.",
            style="Muted.TLabel",
        ).pack(side="left", padx=14)

        self.preview_warning_var = tk.StringVar()
        ttk.Label(
            self.preview_tab,
            textvariable=self.preview_warning_var,
            style="Danger.TLabel",
            wraplength=900,
        ).pack(anchor="w", pady=(6, 6))

        canvas_frame = ttk.Frame(self.preview_tab)
        canvas_frame.pack(fill="both", expand=True)
        self.preview_canvas = tk.Canvas(canvas_frame, background="white", highlightthickness=1, highlightbackground="#B7B7B7")
        xscroll = ttk.Scrollbar(canvas_frame, orient="horizontal", command=self.preview_canvas.xview)
        yscroll = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.preview_canvas.yview)
        self.preview_canvas.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)
        self.preview_canvas.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        canvas_frame.rowconfigure(0, weight=1)
        canvas_frame.columnconfigure(0, weight=1)

        question = ttk.Frame(self.preview_tab, padding=(0, 10, 0, 0))
        question.pack(fill="x")
        ttk.Label(question, text="Does this look correct?", style="Heading.TLabel").pack(side="left")
        ttk.Button(question, text="No — revise answers", command=self.return_to_resolve).pack(side="right")
        ttk.Button(question, text="Yes — save and continue", command=self.confirm_preview).pack(side="right", padx=(0, 8))

    def rescan(self, *, select_first: bool) -> None:
        try:
            self.model.load()
        except Exception as exc:
            messagebox.showerror("Could not reload workbook", str(exc), parent=self)
            return
        self.review_items = self.model.scan_review_items(self.session_skips)
        self._populate_queue()
        self.status_var.set(
            f"{len(self.review_items)} unresolved relationship reference(s) • {len(self.model.people)} person row(s)"
        )
        if select_first and self.review_items:
            first = self.queue.get_children()[0]
            self.queue.selection_set(first)
            self.queue.focus(first)
            self.queue.see(first)
            self._load_item(self.review_items[0])
        elif not self.review_items:
            self._clear_resolution()
            messagebox.showinfo(
                "Review complete",
                "No unresolved VET/RAT references remain in this worksheet (excluding anything skipped for this session).",
                parent=self,
            )

    def _populate_queue(self) -> None:
        for child in self.queue.get_children():
            self.queue.delete(child)
        for index, item in enumerate(self.review_items):
            self.queue.insert(
                "",
                "end",
                iid=str(index),
                values=(item.source_name, item.column_label, item.reference.name or item.raw),
            )

    def _queue_selected(self, _event=None) -> None:
        selection = self.queue.selection()
        if not selection:
            return
        index = int(selection[0])
        if 0 <= index < len(self.review_items):
            self._load_item(self.review_items[index])

    def _load_item(self, item: ReviewItem) -> None:
        self.current_item = item
        self.notebook.select(self.resolve_tab)
        exact_note = (
            f"{len(item.exact_candidate_rows)} exact-name candidate rows were found, so the reference is not unique."
            if item.exact_candidate_rows
            else "No unique exact person row was found."
        )
        self.item_title_var.set(
            f"{item.source_name} • {item.column_label} • Excel row {item.source_row}"
        )
        self.item_raw_var.set(f"Current cell: {item.raw}    —    {exact_note}")
        self.ref_name_var.set(item.reference.name)
        self.ref_year_var.set(str(item.reference.year or ""))
        self.ref_instrument_var.set(item.reference.instrument)
        self.action_var.set("existing")
        self.target_given_var.set("")
        self.target_family_var.set("")
        self.target_year_var.set("")
        self.target_instrument_var.set("")
        self.ensure_reciprocal_var.set(True)
        self.update_all_refs_var.set(True)
        self.replace_conflicting_vet_var.set(False)
        self.resolve_message_var.set("")
        self.refresh_candidates_from_reference(auto_select=True)

    def _clear_resolution(self) -> None:
        self.current_item = None
        self.item_title_var.set("Nothing currently selected")
        self.item_raw_var.set("")
        self.resolve_message_var.set("")
        for tree in (self.candidates,):
            for child in tree.get_children():
                tree.delete(child)

    def refresh_candidates_from_reference(self, auto_select: bool = False) -> None:
        if not self.current_item:
            return
        ref = RelationReference(
            raw=self.current_item.raw,
            name=normalize_spaces(self.ref_name_var.get()),
            year=parse_year(self.ref_year_var.get()),
            instrument=normalize_spaces(self.ref_instrument_var.get()),
        )
        self.current_candidates = self.model.fuzzy_candidates(ref, limit=12)
        for child in self.candidates.get_children():
            self.candidates.delete(child)
        for index, match in enumerate(self.current_candidates):
            person = match.person
            self.candidates.insert(
                "",
                "end",
                iid=str(index),
                values=(
                    f"{match.score:.0f}",
                    person.name,
                    person.year or "",
                    person.instrument,
                    person.row,
                    match.explanation,
                ),
            )
        if auto_select and self.current_candidates:
            # Auto-highlight only when there is a very strong candidate; this does
            # not accept it or save anything.
            best = self.current_candidates[0]
            if best.score >= 90:
                self.candidates.selection_set("0")
                self.candidates.focus("0")
                self._candidate_selected()
            else:
                self._prefill_new_from_reference()
                self.action_var.set("new")
        elif not self.current_candidates:
            self._prefill_new_from_reference()
            self.action_var.set("new")

    def _candidate_selected(self, _event=None) -> None:
        selection = self.candidates.selection()
        if not selection:
            return
        index = int(selection[0])
        if not (0 <= index < len(self.current_candidates)):
            return
        person = self.current_candidates[index].person
        self.action_var.set("existing")
        self.target_given_var.set(person.given)
        self.target_family_var.set(person.family)
        self.target_year_var.set(str(person.year or ""))
        self.target_instrument_var.set(person.instrument)
        self.resolve_message_var.set("")
        self._update_conflict_state(person.row)

    def _action_changed(self) -> None:
        if self.action_var.get() == "new":
            self._prefill_new_from_reference()
            self.replace_conflicting_vet_var.set(False)
        else:
            selection = self.candidates.selection()
            if selection:
                self._candidate_selected()

    def _prefill_new_from_reference(self) -> None:
        name = normalize_spaces(self.ref_name_var.get())
        given, family = split_name_guess(name)
        self.target_given_var.set(given)
        self.target_family_var.set(family)
        self.target_year_var.set(self.ref_year_var.get())
        self.target_instrument_var.set(self.ref_instrument_var.get())

    def _update_conflict_state(self, target_row: int | None) -> None:
        if not self.current_item or self.current_item.role != "RAT" or not target_row:
            self.replace_conflict_check.state(["disabled"])
            return
        target = self.model.people_by_row.get(target_row)
        source = self.model.people_by_row.get(self.current_item.source_row)
        conflict = False
        if target and source and target.vet_raw:
            ref = parse_relation(target.vet_raw)
            resolved = self.model.resolve_unique(ref) if ref else None
            conflict = bool(resolved and resolved.row != source.row)
        if conflict:
            self.replace_conflict_check.state(["!disabled"])
        else:
            self.replace_conflict_check.state(["disabled"])
            self.replace_conflicting_vet_var.set(False)

    def _build_draft_from_form(self) -> ResolutionDraft | None:
        if not self.current_item:
            return None
        action = self.action_var.get()
        target_row = None
        if action == "existing":
            selection = self.candidates.selection()
            if selection:
                index = int(selection[0])
                if 0 <= index < len(self.current_candidates):
                    target_row = self.current_candidates[index].person.row
        draft = ResolutionDraft(
            action=action,
            target_row=target_row,
            given=normalize_spaces(self.target_given_var.get()),
            family=normalize_spaces(self.target_family_var.get()),
            year=parse_year(self.target_year_var.get()),
            instrument=normalize_spaces(self.target_instrument_var.get()),
            ensure_reciprocal=self.ensure_reciprocal_var.get(),
            replace_conflicting_vet=self.replace_conflicting_vet_var.get(),
            update_all_references=self.update_all_refs_var.get(),
        )
        return draft

    def preview_current(self) -> None:
        if not self.current_item:
            return
        draft = self._build_draft_from_form()
        if draft is None:
            return
        errors = self.model.validate_draft(self.current_item, draft)
        if errors:
            self.resolve_message_var.set(" • ".join(errors))
            return
        self.resolve_message_var.set("")
        self.preview_draft = draft
        preview = self.model.build_preview(self.current_item, draft)
        self._draw_local_tree(preview)
        self.preview_warning_var.set("  ".join(preview.warnings))
        self.notebook.select(self.preview_tab)

    def _draw_local_tree(self, preview: LocalPreview) -> None:
        canvas = self.preview_canvas
        canvas.delete("all")

        all_people = [*preview.vets, preview.center, *preview.rats]
        known_years = sorted({person.year for person in all_people if person.year is not None})
        row_labels: list[int | None] = list(known_years)
        if any(person.year is None for person in all_people):
            row_labels.append(None)
        if not row_labels:
            row_labels = [None]

        # If the year range is modest, show every actual RAT-year row so the
        # local preview resembles the main tree. For very large gaps, keep only
        # occupied years so the confirmation screen remains manageable.
        if known_years and known_years[-1] - known_years[0] <= 20:
            row_labels = list(range(known_years[0], known_years[-1] + 1))
            if any(person.year is None for person in all_people):
                row_labels.append(None)

        rat_count = max(1, len(preview.rats))
        vet_count = max(1, len(preview.vets))
        canvas_width = max(880, TREE_MARGIN_X * 2 + max(rat_count, vet_count) * (CARD_WIDTH + CARD_GAP))
        canvas_height = TREE_MARGIN_Y * 2 + len(row_labels) * YEAR_BAND_HEIGHT
        year_to_y = {
            year: TREE_MARGIN_Y + index * YEAR_BAND_HEIGHT + (YEAR_BAND_HEIGHT - CARD_HEIGHT) // 2
            for index, year in enumerate(row_labels)
        }

        navy = "#003057"
        gold = "#B3A369"
        white = "#FFFFFF"
        for index, year in enumerate(row_labels):
            top = TREE_MARGIN_Y + index * YEAR_BAND_HEIGHT
            pattern = index % 4
            fill = navy if pattern == 1 else gold if pattern == 3 else white
            canvas.create_rectangle(0, top, canvas_width, top + YEAR_BAND_HEIGHT, fill=fill, outline="")
            label_color = "white" if fill in {navy, gold} else "#222222"
            canvas.create_text(50, top + YEAR_BAND_HEIGHT / 2, text=str(year) if year is not None else "?", fill=label_color, font=("Segoe UI", 12, "bold"))

        center_y = year_to_y.get(preview.center.year, year_to_y[row_labels[-1]])

        # Position RATs left-to-right by RAT year then name. Center card is the
        # arithmetic mean of the RAT centers, matching the legacy placement idea.
        placements: dict[str, tuple[float, float, PreviewPerson]] = {}
        rat_centers: list[float] = []
        if preview.rats:
            total_rat_width = len(preview.rats) * CARD_WIDTH + (len(preview.rats) - 1) * CARD_GAP
            start_x = max(TREE_MARGIN_X + 70, (canvas_width - total_rat_width) / 2)
            for index, person in enumerate(preview.rats):
                x = start_x + index * (CARD_WIDTH + CARD_GAP)
                y = year_to_y.get(person.year, year_to_y[row_labels[-1]])
                placements[person.key] = (x, y, person)
                rat_centers.append(x + CARD_WIDTH / 2)
            center_x = sum(rat_centers) / len(rat_centers) - CARD_WIDTH / 2
        else:
            center_x = canvas_width / 2 - CARD_WIDTH / 2
        placements[preview.center.key] = (center_x, center_y, preview.center)

        if preview.vets:
            total_vet_width = len(preview.vets) * CARD_WIDTH + (len(preview.vets) - 1) * CARD_GAP
            vet_start_x = center_x + CARD_WIDTH / 2 - total_vet_width / 2
            for index, person in enumerate(preview.vets):
                x = vet_start_x + index * (CARD_WIDTH + CARD_GAP)
                y = year_to_y.get(person.year, year_to_y[row_labels[-1]])
                placements[person.key] = (x, y, person)

        # Draw connectors before cards.
        center_top = (center_x + CARD_WIDTH / 2, center_y)
        center_bottom = (center_x + CARD_WIDTH / 2, center_y + CARD_HEIGHT)

        # VET(s) -> center. Multiple VETs are only shown when there is a conflict.
        vet_bottoms = []
        for person in preview.vets:
            x, y, _ = placements[person.key]
            vet_bottoms.append((x + CARD_WIDTH / 2, y + CARD_HEIGHT, person))
        if len(vet_bottoms) == 1:
            vx, vy, person = vet_bottoms[0]
            self._draw_elbow_line(canvas, (vx, vy), center_top, dashed=person.proposed and len(preview.vets) > 1)
        elif len(vet_bottoms) > 1:
            bus_y = min(center_top[1] - 18, max(vy for _, vy, _ in vet_bottoms) + 18)
            xs = [vx for vx, _, _ in vet_bottoms]
            canvas.create_line(min(xs), bus_y, max(xs), bus_y, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH)
            for vx, vy, person in vet_bottoms:
                canvas.create_line(vx, vy, vx, bus_y, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH, dash=(6, 4) if person.proposed else None)
            canvas.create_line(center_top[0], bus_y, center_top[0], center_top[1], fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH)

        # Center -> RAT(s): original-style stem, sibling bus, then child stems.
        rat_tops = []
        for person in preview.rats:
            x, y, _ = placements[person.key]
            rat_tops.append((x + CARD_WIDTH / 2, y, person))
        if len(rat_tops) == 1:
            rx, ry, person = rat_tops[0]
            self._draw_elbow_line(canvas, center_bottom, (rx, ry), dashed=person.proposed and person.key.startswith("source-"))
        elif len(rat_tops) > 1:
            earliest_top = min(ry for _, ry, _ in rat_tops)
            bus_y = min(center_bottom[1] + 24, earliest_top - 20)
            if bus_y <= center_bottom[1]:
                bus_y = center_bottom[1] + 18
            xs = [rx for rx, _, _ in rat_tops]
            canvas.create_line(center_bottom[0], center_bottom[1], center_bottom[0], bus_y, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH)
            canvas.create_line(min(xs), bus_y, max(xs), bus_y, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH)
            for rx, ry, person in rat_tops:
                canvas.create_line(rx, bus_y, rx, ry, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH, dash=(6, 4) if person.unresolved else None)

        for x, y, person in placements.values():
            self._draw_card(canvas, x, y, person, is_center=person.key == preview.center.key)

        canvas.configure(scrollregion=(0, 0, canvas_width, canvas_height))
        canvas.xview_moveto(0)
        canvas.yview_moveto(0)

    @staticmethod
    def _draw_elbow_line(canvas: tk.Canvas, start: tuple[float, float], end: tuple[float, float], dashed: bool = False) -> None:
        sx, sy = start
        ex, ey = end
        if abs(sx - ex) < 1:
            canvas.create_line(sx, sy, ex, ey, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH, dash=(6, 4) if dashed else None)
            return
        mid_y = sy + (ey - sy) / 2
        canvas.create_line(sx, sy, sx, mid_y, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH, dash=(6, 4) if dashed else None)
        canvas.create_line(sx, mid_y, ex, mid_y, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH, dash=(6, 4) if dashed else None)
        canvas.create_line(ex, mid_y, ex, ey, fill=CONNECTOR_COLOR, width=CONNECTOR_WIDTH, dash=(6, 4) if dashed else None)

    @staticmethod
    def _draw_card(canvas: tk.Canvas, x: float, y: float, person: PreviewPerson, *, is_center: bool) -> None:
        sections = normalize_sections(person.instrument)
        segment_width = CARD_WIDTH / len(sections)
        for index, section in enumerate(sections):
            x0 = x + index * segment_width
            x1 = x + (index + 1) * segment_width
            canvas.create_rectangle(
                x0,
                y,
                x1,
                y + CARD_HEIGHT,
                fill=SECTION_COLORS.get(section, SECTION_COLORS["unknown"]),
                outline="",
            )
        border = "#111111" if not person.unresolved else "#8b3a3a"
        width = 3 if is_center else 2
        canvas.create_rectangle(x, y, x + CARD_WIDTH, y + CARD_HEIGHT, outline=border, width=width)
        # Wrap name over two lines if needed.
        words = person.name.split()
        if len(words) >= 2:
            line1 = " ".join(words[:-1])
            line2 = words[-1]
        else:
            line1 = person.name
            line2 = ""
        text = line1 if not line2 else f"{line1}\n{line2}"
        canvas.create_text(
            x + CARD_WIDTH / 2,
            y + CARD_HEIGHT / 2 - (7 if person.unresolved else 0),
            text=text,
            justify="center",
            font=("Segoe UI", 10, "bold" if is_center else "normal"),
            fill="#111111",
            width=CARD_WIDTH - 12,
        )
        if person.unresolved:
            canvas.create_text(
                x + CARD_WIDTH / 2,
                y + CARD_HEIGHT - 9,
                text="unresolved reference",
                font=("Segoe UI", 7, "italic"),
                fill="#8b3a3a",
            )

    def return_to_resolve(self) -> None:
        self.notebook.select(self.resolve_tab)

    def confirm_preview(self) -> None:
        if not self.current_item or not self.preview_draft:
            return
        try:
            _, messages = self.model.commit_resolution(self.current_item, self.preview_draft)
        except Exception as exc:
            messagebox.showerror("Could not save changes", str(exc), parent=self)
            return
        messagebox.showinfo(
            "Saved",
            "The confirmed resolution was saved.\n\n" + "\n".join(messages),
            parent=self,
        )
        self.preview_draft = None
        # Rescan from disk; other references may now resolve automatically.
        self.rescan(select_first=True)

    def skip_current(self) -> None:
        if not self.current_item:
            return
        self.session_skips.add(self.current_item.key)
        self.rescan(select_first=True)

    def reset_skips(self) -> None:
        self.session_skips.clear()
        self.rescan(select_first=True)


# ---------------------------------------------------------------------------
# Helpers / CLI
# ---------------------------------------------------------------------------


def split_name_guess(name: str) -> tuple[str, str]:
    """Conservative prefill only; the GUI always lets the user correct it."""
    parts = normalize_spaces(name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    # Keep common multi-word surname forms together when obvious.
    lower = [part.casefold() for part in parts]
    surname_prefixes = {"el", "de", "del", "van", "von", "la", "le", "st.", "saint"}
    if len(parts) >= 3 and lower[-2] in surname_prefixes:
        return " ".join(parts[:-2]), " ".join(parts[-2:])
    return " ".join(parts[:-1]), parts[-1]


def choose_workbook(initial: Path | None = None) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="Select YJMB Trees workbook",
        initialdir=str((initial or SCRIPT_DIR).parent if initial else SCRIPT_DIR),
        filetypes=[("Excel workbooks", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()
    return Path(path).resolve() if path else None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Visual GUI for resolving missing/ambiguous YJMB VET/RAT references.")
    parser.add_argument("--workbook", type=Path, default=None, help="Path to YJMB Trees.xlsx")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Master worksheet (default: {DEFAULT_SHEET!r})")
    parser.add_argument("--scan-only", action="store_true", help="Print unresolved references and exit without opening the GUI")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve() if args.workbook else DEFAULT_WORKBOOK
    if not workbook_path.exists():
        if args.scan_only:
            print(f"Workbook not found: {workbook_path}", file=sys.stderr)
            return 2
        selected = choose_workbook(workbook_path)
        if not selected:
            return 0
        workbook_path = selected

    try:
        model = WorkbookModel(workbook_path, args.sheet)
    except Exception as exc:
        if args.scan_only:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 2
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Could not open workbook", str(exc), parent=root)
        root.destroy()
        return 2

    items = model.scan_review_items()
    if args.scan_only:
        print(f"Workbook: {model.path}")
        print(f"Sheet: {model.sheet_name}")
        print(f"People rows: {len(model.people)}")
        print(f"Unresolved/non-unique references: {len(items)}")
        for item in items:
            candidates = model.fuzzy_candidates(item.reference, limit=3)
            suggestion = "; ".join(
                f"{match.person.name} ({match.person.year or '?'}) score={match.score:.0f}"
                for match in candidates
            )
            print(
                f"- row {item.source_row} {item.source_name} {item.column_label}: "
                f"{item.raw!r} | suggestions: {suggestion or 'none'}"
            )
        return 0

    app = RelationshipReviewApp(model)
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
