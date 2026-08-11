#!/usr/bin/env python3
"""Shared workbook schema helpers for the YJMB family-tree project.

This module contains only column names and mechanical worksheet helpers. It does
not contain member data or secrets and is safe to commit.
"""
from __future__ import annotations

import copy
import re
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl.utils import get_column_letter, range_boundaries


def norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", norm(value).casefold())


# These columns are append-only compatibility fields. Existing workbook columns
# are never renamed or overwritten by the migration helper.
V17_OPTIONAL_COLUMNS: tuple[str, ...] = (
    "Tree Display Name Preference",
    "Section Nicknames",
    "Specific Instruments",
    "Favorite Tech Band Memory",
    "Marching Band Leadership Role(s)",
    "Served in Informal Leadership Position",
    "Informal Leadership Position(s)",
    "Marching Band Leadership History",
    "Served in Band Club Leadership Position",
    "Band Club Leadership Position(s)",
    "Band Club Leadership History",
    "Leadership Position Classification",
    "Has Nickname",
    "Changed Last Name Since Band",
    "Has Been in Multiple Sections",
    "Currently a RAT",
    "RAT/VET Pair System Applied",
)


def find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 10) + 1):
        values = {header_key(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if "givenpreferredname" in values and "familymaidenname" in values:
            return row
    raise RuntimeError("Could not find the People on Tree header row.")


def header_map(ws, header_row: int | None = None) -> dict[str, int]:
    header_row = header_row or find_header_row(ws)
    return {
        header_key(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if norm(ws.cell(header_row, col).value)
    }


def missing_optional_columns(ws, header_row: int | None = None) -> list[str]:
    header_row = header_row or find_header_row(ws)
    existing = set(header_map(ws, header_row))
    return [name for name in V17_OPTIONAL_COLUMNS if header_key(name) not in existing]


def _copy_column_style(ws, source_col: int, target_col: int, header_row: int) -> None:
    # Copy presentation only. No values/formulas are copied into the new field.
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, source_col)
        dst = ws.cell(row, target_col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
        dst.number_format = src.number_format
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    if source_letter in ws.column_dimensions:
        src_dim = ws.column_dimensions[source_letter]
        dst_dim = ws.column_dimensions[target_letter]
        if src_dim.width is not None:
            dst_dim.width = src_dim.width


def ensure_optional_columns(
    ws,
    *,
    header_row: int | None = None,
    columns: Iterable[str] = V17_OPTIONAL_COLUMNS,
) -> tuple[dict[str, int], list[str]]:
    """Append any missing optional columns and return (header map, added names).

    Existing data is never changed. Excel tables/autofilters covering the master
    data region are expanded horizontally so downloaded workbooks keep the new
    fields inside the same visible table.
    """
    header_row = header_row or find_header_row(ws)
    existing = header_map(ws, header_row)
    added: list[str] = []
    old_max_col = ws.max_column
    style_source_col = max(1, old_max_col)

    for name in columns:
        k = header_key(name)
        if k in existing:
            continue
        new_col = ws.max_column + 1
        _copy_column_style(ws, style_source_col, new_col, header_row)
        ws.cell(header_row, new_col).value = name
        existing[k] = new_col
        added.append(name)
        style_source_col = new_col

    if not added:
        return existing, added

    new_max_col = ws.max_column
    for table in ws.tables.values():
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        except ValueError:
            continue
        if min_row <= header_row <= max_row and max_col == old_max_col:
            table.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(new_max_col)}{max_row}"
            )
    if ws.auto_filter and ws.auto_filter.ref:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ws.auto_filter.ref)
            if max_col == old_max_col:
                ws.auto_filter.ref = (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(new_max_col)}{max_row}"
                )
        except ValueError:
            pass

    return existing, added
