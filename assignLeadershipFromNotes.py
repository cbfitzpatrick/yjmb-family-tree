#!/usr/bin/env python3
"""Scan note/comment columns for high-confidence leadership mentions.

Default is scan-only. With --apply, recognized roles are unioned into the
existing structured leadership fields without deleting anything:
  * Marching Band Leadership Role(s)
  * Served in Informal Leadership Position
  * Informal Leadership Position(s)

Exact/high-confidence phrases such as "drum major", "section leader",
"RAT parent", and "hype man" are assignable. Generic words such as "leader",
"leadership", or "captain" are reported for review rather than guessed.
"""
from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from yjmb_taxonomy import canonical_formal_roles, extract_roles_from_notes, informal_roles_from_text, key, norm

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "YJMB Trees.xlsx"
DEFAULT_SHEET = "People on Tree"
FORMAL_HEADER = "Marching Band Leadership Role(s)"
INFORMAL_FLAG_HEADER = "Served in Informal Leadership Position"
INFORMAL_HEADER = "Informal Leadership Position(s)"
AMBIGUOUS_RE = re.compile(r"\b(?:captain|leader|leadership|mentor|coordinator)\b", re.I)


def hk(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", norm(value).casefold())


def discover(ws):
    for row in range(1, min(10, ws.max_row) + 1):
        headers = {c: hk(ws.cell(row, c).value) for c in range(1, ws.max_column + 1)}
        if "givenpreferredname" in headers.values() and "familymaidenname" in headers.values():
            return row, headers
    raise RuntimeError("Could not find People on Tree header row.")


def find_col(headers: dict[int, str], *names: str) -> int | None:
    wanted = {hk(name) for name in names}
    for col, value in headers.items():
        if value in wanted:
            return col
    return None


def note_columns(headers: dict[int, str], include_memory: bool) -> list[int]:
    result = []
    for col, h in headers.items():
        if any(token in h for token in ("note", "notes", "comment", "comments", "remark", "remarks", "additionalinfo", "additionalinformation")):
            result.append(col)
        elif include_memory and h == "favoritetechbandmemory":
            result.append(col)
    return sorted(set(result))


def backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = path.parent / "backups" / "data_cleanup" / stamp
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / path.name
    shutil.copy2(path, target)
    return target


def union_roles(existing: list[str], additions: list[str]) -> list[str]:
    out = list(existing)
    seen = {key(v) for v in out}
    for item in additions:
        if key(item) not in seen:
            out.append(item)
            seen.add(key(item))
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Assign high-confidence leadership roles from note fields.")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--include-memory", action="store_true", help="Also scan Favorite Tech Band Memory.")
    ap.add_argument("--apply", action="store_true", help="Union detected roles into structured leadership fields.")
    args = ap.parse_args()

    path = args.workbook.expanduser().resolve()
    wb = load_workbook(path)
    ws = wb[args.sheet]
    header_row, headers = discover(ws)
    notes = note_columns(headers, args.include_memory)
    if not notes:
        print("No note/comment columns were found. Nothing was changed.")
        wb.close()
        return 0

    formal_col = find_col(headers, FORMAL_HEADER, "Formal Leadership Position(s)", "Formal Leadership Role(s)")
    informal_flag_col = find_col(headers, INFORMAL_FLAG_HEADER, "Informal Leadership")
    informal_col = find_col(headers, INFORMAL_HEADER, "Informal Leadership Role(s)")

    proposals: list[dict[str, object]] = []
    ambiguous: list[str] = []

    for row in range(header_row + 1, ws.max_row + 1):
        combined = "\n".join(norm(ws.cell(row, col).value) for col in notes if norm(ws.cell(row, col).value))
        if not combined:
            continue
        formal_add, informal_add = extract_roles_from_notes(combined)
        if not formal_add and not informal_add:
            if AMBIGUOUS_RE.search(combined):
                ambiguous.append(f"row {row}: generic leadership wording found; review manually")
            continue
        existing_formal = canonical_formal_roles(ws.cell(row, formal_col).value) if formal_col else []
        existing_informal = informal_roles_from_text(ws.cell(row, informal_col).value) if informal_col else []
        new_formal = union_roles(existing_formal, formal_add)
        new_informal = union_roles(existing_informal, informal_add)
        proposals.append({
            "row": row,
            "formal_add": formal_add,
            "informal_add": informal_add,
            "formal": new_formal,
            "informal": new_informal,
        })

    print(f"Workbook: {path}")
    print("Note columns scanned: " + ", ".join(str(ws.cell(header_row,c).value) for c in notes))
    print(f"Rows with high-confidence assignments: {len(proposals)}")
    for item in proposals[:100]:
        additions = [*(f"Formal={x}" for x in item['formal_add']), *(f"Informal={x}" for x in item['informal_add'])]
        print(f"  row {item['row']}: " + ", ".join(additions))
    print(f"Rows needing manual leadership review: {len(ambiguous)}")
    for item in ambiguous[:100]:
        print(f"  ! {item}")

    if not args.apply:
        print("\nScan only. Re-run with --apply to add high-confidence assignments.")
        wb.close()
        return 0

    if not proposals:
        print("\nNothing to apply.")
        wb.close()
        return 0

    # Create structured columns only if a project variant is missing them.
    if formal_col is None:
        formal_col = ws.max_column + 1
        ws.cell(header_row, formal_col).value = FORMAL_HEADER
    if informal_flag_col is None:
        informal_flag_col = ws.max_column + 1
        ws.cell(header_row, informal_flag_col).value = INFORMAL_FLAG_HEADER
    if informal_col is None:
        informal_col = ws.max_column + 1
        ws.cell(header_row, informal_col).value = INFORMAL_HEADER

    backup_path = backup(path)
    for item in proposals:
        row = int(item["row"])
        formal = list(item["formal"])
        informal = list(item["informal"])
        ws.cell(row, formal_col).value = ", ".join(formal) or None
        if informal:
            ws.cell(row, informal_flag_col).value = "Yes"
            ws.cell(row, informal_col).value = ", ".join(informal)
    wb.save(path)
    wb.close()
    print(f"\nApplied high-confidence note assignments. Backup: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
