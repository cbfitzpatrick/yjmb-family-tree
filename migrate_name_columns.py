"""Migrate YJMB workbooks to four canonical name columns.

Current schema, in this exact order at the beginning of each data table:
    1. Given/Preferred Name
    2. Nickname
    3. Family/Maiden Name
    4. Married Name

The migration supports both the original single ``Name`` column and the earlier
two-column migration. Existing Nickname values are moved into column 2 rather
than duplicated. Married Name remains blank unless the source explicitly marks
it or the user confirms that the person adopted a spouse's surname.
"""

from __future__ import annotations

import argparse
from copy import copy
from dataclasses import dataclass
from pathlib import Path
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import TableColumn

from name_tools import (
    COMPOUND_FAMILY_NAMES,
    InteractiveResolver,
    NameParts,
    SkipRecord,
    UserQuit,
    has_ambiguous_annotation,
    normalize_spaces,
    normalized_header,
    suspicious_case_or_spacing,
)
from tree_workbook import (
    FAMILY_HEADER,
    GIVEN_HEADER,
    MARRIED_HEADER,
    NICKNAME_HEADER,
    backup_file,
    find_header_row,
    header_map,
    save_workbook_atomic,
    timestamped_backup_directory,
)


SKIP_DIRECTORY_NAMES = {".git", ".venv", "venv", "backups", "node_modules", "__pycache__"}
CANONICAL_HEADERS = (GIVEN_HEADER, NICKNAME_HEADER, FAMILY_HEADER, MARRIED_HEADER)
NAME_HEADER_KEYS = {
    "name",
    "fullname",
    "givenpreferredname",
    "givenname",
    "preferredname",
    "firstname",
    "nickname",
    "nick",
    "preferrednickname",
    "familymaidenname",
    "familyname",
    "maidenname",
    "lastname",
    "surname",
    "marriedname",
    "marriedsurname",
    "spousesurname",
    "currentlastname",
}


@dataclass
class ColumnSnapshot:
    width: float | None
    hidden: bool
    best_fit: bool
    outline_level: int
    collapsed: bool
    styles: list[object]


def discover_workbooks(root: Path) -> list[Path]:
    results: list[Path] = []
    for path in root.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if any(part in SKIP_DIRECTORY_NAMES for part in path.relative_to(root).parts):
            continue
        results.append(path)
    return sorted(results, key=lambda item: str(item).casefold())


def _worksheet_headers(ws, header_row: int) -> dict[str, int]:
    return {
        normalized_header(ws.cell(header_row, column).value): column
        for column in range(1, ws.max_column + 1)
        if normalize_spaces(ws.cell(header_row, column).value)
    }


def _all_name_columns(ws, header_row: int) -> list[int]:
    columns = []
    for column in range(1, ws.max_column + 1):
        if normalized_header(ws.cell(header_row, column).value) in NAME_HEADER_KEYS:
            columns.append(column)
    return sorted(set(columns))


def _capture_column(ws, column: int, max_row: int) -> ColumnSnapshot:
    letter = get_column_letter(column)
    dimension = ws.column_dimensions[letter]
    return ColumnSnapshot(
        width=dimension.width,
        hidden=bool(dimension.hidden),
        best_fit=bool(dimension.bestFit),
        outline_level=int(dimension.outlineLevel or 0),
        collapsed=bool(dimension.collapsed),
        styles=[copy(ws.cell(row, column)._style) for row in range(1, max_row + 1)],
    )


def _apply_column_snapshot(ws, column: int, snapshot: ColumnSnapshot, max_row: int) -> None:
    letter = get_column_letter(column)
    dimension = ws.column_dimensions[letter]
    dimension.width = snapshot.width
    dimension.hidden = snapshot.hidden
    dimension.bestFit = snapshot.best_fit
    dimension.outlineLevel = snapshot.outline_level
    dimension.collapsed = snapshot.collapsed
    for row in range(1, max_row + 1):
        if row - 1 < len(snapshot.styles):
            ws.cell(row, column)._style = copy(snapshot.styles[row - 1])


def _transform_column(original_column: int, removed_columns: list[int], base_column: int) -> int | None:
    if original_column in removed_columns:
        return None
    after_delete = original_column - sum(1 for removed in removed_columns if removed < original_column)
    if after_delete >= base_column:
        after_delete += len(CANONICAL_HEADERS)
    return after_delete


def _transform_reference(ref: str, removed_columns: list[int], base_column: int) -> str:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    transformed = [
        result
        for column in range(min_col, max_col + 1)
        if (result := _transform_column(column, removed_columns, base_column)) is not None
    ]
    if any(min_col <= column <= max_col for column in removed_columns):
        transformed.extend(range(base_column, base_column + len(CANONICAL_HEADERS)))
    if not transformed:
        transformed = list(range(base_column, base_column + len(CANONICAL_HEADERS)))
    return (
        f"{get_column_letter(min(transformed))}{min_row}:"
        f"{get_column_letter(max(transformed))}{max_row}"
    )


def _sync_table_columns(ws, table) -> None:
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    table.tableColumns = []
    used_names: dict[str, int] = {}
    for index, column in enumerate(range(min_col, max_col + 1), start=1):
        name = normalize_spaces(ws.cell(min_row, column).value) or f"Column {index}"
        count = used_names.get(name.casefold(), 0) + 1
        used_names[name.casefold()] = count
        if count > 1:
            name = f"{name} {count}"
            ws.cell(min_row, column).value = name
        table.tableColumns.append(TableColumn(id=index, name=name))


def _known_compound_suffix(raw: str) -> bool:
    lowered = raw.casefold()
    return any(lowered.endswith(" " + family.casefold()) for family in COMPOUND_FAMILY_NAMES)


def _resolved_rows(
    ws,
    mapping: dict[str, int],
    resolver: InteractiveResolver,
    *,
    workbook_label: str,
    header_row: int,
) -> tuple[dict[int, NameParts], int]:
    old_name_col = mapping.get("old_name")
    given_col = mapping.get("given")
    nickname_col = mapping.get("nickname")
    family_col = mapping.get("family")
    married_col = mapping.get("married")

    raw_names: list[str] = []
    for row in range(header_row + 1, ws.max_row + 1):
        if old_name_col:
            candidate = normalize_spaces(ws.cell(row, old_name_col).value)
        else:
            given = normalize_spaces(ws.cell(row, given_col).value) if given_col else ""
            family = normalize_spaces(ws.cell(row, family_col).value) if family_col else ""
            married = normalize_spaces(ws.cell(row, married_col).value) if married_col else ""
            candidate = normalize_spaces(f"{given} {family or married}")
        if candidate:
            raw_names.append(candidate)

    resolved: dict[int, NameParts] = {}
    skipped = 0
    for row in range(header_row + 1, ws.max_row + 1):
        existing_nickname = normalize_spaces(ws.cell(row, nickname_col).value) if nickname_col else ""
        context = f"{workbook_label} / {ws.title} / row {row}"
        try:
            if old_name_col:
                raw = normalize_spaces(ws.cell(row, old_name_col).value)
                if not raw:
                    continue
                force = (
                    (len(raw.split()) >= 3 and not _known_compound_suffix(raw))
                    or has_ambiguous_annotation(raw)
                )
                parts = resolver.resolve_name(
                    raw,
                    context=context,
                    existing_names=raw_names,
                    provided_nickname=existing_nickname,
                    force_confirm=force,
                )
            else:
                given = normalize_spaces(ws.cell(row, given_col).value) if given_col else ""
                family = normalize_spaces(ws.cell(row, family_col).value) if family_col else ""
                married = normalize_spaces(ws.cell(row, married_col).value) if married_col else ""
                if not any((given, existing_nickname, family, married)):
                    continue

                family_needs_split = bool(family) and has_ambiguous_annotation(family)
                suspicious = any(
                    suspicious_case_or_spacing(value)
                    for value in (given, existing_nickname, family, married)
                    if value
                )
                if family_needs_split:
                    raw = normalize_spaces(f"{given} {family}")
                    parts = resolver.resolve_name(
                        raw,
                        context=context,
                        existing_names=raw_names,
                        provided_given=given,
                        provided_nickname=existing_nickname,
                        force_confirm=True,
                    )
                elif suspicious or married:
                    raw = normalize_spaces(
                        f"{given} ({family}) {married}" if married and family else f"{given} {family or married}"
                    )
                    parts = resolver.resolve_name(
                        raw,
                        context=context,
                        existing_names=raw_names,
                        provided_given=given,
                        provided_nickname=existing_nickname,
                        provided_family=family,
                        provided_married=married,
                        force_confirm=suspicious or bool(married),
                    )
                else:
                    parts = NameParts(given, existing_nickname, family, "")
        except SkipRecord:
            skipped += 1
            if old_name_col:
                raw = normalize_spaces(ws.cell(row, old_name_col).value)
                parts = NameParts(raw, existing_nickname, "", "")
            else:
                parts = NameParts(given, existing_nickname, family, married)
        resolved[row] = parts
    return resolved, skipped


def migrate_sheet(ws, resolver: InteractiveResolver, *, workbook_label: str) -> tuple[bool, int, int]:
    """Return ``(sheet_changed, rows_migrated, rows_skipped)``."""
    header_row = find_header_row(ws)
    if header_row is None:
        return False, 0, 0
    mapping = header_map(ws, header_row)
    if "old_name" not in mapping and "given" not in mapping:
        return False, 0, 0

    base_column = mapping.get("old_name") or mapping.get("given")
    expected = {
        "given": base_column,
        "nickname": base_column + 1,
        "family": base_column + 2,
        "married": base_column + 3,
    }
    headers = _worksheet_headers(ws, header_row)
    if all(mapping.get(field) == column for field, column in expected.items()) and all(
        normalized_header(ws.cell(header_row, base_column + offset).value)
        == normalized_header(header)
        for offset, header in enumerate(CANONICAL_HEADERS)
    ):
        return False, 0, 0

    resolved, skipped = _resolved_rows(
        ws,
        mapping,
        resolver,
        workbook_label=workbook_label,
        header_row=header_row,
    )

    original_max_row = ws.max_row
    name_columns = _all_name_columns(ws, header_row)
    if not name_columns:
        return False, 0, skipped
    base_column = min(base_column, min(name_columns))

    source_columns = {
        "given": mapping.get("given") or mapping.get("old_name") or name_columns[0],
        "nickname": mapping.get("nickname") or mapping.get("given") or mapping.get("old_name") or name_columns[0],
        "family": mapping.get("family") or mapping.get("given") or mapping.get("old_name") or name_columns[0],
        "married": mapping.get("married") or mapping.get("family") or mapping.get("given") or mapping.get("old_name") or name_columns[0],
    }
    snapshots = {
        field: _capture_column(ws, column, original_max_row)
        for field, column in source_columns.items()
    }

    original_table_refs = {table.name: table.ref for table in ws.tables.values()}
    original_auto_filter = ws.auto_filter.ref

    for column in sorted(name_columns, reverse=True):
        ws.delete_cols(column, 1)
    ws.insert_cols(base_column, len(CANONICAL_HEADERS))

    for offset, field in enumerate(("given", "nickname", "family", "married")):
        destination = base_column + offset
        _apply_column_snapshot(ws, destination, snapshots[field], original_max_row)
        ws.cell(header_row, destination).value = CANONICAL_HEADERS[offset]

    for row, parts in resolved.items():
        ws.cell(row, base_column).value = parts.given
        ws.cell(row, base_column + 1).value = parts.nickname
        ws.cell(row, base_column + 2).value = parts.family
        ws.cell(row, base_column + 3).value = parts.married

    minimum_widths = (23, 18, 23, 23)
    for offset, minimum in enumerate(minimum_widths):
        letter = get_column_letter(base_column + offset)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 10, minimum)

    for table_name, ref in original_table_refs.items():
        if table_name in ws.tables:
            table = ws.tables[table_name]
            table.ref = _transform_reference(ref, name_columns, base_column)
            _sync_table_columns(ws, table)
    if original_auto_filter:
        ws.auto_filter.ref = _transform_reference(original_auto_filter, name_columns, base_column)

    return True, len(resolved), skipped


def migrate_workbook(
    path: Path,
    resolver: InteractiveResolver,
    *,
    backup_dir: Path,
    dry_run: bool,
) -> tuple[int, int, int]:
    workbook = load_workbook(path)
    changed_sheets = 0
    migrated_rows = 0
    skipped_rows = 0
    for ws in workbook.worksheets:
        changed, count, skipped = migrate_sheet(ws, resolver, workbook_label=path.name)
        if changed:
            changed_sheets += 1
            migrated_rows += count
            skipped_rows += skipped
    if changed_sheets and not dry_run:
        backup_file(path, backup_dir / path.parent.name)
        save_workbook_atomic(workbook, path)
    workbook.close()
    return changed_sheets, migrated_rows, skipped_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--root",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="Project directory to scan recursively (default: script directory).",
    )
    parser.add_argument(
        "--cache",
        type=Path,
        help="Decision-cache JSON path. Default: <root>/.name_resolution_cache.json.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Analyze and prompt without writing files.")
    parser.add_argument("--non-interactive", action="store_true", help="Never prompt; fail on ambiguous names.")
    parser.add_argument(
        "--accept-auto",
        action="store_true",
        help="In non-interactive mode, accept automatic proposals even when suspicious.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = args.root.expanduser().resolve()
    if not root.exists():
        print(f"Project root does not exist: {root}", file=sys.stderr)
        return 2

    workbooks = discover_workbooks(root)
    if not workbooks:
        print(f"No .xlsx workbooks found beneath {root}")
        return 0

    cache_path = args.cache.expanduser().resolve() if args.cache else root / ".name_resolution_cache.json"
    resolver = InteractiveResolver(
        cache_path,
        interactive=not args.non_interactive,
        accept_auto=args.accept_auto,
    )
    backup_dir = timestamped_backup_directory(root, "before_four_name_column_migration")

    print(f"Found {len(workbooks)} workbook(s).")
    if args.dry_run:
        print("DRY RUN: no workbooks will be changed.")

    totals = {"workbooks": 0, "sheets": 0, "rows": 0, "skipped": 0}
    try:
        for path in workbooks:
            sheets, rows, skipped = migrate_workbook(
                path,
                resolver,
                backup_dir=backup_dir,
                dry_run=args.dry_run,
            )
            if sheets:
                totals["workbooks"] += 1
                totals["sheets"] += sheets
                totals["rows"] += rows
                totals["skipped"] += skipped
                action = "Would migrate" if args.dry_run else "Migrated"
                print(f"{action}: {path.relative_to(root)} ({sheets} sheet(s), {rows} row(s))")
            else:
                print(f"Skipped (already current or no supported name table): {path.relative_to(root)}")
    except UserQuit as exc:
        print(f"Stopped: {exc}")
        return 130
    finally:
        resolver.save()

    print("\nSummary")
    print(f"  Workbooks changed: {totals['workbooks']}")
    print(f"  Worksheets changed: {totals['sheets']}")
    print(f"  Name rows migrated: {totals['rows']}")
    print(f"  Rows deliberately skipped: {totals['skipped']}")
    if totals["workbooks"] and not args.dry_run:
        print(f"  Backups: {backup_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
