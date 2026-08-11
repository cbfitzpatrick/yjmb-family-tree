#!/usr/bin/env python3
"""Classify YJMB leadership data as Formal or Informal.

Current structured fields are treated as authoritative:
  * Marching Band Leadership Role(s) -> formal
  * Served in Informal Leadership Position -> informal flag
  * Informal Leadership Position(s) -> informal free text

Default is read-only. With --apply the script canonicalizes known formal-role
spellings and adds/updates a human-readable "Leadership Position Classification"
column such as:
    Formal: Drum Major; Formal: Section Leader; Informal: Hype Man
"""
from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from yjmb_taxonomy import canonical_formal_roles, extract_roles_from_notes, informal_roles_from_text, key, norm, truthy

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "YJMB Trees.xlsx"
DEFAULT_SHEET = "People on Tree"
CLASSIFICATION_HEADER = "Leadership Position Classification"


def hk(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", norm(value).casefold())


def discover(ws):
    for row in range(1, min(10, ws.max_row) + 1):
        headers = {c: hk(ws.cell(row, c).value) for c in range(1, ws.max_column + 1)}
        if "givenpreferredname" in headers.values() and "familymaidenname" in headers.values():
            return row, headers
    raise RuntimeError("Could not find People on Tree header row.")


def find_col(headers: dict[int, str], *names: str) -> int | None:
    wanted = {hk(name) for name in names}
    for col, value in headers.items():
        if value in wanted:
            return col
    return None


def backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    folder = path.parent / "backups" / "data_cleanup" / stamp
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / path.name
    shutil.copy2(path, target)
    return target


def main() -> int:
    ap = argparse.ArgumentParser(description="Classify formal/informal YJMB leadership positions.")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    path = args.workbook.expanduser().resolve()
    wb = load_workbook(path)
    ws = wb[args.sheet]
    header_row, headers = discover(ws)

    formal_col = find_col(headers, "Marching Band Leadership Role(s)", "Formal Leadership Position(s)", "Formal Leadership Role(s)")
    informal_flag_col = find_col(headers, "Served in Informal Leadership Position", "Informal Leadership")
    informal_col = find_col(headers, "Informal Leadership Position(s)", "Informal Leadership Role(s)")
    instrument_col = find_col(headers, "Instrument", "Instruments", "Section")
    class_col = find_col(headers, CLASSIFICATION_HEADER)

    if not any((formal_col, informal_flag_col, informal_col)):
        raise SystemExit("No leadership columns were found.")

    proposals: list[tuple[int, str, str, str]] = []
    formal_updates: list[tuple[int, str, str]] = []

    for row in range(header_row + 1, ws.max_row + 1):
        formal_raw = norm(ws.cell(row, formal_col).value) if formal_col else ""
        informal_flag = norm(ws.cell(row, informal_flag_col).value) if informal_flag_col else ""
        informal_raw = norm(ws.cell(row, informal_col).value) if informal_col else ""
        formal = canonical_formal_roles(formal_raw)
        legacy_instrument = norm(ws.cell(row, instrument_col).value) if instrument_col else ""
        legacy_formal, _ = extract_roles_from_notes(legacy_instrument)
        seen = {key(role) for role in formal}
        for role in legacy_formal:
            if key(role) not in seen:
                formal.append(role)
                seen.add(key(role))
        informal = informal_roles_from_text(informal_raw)
        if truthy(informal_flag) and not informal:
            informal = ["Informal leadership (unspecified)"]

        classified = [*(f"Formal: {role}" for role in formal), *(f"Informal: {role}" for role in informal)]
        classification = "; ".join(classified)
        if not classification and not any((formal_raw, informal_flag, informal_raw)):
            continue

        canonical_formal = ", ".join(formal)
        if canonical_formal and canonical_formal != formal_raw:
            formal_updates.append((row, formal_raw, canonical_formal))
        old_class = norm(ws.cell(row, class_col).value) if class_col else ""
        if classification != old_class:
            proposals.append((row, old_class, classification, formal_raw))

    print(f"Workbook: {path}")
    print(f"Classification rows to update: {len(proposals)}")
    print(f"Formal-role field updates: {len(formal_updates)}")
    for row, _, classification, _ in proposals[:100]:
        print(f"  row {row}: {classification or '(no classified leadership)'}")

    if not args.apply:
        print("\nScan only. Re-run with --apply to write classifications.")
        wb.close()
        return 0

    if class_col is None:
        class_col = ws.max_column + 1
        ws.cell(header_row, class_col).value = CLASSIFICATION_HEADER
    if formal_col is None and formal_updates:
        formal_col = ws.max_column + 1
        ws.cell(header_row, formal_col).value = "Marching Band Leadership Role(s)"
    backup_path = backup(path)
    for row, old, new, formal_raw in proposals:
        ws.cell(row, class_col).value = new or None
    if formal_col:
        for row, old, new in formal_updates:
            ws.cell(row, formal_col).value = new
    wb.save(path)
    wb.close()
    print(f"\nApplied classifications. Backup: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
