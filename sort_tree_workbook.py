"""Sort a four-name-column YJMB workbook without fixed column numbers."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
import shutil

from openpyxl import load_workbook

from name_tools import normalize_spaces
from tree_workbook import find_header_row, header_map, is_split_name_sheet, save_workbook_atomic


def year_key(value: object) -> int:
    text = normalize_spaces(value)
    return int(text[:4]) if len(text) >= 4 and text[:4].isdigit() else 9999


def sort_workbook(input_path: Path, output_path: Path, sheet_name: str | None = None) -> None:
    if input_path.resolve() != output_path.resolve():
        shutil.copy2(input_path, output_path)
    workbook = load_workbook(output_path)
    ws = workbook[sheet_name] if sheet_name else workbook.active
    if not is_split_name_sheet(ws):
        workbook.close()
        raise ValueError(
            f"{input_path.name} / {ws.title} does not use the four canonical name columns. Run migrate_name_columns.py first."
        )
    header_row = find_header_row(ws) or 1
    mapping = header_map(ws, header_row)

    rows = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        given = normalize_spaces(ws.cell(row_number, mapping["given"]).value)
        nickname = normalize_spaces(ws.cell(row_number, mapping["nickname"]).value)
        family = normalize_spaces(ws.cell(row_number, mapping["family"]).value)
        married = normalize_spaces(ws.cell(row_number, mapping["married"]).value)
        if not given and not family and not married:
            continue
        values = [ws.cell(row_number, col).value for col in range(1, ws.max_column + 1)]
        year = ws.cell(row_number, mapping["rat_year"]).value
        instrument = normalize_spaces(ws.cell(row_number, mapping["instrument"]).value)
        rows.append((year_key(year), 0 if "trumpet" in instrument.casefold() else 1, (family or married).casefold(), given.casefold(), nickname.casefold(), married.casefold(), instrument.casefold(), values))

    rows.sort(key=lambda item: item[:-1])
    for offset, item in enumerate(rows, start=header_row + 1):
        values = item[-1]
        for col, value in enumerate(values, start=1):
            ws.cell(offset, col).value = value

    # Clear any trailing rows that were previously occupied but are now beyond the data.
    first_empty = header_row + 1 + len(rows)
    for row_number in range(first_empty, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row_number, col).value = None

    save_workbook_atomic(workbook, output_path)
    workbook.close()
    print(f"Sorted workbook saved to: {output_path}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--sheet")
    args = parser.parse_args()
    sort_workbook(args.input.expanduser().resolve(), args.output.expanduser().resolve(), args.sheet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
