#!/usr/bin/env python3
"""Generate full-band YJMB family trees from ``YJMB Trees.xlsx``.

Default output:

* One complete full-band tree as PNG and SVG.
* One tree per disconnected family/root as PNG and SVG.
* Recolored per-person card PNGs made from ``blank_name_card.png``.
* A color legend and a validation/report text file.

The script expects the four-column name schema:

1. Given/Preferred Name
2. Nickname
3. Family/Maiden Name
4. Married Name

Lineage identity uses Given/Preferred Name + Family/Maiden Name. Card labels may use the Nickname when Tree Display Name Preference requests it.
The relationship columns are discovered by header, with VET and RAT columns
expected in J:P in the current workbook.
"""

from __future__ import annotations

import argparse
import base64
import difflib
import html
import json
import math
import os
import re
import shutil
import sys
import textwrap
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Iterator, Sequence

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from yjmb_taxonomy import leadership_icon_flags, truthy

# Large full-band charts can exceed Pillow's default decompression warning size.
Image.MAX_IMAGE_PIXELS = None

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_SOURCE_DIR = SCRIPT_DIR.parent / "trumpettree"
SOURCE_DIR = Path(os.environ.get("YJMB_SOURCE_DIR", DEFAULT_SOURCE_DIR)).expanduser().resolve()

MASTER_WORKBOOK = "YJMB Trees.xlsx"
DEFAULT_SHEET = "People on Tree"
DEFAULT_TEMPLATE = "blank_name_card.png"

FULL_TREE_HEADER_HEIGHT = 0
FAMILY_HEADER_HEIGHT = 300
YEAR_STRIP_HEIGHT = 100
LEFT_MARGIN = 200
RIGHT_MARGIN = 80
TOP_CARD_PADDING = 0
LEAF_GAP = 20
FAMILY_GAP = 200
# People with neither a VET nor a RAT are not visually separated as if they
# were full connected family trees. Connected families retain FAMILY_GAP.
ISOLATED_PERSON_GAP = 55
MIN_CANVAS_WIDTH = 900
CONNECTOR_COLOR = "#777777"
CONNECTOR_WIDTH = 9
# v17.4 restores the historical single gray connector stroke.
CONNECTOR_OUTLINE_COLOR = None
CONNECTOR_OUTLINE_EXTRA = 0
UNKNOWN_COLOR = "#D3D3D3"
DIVIDER_COLOR = "#555555"
BACKGROUND_HEADER_COLOR = "#B3A369"
YEAR_STRIP_COLORS = ("#B3A369", "#FFFFFF", "#003057", "#FFFFFF")

# Static GitHub Pages encryption.  The public repository contains only AES-GCM
# ciphertext.  The 256-bit data key is supplied after server-side access
# verification and is never derived from or bundled with the knowledge answers.
ENCRYPTED_DATA_FILENAME = "tree_data.enc"
ACCESS_SECRETS_FILENAME = "access_secrets.json"

# Order here is used by the legend only. A person's card follows the order in
# that person's Instrument cell, as requested.
SECTION_COLORS: dict[str, str] = {
    "flute/piccolo": "#f5c1ce",
    "clarinet": "#ECD5E3",
    "sax/saxophone": "#F3B0C3",
    "trumpet": "#fa9078",
    "mellophone": "#fcb17e",
    "trombone": "#7cf7b7",
    "baritone": "#8bf0e6",
    "sousaphone": "#f797d2",
    "front ensemble": "#C6DBDA",
    "battery": "#D4F0F0",
    "guard": "#FF968A",
    "goldrush": "#F6EAC2",
    "golden girl": "#FFFFB5",
    "unknown": UNKNOWN_COLOR,
}

# Patterns are intentionally ordered from more specific to less specific.
# Overlapping matches are resolved by keeping the longest match.
INSTRUMENT_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("front ensemble", re.compile(r"\bfront\s+en(?:s|c)emble\b|\bpit\b|\bmarimbas?\b|\bvibraphones?\b|\bvibes?\b|\bxylophones?\b|\bglockenspiels?\b|\bbells?\b|\btimpani\b|\bkettledrums?\b|\brack\b|\baux(?:iliary)?\s+percussion\b|\bkeyboards?\b|\bsynth(?:esizer)?s?\b", re.I)),
    ("golden girl", re.compile(r"\bgolden\s+girls?\b", re.I)),
    ("goldrush", re.compile(r"\bgold\s*rush\b|\bgoldrush\b", re.I)),
    ("guard", re.compile(r"\bcolor\s*guard\b|\bcolorguard\b|\bguard\b|\bflags?\b|\brifles?\b|\bsab(?:er|re)s?\b", re.I)),
    (
        "battery",
        re.compile(
            r"\bbattery\b|\bdrum\s*line\b|\bdrumline\b|\bsnares?\b|"
            r"\btenors\b|\btenor\s+drums?\b|\bquads?\b|\bquints?\b|"
            r"\bbass\s+drums?\b|\bcymbals?\b",
            re.I,
        ),
    ),
    (
        "sax/saxophone",
        re.compile(
            r"\b(?:(?:alto|tenor|baritone|bari|soprano)\s+)?sax(?:ophone)?s?\b",
            re.I,
        ),
    ),
    ("flute/piccolo", re.compile(r"\bflutes?\b|\bpiccolos?\b", re.I)),
    ("clarinet", re.compile(r"\b(?:bass\s+)?clarinets?\b", re.I)),
    ("trumpet", re.compile(r"\btrumpets?\b", re.I)),
    ("mellophone", re.compile(r"\bmellophones?\b|\bmellos?\b", re.I)),
    ("trombone", re.compile(r"\btrombones?\b|\bbones?\b", re.I)),
    ("baritone", re.compile(r"\bbaritones?\b|\beuphoniums?\b", re.I)),
    ("sousaphone", re.compile(r"\bsousaphones?\b|\btubas?\b", re.I)),
)

IGNORED_ROLE_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"\bdrum\s+majors?\b", re.I),
    re.compile(r"\bsection\s+leaders?\b", re.I),
    re.compile(r"\b(?:color\s+guard\s+)?guard\s+captains?\b|\bcolor\s+guard\s+captains?\b", re.I),
    re.compile(r"\bcaptains?\b", re.I),
    re.compile(r"\bleadership\b", re.I),
)

RESIDUAL_STOPWORDS = {
    "and",
    "or",
    "the",
    "a",
    "an",
    "section",
    "line",
    "member",
    "members",
    "former",
    "current",
    "primary",
    "secondary",
    "then",
    "later",
    "also",
    "plus",
}

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "given": ("givenpreferredname", "givenname", "preferredname", "firstname"),
    "nickname": ("nickname", "nick name"),
    "family": ("familymaidenname", "familyname", "maidenname", "lastname", "surname"),
    "married": ("marriedname", "marriedsurname", "spousesurname", "currentlastname"),
    "year": ("ratyear", "year"),
    "instrument": ("instrument", "instruments", "section"),
    "tree_name_preference": ("treedisplaynamepreference", "treenamepreference", "cardnamepreference"),
    "vet": ("vet", "vetsnameratyearandinstruments"),
}

RELATION_RE = re.compile(r"^\s*(.*?)\s*\((.*)\)\s*\(([^()]*)\)\s*$")
YEAR_RE = re.compile(r"(?<!\d)((?:19|20)\d{2})(?!\d)")
TWO_DIGIT_YEAR_RE = re.compile(r"(?<!\d)(\d{2})(?!\d)")


class TreeDataError(RuntimeError):
    """Raised when the workbook cannot be interpreted safely."""


def normalize_spaces(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u00a0", " ").replace("’", "'").replace("`", "'")
    return re.sub(r"\s+", " ", text).strip()


def normalized_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_spaces(value).casefold())


def strict_name_key(value: object) -> str:
    text = normalize_spaces(value).casefold()
    return re.sub(r"\s+", " ", text)


def loose_name_key(value: object) -> str:
    text = unicodedata.normalize("NFKD", normalize_spaces(value).casefold())
    return "".join(char for char in text if char.isalnum())


def safe_filename(value: object, *, max_length: int = 120) -> str:
    text = normalize_spaces(value)
    text = re.sub(r"[<>:\"/\\|?*\x00-\x1F]", "_", text)
    text = text.rstrip(" .") or "unnamed"
    reserved = {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        *(f"COM{i}" for i in range(1, 10)),
        *(f"LPT{i}" for i in range(1, 10)),
    }
    if text.upper() in reserved:
        text = f"_{text}"
    return text[:max_length].rstrip(" .") or "unnamed"


def find_existing_file(filename: str, explicit: Path | None = None) -> Path:
    candidates: list[Path] = []
    if explicit is not None:
        candidates.append(explicit.expanduser().resolve())
    candidates.extend((SCRIPT_DIR / filename, SOURCE_DIR / filename))
    for candidate in candidates:
        if candidate.exists():
            return candidate
    locations = "\n".join(f"  - {candidate}" for candidate in candidates)
    raise FileNotFoundError(f"Could not find {filename!r}. Checked:\n{locations}")


def load_font(size: int, *, bold: bool = False) -> ImageFont.ImageFont:
    candidates: list[Path] = []
    if bold:
        candidates.extend(
            (
                Path("C:/Windows/Fonts/calibrib.ttf"),
                Path("C:/Windows/Fonts/arialbd.ttf"),
                Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
            )
        )
    else:
        candidates.extend(
            (
                Path("C:/Windows/Fonts/calibri.ttf"),
                Path("C:/Windows/Fonts/arial.ttf"),
                Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            )
        )
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size)
    try:
        return ImageFont.truetype(
            "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf", size
        )
    except OSError:
        return ImageFont.load_default()


@dataclass
class IssueLog:
    warnings: list[str] = field(default_factory=list)
    decisions: list[str] = field(default_factory=list)
    skipped_relations: list[str] = field(default_factory=list)
    skipped_people: list[str] = field(default_factory=list)
    skipped_ambiguities: list[str] = field(default_factory=list)
    unknown_instruments: Counter[str] = field(default_factory=Counter)

    def warn(self, message: str) -> None:
        self.warnings.append(message)
        print(f"WARNING: {message}")

    def decision(self, message: str) -> None:
        self.decisions.append(message)
        print(f"Resolved: {message}")


class PromptSession:
    """Interactive decisions with a persistent JSON cache."""

    def __init__(
        self,
        cache_path: Path,
        *,
        interactive: bool,
        issues: IssueLog,
        skip_ambiguities: bool = False,
    ) -> None:
        self.cache_path = cache_path
        self.interactive = interactive and sys.stdin.isatty()
        self.skip_ambiguities = skip_ambiguities
        self.issues = issues
        self.cache: dict[str, dict[str, object]] = {
            "years": {},
            "instruments": {},
            "relations": {},
            "parents": {},
            "edges": {},
        }
        if cache_path.exists():
            try:
                loaded = json.loads(cache_path.read_text(encoding="utf-8"))
                if isinstance(loaded, dict):
                    for section in self.cache:
                        if isinstance(loaded.get(section), dict):
                            self.cache[section].update(loaded[section])
            except (OSError, json.JSONDecodeError) as exc:
                self.issues.warn(f"Could not read resolution cache {cache_path}: {exc}")

    def save(self) -> None:
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        self.cache_path.write_text(
            json.dumps(self.cache, indent=2, ensure_ascii=False, sort_keys=True),
            encoding="utf-8",
        )

    def ask(self, prompt: str, *, default: str | None = None) -> str:
        if not self.interactive:
            return default or ""
        suffix = f" [{default}]" if default else ""
        try:
            answer = input(f"{prompt}{suffix}: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nInput cancelled.")
            raise SystemExit(130)
        return answer or (default or "")

    def yes_no(self, prompt: str, *, default: bool = True) -> bool:
        default_text = "Y/n" if default else "y/N"
        if not self.interactive:
            return default
        while True:
            answer = self.ask(f"{prompt} ({default_text})").casefold()
            if not answer:
                return default
            if answer in {"y", "yes"}:
                return True
            if answer in {"n", "no"}:
                return False
            print("Please enter y or n.")

    def yes_no_skip(self, prompt: str, *, default: bool = True) -> bool | None:
        """Return True/False, or None when the user chooses to skip ambiguity.

        Skip is deliberately a third state rather than being treated as "no" so
        callers can omit the uncertain datum/edge instead of making a guessed
        correction.  --skip-ambiguities chooses this state automatically.
        """
        if self.skip_ambiguities:
            print(f"Skipped ambiguity: {prompt}")
            return None
        if not self.interactive:
            return default
        default_text = "Y/n/s" if default else "y/N/s"
        while True:
            answer = self.ask(f"{prompt} ({default_text}; s=skip)").casefold()
            if not answer:
                return default
            if answer in {"y", "yes"}:
                return True
            if answer in {"n", "no"}:
                return False
            if answer in {"s", "skip", "u", "unknown"}:
                return None
            print("Please enter y, n, or s to skip this ambiguity.")

    def choose(
        self,
        prompt: str,
        options: Sequence[str],
        *,
        default_index: int = 0,
        allow_skip: bool = False,
    ) -> int | None:
        if not options:
            return None
        if allow_skip and self.skip_ambiguities:
            print(f"Skipped ambiguity: {prompt}")
            return None
        if not self.interactive:
            return default_index
        print(prompt)
        for index, option in enumerate(options, start=1):
            print(f"  {index}. {option}")
        if allow_skip:
            print("  0. Skip")
        while True:
            answer = self.ask("Choose a number", default=str(default_index + 1))
            if allow_skip and answer == "0":
                return None
            if answer.isdigit() and 1 <= int(answer) <= len(options):
                return int(answer) - 1
            print("Please choose one of the listed numbers.")

    def resolve_year(self, raw_value: object, *, person_name: str, row: int) -> int | None:
        raw = normalize_spaces(raw_value)
        cache_key = f"row {row}|{person_name}|{raw}"
        cached = self.cache["years"].get(cache_key)
        if isinstance(cached, int):
            return cached
        if cached is None and cache_key in self.cache["years"]:
            return None

        candidates = [int(value) for value in YEAR_RE.findall(raw)]
        candidates = list(dict.fromkeys(candidates))
        current_year = datetime.now().year

        if len(candidates) == 1:
            result = candidates[0]
        elif len(candidates) > 1:
            prompt = (
                f'Row {row}, {person_name}: RAT Year {raw!r} contains multiple years. '
                "Which one is the RAT year?"
            )
            selected = self.choose(prompt, [str(year) for year in candidates], allow_skip=True)
            result = candidates[selected] if selected is not None else None
            if not self.interactive and not self.skip_ambiguities:
                self.issues.warn(
                    f"{person_name} has ambiguous RAT Year {raw!r}; used {result}."
                )
            elif selected is None:
                self.issues.skipped_ambiguities.append(
                    f"Row {row}, {person_name}: skipped ambiguous RAT Year {raw!r}."
                )
        else:
            short_years = [int(value) for value in TWO_DIGIT_YEAR_RE.findall(raw)]
            short_years = list(dict.fromkeys(short_years))
            proposed: int | None = None
            if len(short_years) == 1:
                short = short_years[0]
                proposed = 2000 + short if short <= (current_year + 1) % 100 else 1900 + short
            if proposed is not None:
                decision = self.yes_no_skip(
                    f'Row {row}, {person_name}: did you mean RAT Year {proposed} for {raw!r}?',
                    default=True,
                )
                if decision is True:
                    result = proposed
                elif decision is False:
                    result = self._request_year(person_name, row, raw)
                else:
                    result = None
                    self.issues.skipped_ambiguities.append(
                        f"Row {row}, {person_name}: skipped ambiguous RAT Year {raw!r}."
                    )
            elif raw:
                result = self._request_year(person_name, row, raw)
            else:
                result = self._request_year(person_name, row, "blank")

        if result is not None and not (1950 <= result <= current_year + 2):
            decision = self.yes_no_skip(
                f"Row {row}, {person_name}: {result} is an unusual RAT year. Keep it?",
                default=False,
            )
            if decision is False:
                result = self._request_year(person_name, row, raw)
            elif decision is None:
                self.issues.skipped_ambiguities.append(
                    f"Row {row}, {person_name}: skipped unusual RAT Year {result}."
                )
                result = None

        self.cache["years"][cache_key] = result
        return result

    def _request_year(self, person_name: str, row: int, raw: str) -> int | None:
        if self.skip_ambiguities:
            self.issues.skipped_ambiguities.append(
                f"Row {row}, {person_name}: skipped unresolved RAT Year {raw!r}."
            )
            return None
        if not self.interactive:
            self.issues.warn(
                f"Row {row}, {person_name}: could not determine RAT Year from {raw!r}; "
                "left unresolved."
            )
            return None
        while True:
            answer = self.ask(
                f"Row {row}, {person_name}: enter the four-digit RAT year, or S to skip this person"
            )
            if answer.casefold() in {"s", "skip", "u", "unknown"}:
                self.issues.skipped_ambiguities.append(
                    f"Row {row}, {person_name}: skipped unresolved RAT Year {raw!r}."
                )
                return None
            if answer.isdigit() and len(answer) == 4:
                return int(answer)
            print("Please enter a four-digit year or S to skip this person.")

    def resolve_unknown_instrument(
        self,
        fragment: str,
        *,
        person_name: str,
        row: int,
        full_text: str,
    ) -> str | None:
        key = normalize_spaces(fragment).casefold()
        cached = self.cache["instruments"].get(key)
        if isinstance(cached, str):
            return None if cached == "__ignore__" else cached

        choices = list(SECTION_COLORS)
        close = difflib.get_close_matches(key, choices[:-1], n=3, cutoff=0.35)
        if self.skip_ambiguities:
            result = None
            self.issues.skipped_ambiguities.append(
                f"Row {row}, {person_name}: ignored ambiguous section fragment {fragment!r}."
            )
            print(
                f"Skipped ambiguous section text for {person_name}: {fragment!r}."
            )
        elif not self.interactive:
            result = "unknown"
            self.issues.unknown_instruments[fragment] += 1
            self.issues.warn(
                f"Row {row}, {person_name}: unrecognized instrument text {fragment!r} "
                f"in {full_text!r}; colored as unknown."
            )
        else:
            print(
                f"\nRow {row}, {person_name}: unrecognized instrument/section text "
                f"{fragment!r} in {full_text!r}."
            )
            decision = (
                    self.yes_no_skip(f'Did you mean "{close[0]}"?', default=True)
                    if close else False
                )
            if decision is True:
                result = close[0]
            elif decision is None:
                result = None
                self.issues.skipped_ambiguities.append(
                    f"Row {row}, {person_name}: ignored ambiguous section fragment {fragment!r}."
                )
            else:
                menu = choices + ["ignore this fragment"]
                selected = self.choose(
                    "How should this text be colored?", menu, allow_skip=True
                )
                if selected is None or selected == len(menu) - 1:
                    result = None
                    self.issues.skipped_ambiguities.append(
                        f"Row {row}, {person_name}: ignored ambiguous section fragment {fragment!r}."
                    )
                else:
                    result = menu[selected]
        self.cache["instruments"][key] = result if result is not None else "__ignore__"
        return result

    def resolve_relation_target(
        self,
        *,
        raw_reference: str,
        parsed_name: str,
        source_description: str,
        candidates: Sequence["Person"],
    ) -> "Person" | None:
        cache_key = f"{source_description}|{raw_reference}"
        cached = self.cache["relations"].get(cache_key)
        if isinstance(cached, str):
            for candidate in candidates:
                if candidate.person_id == cached:
                    return candidate
            # A cached skip is an old ambiguity decision, not authoritative data.
            # If the workbook now resolves to one exact candidate (for example after
            # a person was added/fixed online), allow the current data to win.
            if cached == "__skip__" and len(candidates) != 1:
                return None

        if not candidates:
            self.issues.skipped_relations.append(
                f"{source_description}: no person matches {raw_reference!r}"
            )
            if self.interactive:
                print(
                    f"\n{source_description}: could not match relationship entry "
                    f"{raw_reference!r}."
                )
                print(
                    "Correct the name in the workbook, or add that person to the master sheet, "
                    "then rerun the generator."
                )
                self.ask("Press Enter to skip this relationship for the current run", default="")
            self.cache["relations"][cache_key] = "__skip__"
            return None

        if len(candidates) == 1:
            candidate = candidates[0]
            if strict_name_key(candidate.stable_name) != strict_name_key(parsed_name):
                if self.skip_ambiguities:
                    decision = None
                elif self.interactive:
                    decision = self.yes_no_skip(
                        f'{source_description}: did you mean "{candidate.stable_name}" for '
                        f'{parsed_name!r}?',
                        default=True,
                    )
                else:
                    decision = True
                if decision is not True:
                    self.cache["relations"][cache_key] = "__skip__"
                    self.issues.skipped_relations.append(
                        f"{source_description}: skipped ambiguous match {raw_reference!r}."
                    )
                    return None
            self.cache["relations"][cache_key] = candidate.person_id
            return candidate

        selected = self.choose(
            f"{source_description}: which person does {raw_reference!r} mean?",
            [f"{person.stable_name} ({person.year_label}, row {person.row})" for person in candidates],
            allow_skip=True,
        )
        if selected is None:
            self.cache["relations"][cache_key] = "__skip__"
            return None
        result = candidates[selected]
        self.cache["relations"][cache_key] = result.person_id
        return result

    def choose_parent(self, child: "Person", parents: Sequence["Person"]) -> "Person" | None:
        key = child.person_id
        cached = self.cache["parents"].get(key)
        if isinstance(cached, str):
            if cached == "__skip__" and len(parents) > 1:
                return None
            for parent in parents:
                if parent.person_id == cached:
                    return parent
        selected = self.choose(
            f"{child.stable_name} has multiple possible VETs. Which relationship is correct?",
            [f"{parent.stable_name} ({parent.year_label})" for parent in parents],
            allow_skip=True,
        )
        if selected is None:
            self.cache["parents"][key] = "__skip__"
            self.issues.skipped_relations.append(
                f"{child.stable_name}: skipped multiple possible VETs; no parent edge was used."
            )
            return None
        result = parents[selected]
        self.cache["parents"][key] = result.person_id
        return result


@dataclass(frozen=True)
class PixelPoint:
    """One exact pixel coordinate in either card-local or scene-global space."""

    x: int
    y: int

    def translated(self, dx: int, dy: int) -> "PixelPoint":
        return PixelPoint(self.x + dx, self.y + dy)

    def as_dict(self) -> dict[str, int]:
        return {"x": self.x, "y": self.y}


@dataclass(frozen=True)
class CardPlacement:
    """Placement of one card inside one rendered scene."""

    scene_id: str
    top_left: PixelPoint
    vet_connection: PixelPoint
    rat_connection: PixelPoint
    width: int
    height: int

    @property
    def x(self) -> int:
        return self.top_left.x

    @property
    def y(self) -> int:
        return self.top_left.y

    @property
    def center_x(self) -> int:
        return self.top_left.x + self.width // 2

    @property
    def bounds(self) -> tuple[int, int, int, int]:
        return (self.x, self.y, self.x + self.width, self.y + self.height)


@dataclass
class CardObject:
    """Rendered name card plus geometry shared by PNG, SVG, and the web export.

    ``local_vet_connection`` is the top-center connection pixel on the card
    image. ``local_rat_connection`` is the bottom-center connection pixel.
    The two ``global_*`` fields are reserved for the card's placement in the
    complete full-band scene, exactly as requested; per-family placements are
    retained separately in ``placements`` and never overwrite those globals.
    """

    person_id: str
    image: Image.Image = field(repr=False)
    section_colors: list[str] = field(default_factory=list)
    local_vet_connection: PixelPoint = field(init=False)
    local_rat_connection: PixelPoint = field(init=False)
    global_top_left: PixelPoint | None = None
    global_vet_connection: PixelPoint | None = None
    global_rat_connection: PixelPoint | None = None
    placements: dict[str, CardPlacement] = field(default_factory=dict, repr=False)

    def __post_init__(self) -> None:
        # The old renderer deliberately let connector rectangles touch a pixel
        # inside the card edge. Keeping y=1 and y=height-2 preserves that look.
        center_x = self.image.width // 2
        self.local_vet_connection = PixelPoint(center_x, 1)
        self.local_rat_connection = PixelPoint(center_x, self.image.height - 2)

    @property
    def width(self) -> int:
        return self.image.width

    @property
    def height(self) -> int:
        return self.image.height

    def place(self, scene_id: str, x: int, y: int, *, is_full_tree: bool = False) -> CardPlacement:
        top_left = PixelPoint(int(x), int(y))
        placement = CardPlacement(
            scene_id=scene_id,
            top_left=top_left,
            vet_connection=self.local_vet_connection.translated(top_left.x, top_left.y),
            rat_connection=self.local_rat_connection.translated(top_left.x, top_left.y),
            width=self.width,
            height=self.height,
        )
        self.placements[scene_id] = placement
        if is_full_tree:
            self.global_top_left = placement.top_left
            self.global_vet_connection = placement.vet_connection
            self.global_rat_connection = placement.rat_connection
        return placement


@dataclass(frozen=True)
class ConnectorSegment:
    """Axis-aligned centerline segment used to build the old rectangle style."""

    start: PixelPoint
    end: PixelPoint

    @property
    def orientation(self) -> str:
        return "vertical" if self.start.x == self.end.x else "horizontal"

    def as_rect(self, width: int = CONNECTOR_WIDTH) -> tuple[int, int, int, int]:
        half = width // 2
        if self.orientation == "vertical":
            x = self.start.x - half
            y = min(self.start.y, self.end.y)
            height = max(1, abs(self.end.y - self.start.y))
            return (x, y, width, height)
        x = min(self.start.x, self.end.x)
        y = self.start.y - half
        rect_width = max(1, abs(self.end.x - self.start.x) + width)
        return (x, y, rect_width, width)

    def as_dict(self) -> dict[str, object]:
        return {
            "start": self.start.as_dict(),
            "end": self.end.as_dict(),
            "orientation": self.orientation,
        }


@dataclass
class FamilyConnector:
    parent_id: str
    child_ids: list[str]
    parent_stem: ConnectorSegment
    child_stems: list[ConnectorSegment]
    sibling_bus: ConnectorSegment | None = None

    @property
    def segments(self) -> list[ConnectorSegment]:
        result = [self.parent_stem]
        if self.sibling_bus is not None:
            result.append(self.sibling_bus)
        result.extend(self.child_stems)
        return result


@dataclass
class Person:
    person_id: str
    row: int
    given: str
    nickname: str
    family: str
    married: str
    tree_name_preference: str
    year_raw: object
    year: int | None
    instrument_raw: str
    instruments: list[str]
    vet_raw: str
    rat_raws: list[tuple[str, str]]
    source_fields: list[dict[str, str]] = field(default_factory=list)
    parent_id: str | None = None
    children_ids: list[str] = field(default_factory=list)

    @property
    def stable_name(self) -> str:
        return normalize_spaces(f"{self.given} {self.family}") or self.given or self.family

    @property
    def current_name(self) -> str:
        married = normalize_spaces(self.married)
        if married:
            # Historical rows are not perfectly uniform: most Married Name
            # cells contain only the current surname, but a few may already
            # contain a full name. Avoid duplicating the given name in that case.
            married_key = loose_name_key(married)
            given_key = loose_name_key(self.given)
            nickname_key = loose_name_key(self.nickname)
            if (given_key and married_key.startswith(given_key)) or (nickname_key and married_key.startswith(nickname_key)):
                return married
        surname = married or self.family
        return normalize_spaces(f"{self.given} {surname}")

    @property
    def card_given_name(self) -> str:
        preference = normalized_header(self.tree_name_preference)
        if preference in {"nickname", "usenickname"} and self.nickname:
            return f'"{self.nickname}"'
        if preference in {"both", "givenpreferrednickname", "givenpreferredandnickname", "firstnickname", "firstnameandnickname"} and self.nickname:
            return normalize_spaces(f'{self.given} "{self.nickname}"')
        return self.given or self.nickname

    @property
    def card_display_name(self) -> str:
        return normalize_spaces(f"{self.card_given_name} {self.family}") or self.stable_name

    @property
    def currently_rat(self) -> bool:
        return truthy(self.source_field_value("Currently a RAT", "Current RAT"))

    @property
    def band_club_leadership(self) -> bool:
        return bool(
            truthy(self.source_field_value("Served in Band Club Leadership Position", "Band Club Leadership"))
            or self.source_field_value("Band Club Leadership Position(s)", "Band Club Leadership History")
        )

    @property
    def favorite_tech_band_memory(self) -> str:
        return self.source_field_value("Favorite Tech Band Memory")

    @property
    def section_nicknames(self) -> str:
        return self.source_field_value("Section Nicknames", "Section Nickname(s)")

    @property
    def marching_band_leadership_history(self) -> str:
        return self.source_field_value("Marching Band Leadership History")

    @property
    def band_club_leadership_history(self) -> str:
        return self.source_field_value("Band Club Leadership History")

    def source_field_value(self, *labels: str) -> str:
        wanted = {normalized_header(label) for label in labels}
        for item in self.source_fields:
            if normalized_header(item.get("label", "")) in wanted:
                return normalize_spaces(item.get("value", ""))
        return ""

    @property
    def leadership_icons(self) -> list[str]:
        formal = self.source_field_value(
            "Marching Band Leadership Role(s)",
            "Formal Leadership Position(s)",
            "Formal Leadership Role(s)",
        )
        informal = self.source_field_value(
            "Informal Leadership Position(s)",
            "Informal Leadership Role(s)",
        )
        informal_flag = self.source_field_value(
            "Served in Informal Leadership Position",
            "Informal Leadership",
        )
        return leadership_icon_flags(formal, informal, informal_flag)

    @property
    def year_label(self) -> str:
        return str(self.year) if self.year is not None else "Unknown Year"

    @property
    def name_aliases(self) -> list[str]:
        aliases = {self.stable_name, self.current_name}
        if self.nickname:
            aliases.add(normalize_spaces(f"{self.nickname} {self.family}"))
            aliases.add(normalize_spaces(f"{self.given} {self.nickname} {self.family}"))
        return sorted((alias for alias in aliases if alias), key=str.casefold)


@dataclass(frozen=True)
class RelationReference:
    raw: str
    name: str
    year_candidates: tuple[int, ...]
    instrument_text: str


@dataclass
class EdgeSource:
    parent_id: str
    child_id: str
    source: str
    order: tuple[int, int, int]


@dataclass
class Scene:
    scene_id: str
    title: str
    header_height: int
    people: list[Person]
    roots: list[Person]
    width: int
    height: int
    min_year: int | None
    max_year: int | None
    unknown_year: bool
    placements: dict[str, CardPlacement]
    connectors: list[FamilyConnector]

    @property
    def connector_rects(self) -> list[tuple[int, int, int, int]]:
        return [
            segment.as_rect(CONNECTOR_WIDTH)
            for connector in self.connectors
            for segment in connector.segments
        ]

    @property
    def connector_outline_rects(self) -> list[tuple[int, int, int, int]]:
        # v17.4: no white halo; keep this property for renderer compatibility.
        return []


def discover_headers(ws) -> tuple[int, dict[str, int], list[tuple[str, int]]]:
    header_row: int | None = None
    normalized_by_column: dict[int, str] = {}
    for row in range(1, min(ws.max_row, 10) + 1):
        candidate = {
            column: normalized_header(ws.cell(row, column).value)
            for column in range(1, ws.max_column + 1)
        }
        values = set(candidate.values())
        if "givenpreferredname" in values and "familymaidenname" in values:
            header_row = row
            normalized_by_column = candidate
            break
    if header_row is None:
        raise TreeDataError(
            f"Worksheet {ws.title!r} does not contain the four-column name headers. "
            "Run migrate_name_columns.py first."
        )

    mapping: dict[str, int] = {}
    for field_name, aliases in HEADER_ALIASES.items():
        alias_keys = {normalized_header(alias) for alias in aliases}
        for column, header in normalized_by_column.items():
            if header in alias_keys:
                mapping[field_name] = column
                break

    required = ("given", "nickname", "family", "married", "year", "instrument", "vet")
    missing = [field_name for field_name in required if field_name not in mapping]
    if missing:
        raise TreeDataError(
            f"Worksheet {ws.title!r} is missing required columns: {', '.join(missing)}"
        )

    rat_columns: list[tuple[str, int]] = []
    for column, header in normalized_by_column.items():
        match = re.fullmatch(r"rat(\d+)", header)
        if match:
            rat_columns.append((f"RAT {int(match.group(1))}", column))
    rat_columns.sort(key=lambda item: int(item[0].split()[1]))
    if not rat_columns:
        raise TreeDataError(f"Worksheet {ws.title!r} has no RAT relationship columns.")

    return header_row, mapping, rat_columns


def select_non_overlapping_matches(
    text: str,
) -> tuple[list[tuple[int, int, str]], list[tuple[int, int]]]:
    section_matches: list[tuple[int, int, str]] = []
    ignored_matches: list[tuple[int, int]] = []

    raw_matches: list[tuple[int, int, str]] = []
    for category, pattern in INSTRUMENT_PATTERNS:
        for match in pattern.finditer(text):
            raw_matches.append((match.start(), match.end(), category))
    # Longest match wins when patterns overlap (for example, "baritone saxophone").
    raw_matches.sort(key=lambda item: (item[0], -(item[1] - item[0])))
    occupied: set[int] = set()
    for start, end, category in raw_matches:
        if any(position in occupied for position in range(start, end)):
            continue
        section_matches.append((start, end, category))
        occupied.update(range(start, end))

    for pattern in IGNORED_ROLE_PATTERNS:
        for match in pattern.finditer(text):
            if any(position in occupied for position in range(match.start(), match.end())):
                continue
            ignored_matches.append((match.start(), match.end()))
            occupied.update(range(match.start(), match.end()))

    section_matches.sort(key=lambda item: item[0])
    return section_matches, ignored_matches


def normalize_instruments(
    raw_value: object,
    *,
    person_name: str,
    row: int,
    prompts: PromptSession,
) -> list[str]:
    raw = normalize_spaces(raw_value)
    if not raw:
        return ["unknown"]

    matches, ignored = select_non_overlapping_matches(raw)
    covered = [False] * len(raw)
    positioned_categories: list[tuple[int, str]] = []
    for start, end, category in matches:
        for position in range(start, end):
            covered[position] = True
        positioned_categories.append((start, category))
    for start, end in ignored:
        for position in range(start, end):
            covered[position] = True

    # v16 normalized values use an em/en dash to separate the canonical broad
    # section from preserved subsection detail (for example, ``Trumpet —
    # Flugelhorn``). Once a broad section has been recognized, that detail is
    # descriptive data, not a second section that needs a color decision.
    if matches:
        for detail_match in re.finditer(r"[—–][^;]*", raw):
            for position in range(detail_match.start(), detail_match.end()):
                covered[position] = True

    residual_chars = [char if not covered[index] else " " for index, char in enumerate(raw)]
    residual = "".join(residual_chars)
    residual = re.sub(r"[,;/&+|():_—–\-]+", " ", residual)
    words = [word for word in normalize_spaces(residual).split() if word.casefold() not in RESIDUAL_STOPWORDS]

    if words:
        fragment = " ".join(words)
        resolution = prompts.resolve_unknown_instrument(
            fragment,
            person_name=person_name,
            row=row,
            full_text=raw,
        )
        if resolution:
            first_position = next(
                (index for index, is_covered in enumerate(covered) if not is_covered and raw[index].isalnum()),
                len(raw),
            )
            positioned_categories.append((first_position, resolution))

    positioned_categories.sort(key=lambda item: item[0])
    result: list[str] = []
    for _, category in positioned_categories:
        if category not in result:
            result.append(category)
    return result or ["unknown"]


def uncategorized_instrument_text(raw_value: object) -> str:
    """Return meaningful instrument text not explained by the section taxonomy.

    This is exported for the administrator's unknown-instrument queue. Preserved
    subsection detail after a canonical dash (for example ``Trumpet — Flugelhorn``)
    is intentionally not flagged because the broad section is already known.
    """
    raw = normalize_spaces(raw_value)
    if not raw:
        return ""
    matches, ignored = select_non_overlapping_matches(raw)
    covered = [False] * len(raw)
    for start, end, _ in matches:
        for position in range(start, end):
            covered[position] = True
    for start, end in ignored:
        for position in range(start, end):
            covered[position] = True
    if matches:
        for detail_match in re.finditer(r"[—–][^;]*", raw):
            for position in range(detail_match.start(), detail_match.end()):
                covered[position] = True
    residual = "".join(char if not covered[index] else " " for index, char in enumerate(raw))
    residual = re.sub(r"[,;/&+|():_—–\-]+", " ", residual)
    words = [word for word in normalize_spaces(residual).split() if word.casefold() not in RESIDUAL_STOPWORDS]
    return " ".join(words)


def parse_relation(raw_value: object) -> RelationReference | None:
    raw = normalize_spaces(raw_value)
    if not raw:
        return None
    match = RELATION_RE.match(raw)
    if match:
        name = normalize_spaces(match.group(1))
        year_text = normalize_spaces(match.group(2))
        instrument_text = normalize_spaces(match.group(3))
    else:
        name = normalize_spaces(raw.split(" (", 1)[0])
        year_text = ""
        instrument_text = ""
    years = tuple(dict.fromkeys(int(value) for value in YEAR_RE.findall(year_text)))
    return RelationReference(raw=raw, name=name, year_candidates=years, instrument_text=instrument_text)


def load_people(
    workbook_path: Path,
    *,
    sheet_name: str,
    prompts: PromptSession,
    issues: IssueLog,
) -> tuple[list[Person], object]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if sheet_name not in workbook.sheetnames:
        if prompts.interactive:
            selected = prompts.choose(
                f"Worksheet {sheet_name!r} was not found. Choose the master sheet:",
                workbook.sheetnames,
            )
            sheet_name = workbook.sheetnames[selected or 0]
        else:
            workbook.close()
            raise TreeDataError(
                f"Worksheet {sheet_name!r} not found. Available sheets: {workbook.sheetnames}"
            )
    ws = workbook[sheet_name]
    header_row, mapping, rat_columns = discover_headers(ws)
    ordered_headers = [
        (column, normalize_spaces(ws.cell(header_row, column).value) or f"Column {column}")
        for column in range(1, ws.max_column + 1)
    ]
    header_keys = {normalized_header(label) for _, label in ordered_headers}
    if normalized_header("Favorite Tech Band Memory") not in header_keys:
        workbook.close()
        raise TreeDataError(
            "The v17 workbook field 'Favorite Tech Band Memory' is missing. "
            "Run migrateWorkbookV17.py --apply first so memories cannot be silently dropped "
            "between the workbook, encrypted website data, and protected workbook export."
        )

    # Confirm the current J:P relationship layout without hard-coding it.
    relationship_columns = [mapping["vet"], *(column for _, column in rat_columns)]
    expected = list(range(10, 10 + len(relationship_columns)))
    if relationship_columns != expected:
        issues.warn(
            "VET/RAT relationship columns are not contiguous starting at J. "
            "The script will use the headers rather than fixed positions."
        )

    people: list[Person] = []
    # Display names are not unique identifiers.  Two or more people may legitimately
    # share the exact same Given/Preferred + Family/Maiden name.  Every card/person
    # is identified internally by its worksheet-row-backed person_id (``row-N``).
    seen_strict: dict[str, list[Person]] = defaultdict(list)
    for row in range(header_row + 1, ws.max_row + 1):
        given = normalize_spaces(ws.cell(row, mapping["given"]).value)
        nickname = normalize_spaces(ws.cell(row, mapping["nickname"]).value)
        family = normalize_spaces(ws.cell(row, mapping["family"]).value)
        married = normalize_spaces(ws.cell(row, mapping["married"]).value)
        tree_name_preference = (
            normalize_spaces(ws.cell(row, mapping["tree_name_preference"]).value)
            if "tree_name_preference" in mapping
            else "Given/Preferred Name"
        )
        if not any((given, nickname, family, married)):
            continue
        stable_name = normalize_spaces(f"{given} {family}") or given or family
        if not stable_name:
            issues.warn(f"Row {row} has name data but no usable stable name; skipped.")
            continue
        year_raw = ws.cell(row, mapping["year"]).value
        year = prompts.resolve_year(year_raw, person_name=stable_name, row=row)
        if year is None:
            # Never remove a workbook person from the website/admin data just because
            # RAT Year is blank or ambiguous.  The layout already supports an
            # Unknown Year band.  Omitting the row here made the encrypted site a
            # lossy projection of the authoritative workbook and could also make
            # otherwise valid RAT/VET references disappear.
            issues.warn(
                f"Row {row}, {stable_name}: RAT Year is unresolved; keeping the person "
                "in the site/admin dataset in the Unknown Year band."
            )
        instrument_raw = normalize_spaces(ws.cell(row, mapping["instrument"]).value)
        instruments = normalize_instruments(
            instrument_raw,
            person_name=stable_name,
            row=row,
            prompts=prompts,
        )
        vet_raw = normalize_spaces(ws.cell(row, mapping["vet"]).value)
        rat_raws = [
            (header, normalize_spaces(ws.cell(row, column).value))
            for header, column in rat_columns
            if normalize_spaces(ws.cell(row, column).value)
        ]
        source_fields: list[dict[str, str]] = []
        for column, label in ordered_headers:
            cell = ws.cell(row, column)
            value = normalize_spaces(cell.value)
            hyperlink = getattr(cell, "hyperlink", None)
            target = normalize_spaces(getattr(hyperlink, "target", "")) if hyperlink else ""
            if target:
                if value and value != target:
                    value = f"{value} ({target})"
                else:
                    value = target
            source_fields.append({"label": label, "value": value})

        person_id = f"row-{row}"
        person = Person(
            person_id=person_id,
            row=row,
            given=given,
            nickname=nickname,
            family=family,
            married=married,
            tree_name_preference=tree_name_preference,
            year_raw=year_raw,
            year=year,
            instrument_raw=instrument_raw,
            instruments=instruments,
            vet_raw=vet_raw,
            rat_raws=rat_raws,
            source_fields=source_fields,
        )
        key = strict_name_key(person.stable_name)
        if seen_strict[key]:
            existing_rows = ", ".join(str(existing.row) for existing in seen_strict[key])
            issues.warn(
                f"Duplicate display/stable name {person.stable_name!r} appears in row(s) "
                f"{existing_rows} and {row}. Keeping all cards as separate row-based people."
            )
        seen_strict[key].append(person)
        people.append(person)

    if not people:
        workbook.close()
        raise TreeDataError(f"No people were found in worksheet {ws.title!r}.")
    return people, workbook


def build_alias_indexes(
    people: Sequence[Person],
) -> tuple[dict[str, list[Person]], dict[str, list[Person]]]:
    strict: dict[str, list[Person]] = defaultdict(list)
    loose: dict[str, list[Person]] = defaultdict(list)
    for person in people:
        for alias in person.name_aliases:
            strict[strict_name_key(alias)].append(person)
            loose[loose_name_key(alias)].append(person)
    return strict, loose


def candidate_people_for_reference(
    reference: RelationReference,
    people: Sequence[Person],
    strict_index: dict[str, list[Person]],
    loose_index: dict[str, list[Person]],
) -> list[Person]:
    candidates = list(strict_index.get(strict_name_key(reference.name), ()))
    if not candidates:
        candidates = list(loose_index.get(loose_name_key(reference.name), ()))
    if candidates and reference.year_candidates:
        year_matches = [person for person in candidates if person.year in reference.year_candidates]
        if year_matches:
            candidates = year_matches
    if len(candidates) > 1 and reference.instrument_text:
        ref_sections = {category for _, _, category in select_non_overlapping_matches(reference.instrument_text)[0]}
        if ref_sections:
            section_matches = [
                person for person in candidates
                if ref_sections.intersection(set(person.instruments) - {"unknown"})
            ]
            if section_matches:
                candidates = section_matches
    if candidates:
        return sorted(candidates, key=lambda person: (person.year or 9999, person.row))

    names = [person.stable_name for person in people]
    close_names = difflib.get_close_matches(reference.name, names, n=5, cutoff=0.58)
    result: list[Person] = []
    for name in close_names:
        result.extend(strict_index.get(strict_name_key(name), ()))
    # Keep unique IDs in suggestion order.
    unique: list[Person] = []
    seen: set[str] = set()
    for person in result:
        if person.person_id not in seen:
            seen.add(person.person_id)
            unique.append(person)
    return unique


def add_edge_source(
    edge_sources: dict[tuple[str, str], list[EdgeSource]],
    parent: Person,
    child: Person,
    *,
    source: str,
    order: tuple[int, int, int],
) -> None:
    if parent.person_id == child.person_id:
        return
    edge_sources[(parent.person_id, child.person_id)].append(
        EdgeSource(parent.person_id, child.person_id, source, order)
    )


def clear_person_source_field(person: Person, label: str) -> None:
    """Blank one exported profile field without modifying the Excel workbook.

    This is used when the local builder skips an ambiguous relationship. The
    source workbook remains untouched, but the generated website/tree will not
    present the unresolved value as though it had been accepted.
    """
    wanted = normalized_header(label)
    for item in person.source_fields:
        if normalized_header(item.get("label", "")) == wanted:
            item["value"] = ""


def build_relationships(
    people: list[Person],
    *,
    prompts: PromptSession,
    issues: IssueLog,
) -> tuple[list[Person], dict[tuple[str, str], list[EdgeSource]]]:
    strict_index, loose_index = build_alias_indexes(people)
    edge_sources: dict[tuple[str, str], list[EdgeSource]] = defaultdict(list)

    for person in people:
        if person.vet_raw:
            reference = parse_relation(person.vet_raw)
            if reference and reference.name:
                candidates = candidate_people_for_reference(
                    reference, people, strict_index, loose_index
                )
                parent = prompts.resolve_relation_target(
                    raw_reference=reference.raw,
                    parsed_name=reference.name,
                    source_description=f"Row {person.row} VET for {person.stable_name}",
                    candidates=candidates,
                )
                if parent:
                    add_edge_source(
                        edge_sources,
                        parent,
                        person,
                        source=f"row {person.row} VET",
                        order=(0, person.row, 0),
                    )
                else:
                    # Preserve the workbook claim in sourceFields/relationshipClaims.
                    # Only the visual edge is omitted when it cannot be resolved.
                    pass
            else:
                issues.skipped_relations.append(
                    f"Row {person.row} VET: could not parse {person.vet_raw!r}"
                )

        for rat_index, (column_name, raw_rat) in enumerate(person.rat_raws, start=1):
            reference = parse_relation(raw_rat)
            if not reference or not reference.name:
                issues.skipped_relations.append(
                    f"Row {person.row} {column_name}: could not parse {raw_rat!r}"
                )
                # Preserve the raw workbook cell for Admin Spreadsheet and the
                # sidebar even when no visual edge can be built.
                continue
            candidates = candidate_people_for_reference(
                reference, people, strict_index, loose_index
            )
            child = prompts.resolve_relation_target(
                raw_reference=reference.raw,
                parsed_name=reference.name,
                source_description=f"Row {person.row} {column_name} for {person.stable_name}",
                candidates=candidates,
            )
            if child:
                add_edge_source(
                    edge_sources,
                    person,
                    child,
                    source=f"row {person.row} {column_name}",
                    order=(1, person.row, rat_index),
                )
            else:
                # Preserve the raw workbook claim; unresolved layout is not a data edit.
                pass
        # Do not replace rat_raws with only visually-resolved relationships.  The
        # website/admin spreadsheet must remain a faithful view of the workbook.

    people_by_id = {person.person_id: person for person in people}
    possible_parents: dict[str, set[str]] = defaultdict(set)
    for parent_id, child_id in edge_sources:
        possible_parents[child_id].add(parent_id)

    selected_edges: set[tuple[str, str]] = set()
    for child_id, parent_ids in possible_parents.items():
        child = people_by_id[child_id]
        parents = [people_by_id[parent_id] for parent_id in parent_ids]
        if len(parents) == 1:
            chosen = parents[0]
        else:
            parents.sort(key=lambda person: (person.year or 9999, person.row))
            chosen = prompts.choose_parent(child, parents)
            if chosen is not None:
                issues.decision(
                    f"Selected {chosen.stable_name} as VET for {child.stable_name}."
                )
        if chosen is not None:
            selected_edges.add((chosen.person_id, child.person_id))
        else:
            # A visual tree has one parent slot, so omit conflicting visual edges.
            # The workbook claims themselves remain intact and are still exported
            # for the sidebar/Admin reciprocity tools.
            for parent_id in list(parent_ids):
                edge_sources.pop((parent_id, child.person_id), None)
            # Keep the child's VET source field/claim visible even though the
            # visual single-parent tree cannot choose among conflicting parents.

    for person in people:
        person.parent_id = None
        person.children_ids.clear()
    for parent_id, child_id in selected_edges:
        people_by_id[child_id].parent_id = parent_id
        people_by_id[parent_id].children_ids.append(child_id)

    def child_order(parent_id: str, child_id: str) -> tuple[int, str, str, int, tuple[int, int, int]]:
        """Deterministic visual sibling order.

        The legacy files happen to have RAT columns already ordered this way,
        but the old renderer relied on that spreadsheet order implicitly.  The
        full-band renderer makes the intended rule explicit: earliest RAT year
        first, then Family/Maiden Name, then Given/Preferred Name.  The source
        RAT-column order is retained only as a final deterministic tiebreaker.
        """
        sources = edge_sources[(parent_id, child_id)]
        best_source = min(
            (source.order for source in sources),
            key=lambda value: (0 if value[0] == 1 else 1, value[1], value[2]),
        )
        child = people_by_id[child_id]
        return (
            child.year or 9999,
            child.family.casefold(),
            child.given.casefold(),
            child.row,
            best_source,
        )

    for parent in people:
        parent.children_ids.sort(key=lambda child_id: child_order(parent.person_id, child_id))

    # Validate chronology and allow the user to discard clearly erroneous links.
    for parent_id, child_id in list(selected_edges):
        parent = people_by_id[parent_id]
        child = people_by_id[child_id]
        if parent.year is not None and child.year is not None and child.year <= parent.year:
            cache_key = f"{parent_id}->{child_id}"
            cached = prompts.cache["edges"].get(cache_key)
            if isinstance(cached, bool):
                keep = cached
            elif cached == "__skip__":
                keep = None
            else:
                keep = prompts.yes_no_skip(
                    f"{child.stable_name} ({child.year}) is not later than VET "
                    f"{parent.stable_name} ({parent.year}). Keep this relationship?",
                    default=False,
                )
            prompts.cache["edges"][cache_key] = "__skip__" if keep is None else keep
            if keep is not True:
                if keep is None:
                    issues.skipped_relations.append(
                        f"Skipped chronologically ambiguous relationship: "
                        f"{parent.stable_name} -> {child.stable_name}."
                    )
                child.parent_id = None
                if child_id in parent.children_ids:
                    parent.children_ids.remove(child_id)
                selected_edges.remove((parent_id, child_id))
                edge_sources.pop((parent_id, child_id), None)
                # Preserve workbook relationship claims even when chronology removes the visual edge.

    cycle = find_cycle(people)
    while cycle:
        descriptions = []
        for parent_id, child_id in cycle:
            descriptions.append(
                f"{people_by_id[parent_id].stable_name} -> {people_by_id[child_id].stable_name}"
            )
        if not prompts.interactive and not prompts.skip_ambiguities:
            raise TreeDataError("Relationship cycle detected: " + " | ".join(descriptions))
        selected = prompts.choose(
            "A relationship cycle was found. Choose the incorrect edge to remove, or Skip to omit all edges in this ambiguous cycle:",
            descriptions,
            allow_skip=True,
        )
        if selected is None:
            for parent_id, child_id in list(cycle):
                parent = people_by_id[parent_id]
                child = people_by_id[child_id]
                if child.parent_id == parent_id:
                    child.parent_id = None
                if child_id in parent.children_ids:
                    parent.children_ids.remove(child_id)
                edge_sources.pop((parent_id, child_id), None)
                # Preserve workbook relationship claims even when chronology removes the visual edge.
            issues.skipped_relations.append(
                "Skipped all edges in relationship cycle: " + " | ".join(descriptions)
            )
        else:
            parent_id, child_id = cycle[selected]
            parent = people_by_id[parent_id]
            child = people_by_id[child_id]
            child.parent_id = None
            parent.children_ids.remove(child_id)
            edge_sources.pop((parent_id, child_id), None)
            # Preserve workbook claims; only this cycle's visual edge is removed.
        cycle = find_cycle(people)

    return people, edge_sources


def find_cycle(people: Sequence[Person]) -> list[tuple[str, str]]:
    people_by_id = {person.person_id: person for person in people}
    state: dict[str, int] = defaultdict(int)
    stack: list[str] = []

    def visit(person_id: str) -> list[tuple[str, str]] | None:
        state[person_id] = 1
        stack.append(person_id)
        for child_id in people_by_id[person_id].children_ids:
            if state[child_id] == 0:
                found = visit(child_id)
                if found:
                    return found
            elif state[child_id] == 1:
                start = stack.index(child_id)
                cycle_nodes = stack[start:] + [child_id]
                return list(zip(cycle_nodes, cycle_nodes[1:]))
        stack.pop()
        state[person_id] = 2
        return None

    for person in people:
        if state[person.person_id] == 0:
            found = visit(person.person_id)
            if found:
                return found
    return []


def visual_person_order(person: Person) -> tuple[int, str, str, int]:
    """Order cards/trees by RAT year, then family name, then given name."""

    return (
        person.year or 9999,
        person.family.casefold(),
        person.given.casefold(),
        person.row,
    )


def root_visual_order(person: Person) -> tuple[int, int]:
    """Match the legacy root ordering: RAT year first, then workbook order."""

    return (person.year or 9999, person.row)


def roots_for_people(people: Sequence[Person]) -> list[Person]:
    roots = [person for person in people if person.parent_id is None]
    return sorted(roots, key=root_visual_order)


def descendants(root: Person, people_by_id: dict[str, Person]) -> list[Person]:
    result: list[Person] = []
    stack = [root.person_id]
    while stack:
        person_id = stack.pop()
        person = people_by_id[person_id]
        result.append(person)
        stack.extend(reversed(person.children_ids))
    return result


def assign_x_positions(
    people: Sequence[Person],
    roots: Sequence[Person],
    *,
    card_width: int,
) -> tuple[dict[str, int], int]:
    """Place leaves left-to-right, then center every VET over direct RATs.

    This is the same bottom-up idea as the original section generators.  Leaf
    cards receive non-overlapping slots first.  Each parent receives the
    arithmetic mean of its direct children's left-edge x coordinates.  Separate
    roots receive FAMILY_GAP in addition to normal leaf spacing.
    """

    people_by_id = {person.person_id: person for person in people}
    included = set(people_by_id)
    x_positions: dict[str, int] = {}
    cursor = LEFT_MARGIN

    def place(person: Person) -> None:
        nonlocal cursor
        children = [
            people_by_id[child_id]
            for child_id in person.children_ids
            if child_id in included
        ]
        children.sort(key=visual_person_order)
        if not children:
            x_positions[person.person_id] = cursor
            cursor += card_width + LEAF_GAP
            return
        for child in children:
            place(child)
        x_positions[person.person_id] = round(
            sum(x_positions[child.person_id] for child in children) / len(children)
        )

    ordered_roots = sorted(roots, key=root_visual_order)

    def isolated(person: Person) -> bool:
        return person.parent_id is None and not any(
            child_id in included for child_id in person.children_ids
        )

    for index, root in enumerate(ordered_roots):
        place(root)
        if index + 1 < len(ordered_roots):
            next_root = ordered_roots[index + 1]
            # An isolated card has neither a VET nor a RAT, so do not give it
            # the same whitespace as two complete family trees. If either side
            # of the boundary is isolated, use the compact gap.
            cursor += ISOLATED_PERSON_GAP if isolated(root) or isolated(next_root) else FAMILY_GAP
    width = max(MIN_CANVAS_WIDTH, cursor + RIGHT_MARGIN)
    return x_positions, width


def assign_y_positions(
    people: Sequence[Person],
    *,
    card_height: int,
    header_height: int,
    global_min_year: int | None = None,
    global_max_year: int | None = None,
) -> tuple[dict[str, int], int | None, int | None, bool, int]:
    """Center each card vertically inside its 100-pixel RAT-year band."""

    years = [person.year for person in people if person.year is not None]
    min_year = global_min_year if global_min_year is not None else (min(years) if years else None)
    max_year = global_max_year if global_max_year is not None else (max(years) if years else None)
    has_unknown = any(person.year is None for person in people)

    normal_band_count = (max_year - min_year + 1) if min_year is not None and max_year is not None else 0
    total_bands = normal_band_count + (1 if has_unknown else 0)
    if total_bands == 0:
        total_bands = 1
        has_unknown = True

    y_positions: dict[str, int] = {}
    for person in people:
        if person.year is None or min_year is None:
            band_index = normal_band_count
        else:
            band_index = person.year - min_year
        band_top = header_height + band_index * YEAR_STRIP_HEIGHT
        y_positions[person.person_id] = (
            band_top + (YEAR_STRIP_HEIGHT - card_height) // 2 + TOP_CARD_PADDING
        )

    height = header_height + total_bands * YEAR_STRIP_HEIGHT
    return y_positions, min_year, max_year, has_unknown, height


def build_card_placements(
    scene_id: str,
    people: Sequence[Person],
    *,
    cards: dict[str, CardObject],
    x_positions: dict[str, int],
    y_positions: dict[str, int],
    is_full_tree: bool,
) -> dict[str, CardPlacement]:
    placements: dict[str, CardPlacement] = {}
    for person in people:
        card = cards[person.person_id]
        placements[person.person_id] = card.place(
            scene_id,
            x_positions[person.person_id],
            y_positions[person.person_id],
            is_full_tree=is_full_tree,
        )
    return placements


def connector_geometry(
    people: Sequence[Person],
    *,
    placements: dict[str, CardPlacement],
) -> list[FamilyConnector]:
    """Build the original VET→RAT elbow/bus connector geometry explicitly."""

    people_by_id = {person.person_id: person for person in people}
    included = set(people_by_id)
    connectors: list[FamilyConnector] = []

    for parent in people:
        children = [
            people_by_id[child_id]
            for child_id in parent.children_ids
            if child_id in included
        ]
        children.sort(key=visual_person_order)
        if not children:
            continue

        parent_place = placements[parent.person_id]
        child_places = [placements[child.person_id] for child in children]
        parent_point = parent_place.rat_connection
        child_points = [placement.vet_connection for placement in child_places]
        nearest_child_y = min(point.y for point in child_points)

        # With one RAT, the recursive placement algorithm centers the VET at the
        # same x coordinate, yielding the original straight vertical connector.
        if len(children) == 1 and child_points[0].x == parent_point.x:
            stem = ConnectorSegment(parent_point, child_points[0])
            connectors.append(
                FamilyConnector(
                    parent_id=parent.person_id,
                    child_ids=[children[0].person_id],
                    parent_stem=stem,
                    child_stems=[],
                    sibling_bus=None,
                )
            )
            continue

        junction_y = parent_point.y + max(10, (nearest_child_y - parent_point.y) // 2)
        parent_junction = PixelPoint(parent_point.x, junction_y)
        parent_stem = ConnectorSegment(parent_point, parent_junction)

        left_x = min(point.x for point in child_points)
        right_x = max(point.x for point in child_points)
        sibling_bus = ConnectorSegment(
            PixelPoint(left_x, junction_y),
            PixelPoint(right_x, junction_y),
        )
        child_stems = [
            ConnectorSegment(PixelPoint(point.x, junction_y), point)
            for point in child_points
        ]
        connectors.append(
            FamilyConnector(
                parent_id=parent.person_id,
                child_ids=[child.person_id for child in children],
                parent_stem=parent_stem,
                child_stems=child_stems,
                sibling_bus=sibling_bus,
            )
        )
    return connectors


def make_scene(
    scene_id: str,
    title: str,
    people: list[Person],
    roots: list[Person],
    *,
    cards: dict[str, CardObject],
    card_width: int,
    card_height: int,
    global_years: tuple[int | None, int | None] | None = None,
    is_full_tree: bool = False,
) -> Scene:
    x_positions, width = assign_x_positions(people, roots, card_width=card_width)
    header_height = FULL_TREE_HEADER_HEIGHT if is_full_tree else FAMILY_HEADER_HEIGHT
    y_positions, min_year, max_year, unknown_year, height = assign_y_positions(
        people,
        card_height=card_height,
        header_height=header_height,
        global_min_year=global_years[0] if global_years else None,
        global_max_year=global_years[1] if global_years else None,
    )
    placements = build_card_placements(
        scene_id,
        people,
        cards=cards,
        x_positions=x_positions,
        y_positions=y_positions,
        is_full_tree=is_full_tree,
    )
    connectors = connector_geometry(people, placements=placements)
    return Scene(
        scene_id=scene_id,
        title=title,
        header_height=header_height,
        people=people,
        roots=roots,
        width=width,
        height=height,
        min_year=min_year,
        max_year=max_year,
        unknown_year=unknown_year,
        placements=placements,
        connectors=connectors,
    )


def template_fill_mask(template: Image.Image) -> tuple[Image.Image, tuple[int, int, int, int]]:
    rgba = template.convert("RGBA")
    opaque_colors = Counter(pixel for pixel in rgba.get_flattened_data() if pixel[3] > 0)
    if not opaque_colors:
        raise TreeDataError("The card template contains no visible pixels.")
    dominant = opaque_colors.most_common(1)[0][0]
    mask = Image.new("1", rgba.size, 0)
    mask_pixels = mask.load()
    source_pixels = rgba.load()
    xs: list[int] = []
    ys: list[int] = []
    for y in range(rgba.height):
        for x in range(rgba.width):
            pixel = source_pixels[x, y]
            distance = sum(abs(pixel[index] - dominant[index]) for index in range(3))
            if pixel[3] > 0 and distance <= 6:
                mask_pixels[x, y] = 1
                xs.append(x)
                ys.append(y)
    if not xs:
        raise TreeDataError("Could not identify the fill area of the card template.")
    return mask, (min(xs), min(ys), max(xs), max(ys))


def fit_card_font(
    draw: ImageDraw.ImageDraw,
    lines: Sequence[str],
    *,
    max_width: int,
    max_height: int,
) -> tuple[ImageFont.ImageFont, list[int], int]:
    for size in range(22, 9, -1):
        font = load_font(size)
        heights: list[int] = []
        widths: list[int] = []
        for line in lines:
            box = draw.textbbox((0, 0), line, font=font)
            widths.append(box[2] - box[0])
            heights.append(box[3] - box[1])
        total_height = sum(heights) + 6 * max(0, len(lines) - 1)
        if max(widths, default=0) <= max_width and total_height <= max_height:
            return font, heights, total_height
    font = load_font(10)
    heights = []
    for line in lines:
        box = draw.textbbox((0, 0), line, font=font)
        heights.append(box[3] - box[1])
    total_height = sum(heights) + 4 * max(0, len(lines) - 1)
    return font, heights, total_height


def name_lines(person: Person) -> list[str]:
    given = person.card_given_name or person.stable_name
    family = person.family
    if family:
        return [given, family]
    words = person.stable_name.split()
    if len(words) <= 1:
        return words or ["Unknown"]
    midpoint = math.ceil(len(words) / 2)
    return [" ".join(words[:midpoint]), " ".join(words[midpoint:])]


LEADERSHIP_ICON_SIZE = 15
LEADERSHIP_ICON_MARGIN = 5


def _icon_box(card: Image.Image, kind: str) -> tuple[int, int, int, int]:
    size = LEADERSHIP_ICON_SIZE
    margin = LEADERSHIP_ICON_MARGIN
    positions = {
        "section-leader": (margin, margin),
        # Guard Captain is the color guard analogue of Section Leader and
        # shares the upper-left leadership area as an inward flag badge.
        "guard-captain": (margin + size + 4, margin),
        "drum-major": (card.width - margin - size, margin),
        "rat-parent": (margin, card.height - margin - size),
        "informal-leadership": (card.width - margin - size, card.height - margin - size),
        # Fifth role shares the lower-right corner as an inward second badge.
        "other-leadership": (card.width - margin - size * 2 - 4, card.height - margin - size),
    }
    x, y = positions[kind]
    return x, y, x + size - 1, y + size - 1


def _draw_section_leader(draw: ImageDraw.ImageDraw, box: tuple[int, int, int, int]) -> None:
    x0, y0, x1, y1 = box
    cx = (x0 + x1) // 2
    for offset in (2, 6, 10):
        y = y0 + offset
        draw.line([(x0 + 2, y + 3), (cx, y), (x1 - 2, y + 3)], fill="black", width=2)


def _draw_guard_captain(draw: ImageDraw.ImageDraw, box: tuple[int, int, int, int]) -> None:
    """Draw a compact black color-guard flag badge."""
    x0, y0, x1, y1 = box
    pole_x = x0 + 4
    draw.line([(pole_x, y0 + 1), (pole_x, y1 - 1)], fill="black", width=2)
    draw.polygon(
        [(pole_x + 1, y0 + 2), (x1 - 1, y0 + 5), (pole_x + 1, y0 + 9)],
        fill="black",
    )
    draw.line([(x0 + 1, y1 - 1), (x0 + 8, y1 - 1)], fill="black", width=1)


def _draw_drum_major(draw: ImageDraw.ImageDraw, box: tuple[int, int, int, int]) -> None:
    x0, y0, x1, y1 = box
    draw.line([(x0 + 3, y1 - 2), (x1 - 3, y0 + 2)], fill="black", width=2)
    draw.ellipse([x1 - 5, y0, x1 - 1, y0 + 4], outline="black", width=1)
    draw.line([(x0 + 1, y0 + 4), (x0 + 5, y0 + 1)], fill="black", width=1)
    draw.line([(x0 + 1, y0 + 8), (x0 + 6, y0 + 5)], fill="black", width=1)


def _draw_informal_leadership(draw: ImageDraw.ImageDraw, box: tuple[int, int, int, int]) -> None:
    x0, y0, x1, y1 = box
    mid = (y0 + y1) // 2
    draw.polygon([(x0 + 1, mid - 3), (x0 + 8, y0 + 2), (x0 + 8, y1 - 2)], outline="black")
    draw.rectangle([x0 + 8, mid - 2, x0 + 10, mid + 2], fill="black")
    draw.line([(x0 + 4, mid + 3), (x0 + 6, y1 - 1)], fill="black", width=2)
    draw.line([(x0 + 11, mid - 4), (x1, mid - 6)], fill="black", width=1)
    draw.line([(x0 + 11, mid), (x1, mid)], fill="black", width=1)
    draw.line([(x0 + 11, mid + 4), (x1, mid + 6)], fill="black", width=1)


def _draw_other_leadership(draw: ImageDraw.ImageDraw, box: tuple[int, int, int, int]) -> None:
    x0, y0, x1, y1 = box
    cx = (x0 + x1) / 2
    cy = (y0 + y1) / 2
    outer = (x1 - x0) * 0.46
    inner = outer * 0.43
    points = []
    for index in range(10):
        angle = -math.pi / 2 + index * math.pi / 5
        radius = outer if index % 2 == 0 else inner
        points.append((cx + math.cos(angle) * radius, cy + math.sin(angle) * radius))
    draw.line(points + [points[0]], fill="black", width=1, joint="curve")


def _draw_rat_parent(draw: ImageDraw.ImageDraw, box: tuple[int, int, int, int]) -> None:
    x0, y0, x1, y1 = box
    cx = (x0 + x1) // 2
    top_y = y0 + 3
    lower_y = y1 - 3
    left_x = x0 + 3
    right_x = x1 - 3
    draw.ellipse([cx - 2, top_y - 2, cx + 2, top_y + 2], fill="black")
    draw.ellipse([left_x - 2, lower_y - 2, left_x + 2, lower_y + 2], fill="black")
    draw.ellipse([right_x - 2, lower_y - 2, right_x + 2, lower_y + 2], fill="black")
    branch_y = (top_y + lower_y) // 2
    draw.line([(cx, top_y + 2), (cx, branch_y)], fill="black", width=1)
    draw.line([(left_x, branch_y), (right_x, branch_y)], fill="black", width=1)
    draw.line([(left_x, branch_y), (left_x, lower_y - 2)], fill="black", width=1)
    draw.line([(right_x, branch_y), (right_x, lower_y - 2)], fill="black", width=1)


def draw_leadership_icons(card: Image.Image, person: Person) -> None:
    draw = ImageDraw.Draw(card)
    drawers = {
        "section-leader": _draw_section_leader,
        "guard-captain": _draw_guard_captain,
        "drum-major": _draw_drum_major,
        "rat-parent": _draw_rat_parent,
        "informal-leadership": _draw_informal_leadership,
        "other-leadership": _draw_other_leadership,
    }
    for kind in person.leadership_icons:
        drawer = drawers.get(kind)
        if drawer:
            drawer(draw, _icon_box(card, kind))


RAT_CAP_ICON_PATH = SCRIPT_DIR / "web_template" / "rat-cap-icon.png"
_RAT_CAP_ICON_CACHE: Image.Image | None = None


def _rat_cap_icon(size: tuple[int, int] = (28, 18)) -> Image.Image | None:
    global _RAT_CAP_ICON_CACHE
    if _RAT_CAP_ICON_CACHE is None:
        if not RAT_CAP_ICON_PATH.exists():
            return None
        _RAT_CAP_ICON_CACHE = Image.open(RAT_CAP_ICON_PATH).convert("RGBA")
    icon = _RAT_CAP_ICON_CACHE.copy()
    icon.thumbnail(size, Image.Resampling.LANCZOS)
    return icon


def _draw_band_club_icon(draw: ImageDraw.ImageDraw, card: Image.Image) -> None:
    """Minimal Band Club badge: a music note inside an organization ring."""
    size = 17
    x0 = card.width - size - 5
    y0 = card.height - size - 5
    x1 = x0 + size
    y1 = y0 + size
    draw.ellipse((x0, y0, x1, y1), outline="black", width=1)
    stem_x = x0 + 10
    draw.line((stem_x, y0 + 4, stem_x, y0 + 11), fill="black", width=2)
    draw.line((stem_x, y0 + 4, x0 + 14, y0 + 3), fill="black", width=2)
    draw.ellipse((x0 + 5, y0 + 10, x0 + 10, y0 + 14), fill="black")


def draw_card_status_icons(card: Image.Image, person: Person) -> None:
    """Draw only the v17 non-leadership status badges.

    Existing section/formal/informal leadership corner icons are intentionally
    disabled. The RAT cap supplied by the project owner marks a current RAT,
    and Band Club leadership uses a separate music-club badge.
    """
    if person.currently_rat:
        icon = _rat_cap_icon()
        if icon is not None:
            card.alpha_composite(icon, (card.width - icon.width - 5, 4))
    if person.band_club_leadership:
        _draw_band_club_icon(ImageDraw.Draw(card), card)

def create_card_geometry(person: Person, size: tuple[int, int]) -> CardObject:
    """Create only card geometry/section metadata for browser-only builds.

    The encrypted website renders names, fills, and status badges in the browser.
    When PNG/SVG/card files are all disabled, there is no reason to render a
    personalized Pillow card locally. A transparent template-sized geometry surface keeps
    the original placement/connector math unchanged without creating member art.
    """
    categories = person.instruments or ["unknown"]
    colors = [SECTION_COLORS.get(category, UNKNOWN_COLOR) for category in categories]
    return CardObject(
        person_id=person.person_id,
        image=Image.new("RGBA", size, (0, 0, 0, 0)),
        section_colors=colors,
    )


def create_person_card(
    person: Person,
    *,
    template: Image.Image,
    fill_mask: Image.Image,
    fill_bounds: tuple[int, int, int, int],
) -> CardObject:
    card = template.convert("RGBA").copy()
    pixels = card.load()
    mask_pixels = fill_mask.load()
    min_x, min_y, max_x, max_y = fill_bounds
    categories = person.instruments or ["unknown"]
    colors = [SECTION_COLORS.get(category, UNKNOWN_COLOR) for category in categories]
    rgb_colors = [tuple(int(color[index : index + 2], 16) for index in (1, 3, 5)) for color in colors]
    interior_width = max_x - min_x + 1

    for y in range(min_y, max_y + 1):
        for x in range(min_x, max_x + 1):
            if not mask_pixels[x, y]:
                continue
            segment = min(len(rgb_colors) - 1, ((x - min_x) * len(rgb_colors)) // interior_width)
            red, green, blue = rgb_colors[segment]
            pixels[x, y] = (red, green, blue, pixels[x, y][3])

    if len(categories) > 1:
        divider_rgb = tuple(int(DIVIDER_COLOR[index : index + 2], 16) for index in (1, 3, 5))
        for division in range(1, len(categories)):
            boundary = min_x + round(interior_width * division / len(categories))
            for x in range(boundary - 1, boundary + 1):
                if not (min_x <= x <= max_x):
                    continue
                for y in range(min_y, max_y + 1):
                    if mask_pixels[x, y]:
                        pixels[x, y] = (*divider_rgb, pixels[x, y][3])

    draw = ImageDraw.Draw(card)
    lines = name_lines(person)
    font, line_heights, total_height = fit_card_font(
        draw,
        lines,
        max_width=card.width - 18,
        max_height=card.height - 16,
    )
    current_y = (card.height - total_height) // 2 - 1
    gap = 6 if getattr(font, "size", 10) >= 14 else 4
    for line, line_height in zip(lines, line_heights):
        box = draw.textbbox((0, 0), line, font=font)
        width = box[2] - box[0]
        x = (card.width - width) // 2
        draw.text((x, current_y), line, font=font, fill="black")
        current_y += line_height + gap

    draw_card_status_icons(card, person)

    return CardObject(
        person_id=person.person_id,
        image=card,
        section_colors=colors,
    )


def year_band_labels(scene: Scene) -> list[tuple[str, int, str, str]]:
    labels: list[tuple[str, int, str, str]] = []
    band_index = 0
    if scene.min_year is not None and scene.max_year is not None:
        for year in range(scene.min_year, scene.max_year + 1):
            color = YEAR_STRIP_COLORS[band_index % len(YEAR_STRIP_COLORS)]
            text_color = "#003057" if color == "#FFFFFF" else "#FFFFFF"
            labels.append((str(year), scene.header_height + band_index * YEAR_STRIP_HEIGHT, color, text_color))
            band_index += 1
    if scene.unknown_year or not labels:
        color = YEAR_STRIP_COLORS[band_index % len(YEAR_STRIP_COLORS)]
        text_color = "#003057" if color == "#FFFFFF" else "#FFFFFF"
        labels.append(("Unknown", scene.header_height + band_index * YEAR_STRIP_HEIGHT, color, text_color))
    return labels


def title_font_size(title: str, width: int) -> int:
    approximate = int(width / max(1, len(title)) * 1.55)
    return max(34, min(120, approximate))


def estimate_png_memory_mb(scene: Scene, scale: float) -> float:
    width = max(1, round(scene.width * scale))
    height = max(1, round(scene.height * scale))
    return width * height * 3 / (1024 * 1024) * 1.35


def choose_png_scale(
    scene: Scene,
    requested_scale: float,
    *,
    max_pixels: int,
    prompts: PromptSession,
) -> float | None:
    pixels = scene.width * scene.height * requested_scale * requested_scale
    if pixels <= max_pixels:
        return requested_scale
    recommended = requested_scale * math.sqrt(max_pixels / pixels)
    recommended = max(0.1, min(requested_scale, math.floor(recommended * 100) / 100))
    message = (
        f"The PNG {scene.title!r} would be approximately "
        f"{scene.width * requested_scale:,.0f} × {scene.height * requested_scale:,.0f} "
        f"pixels and use about {estimate_png_memory_mb(scene, requested_scale):,.0f} MB while rendering."
    )
    if not prompts.interactive:
        prompts.issues.warn(f"{message} Automatically scaling to {recommended:.2f}.")
        return recommended
    print("\n" + message)
    selected = prompts.choose(
        "Choose how to render this PNG:",
        [
            f"Continue at full requested scale ({requested_scale:.2f})",
            f"Scale down to {recommended:.2f}",
            "Skip this PNG (SVG will still be generated)",
        ],
        default_index=1,
    )
    if selected == 0:
        return requested_scale
    if selected == 1:
        return recommended
    return None


def scaled_rect(rect: tuple[int, int, int, int], scale: float) -> tuple[int, int, int, int]:
    x, y, width, height = rect
    return (
        round(x * scale),
        round(y * scale),
        max(1, round(width * scale)),
        max(1, round(height * scale)),
    )


def render_png(
    scene: Scene,
    output_path: Path,
    *,
    cards: dict[str, CardObject],
    scale: float,
) -> None:
    width = max(1, round(scene.width * scale))
    height = max(1, round(scene.height * scale))
    image = Image.new("RGB", (width, height), BACKGROUND_HEADER_COLOR)
    draw = ImageDraw.Draw(image)

    for label, y, color, text_color in year_band_labels(scene):
        sy = round(y * scale)
        strip_height = max(1, round(YEAR_STRIP_HEIGHT * scale))
        draw.rectangle((0, sy, width, sy + strip_height), fill=color)
        font = load_font(max(12, round(44 * scale)), bold=True)
        box = draw.textbbox((0, 0), label, font=font)
        text_y = sy + (strip_height - (box[3] - box[1])) // 2 - round(3 * scale)
        draw.text((round(25 * scale), text_y), label, font=font, fill=text_color)

    for rect in scene.connector_outline_rects:
        x, y, rect_width, rect_height = scaled_rect(rect, scale)
        draw.rectangle((x, y, x + rect_width - 1, y + rect_height - 1), fill=CONNECTOR_OUTLINE_COLOR)
    for rect in scene.connector_rects:
        x, y, rect_width, rect_height = scaled_rect(rect, scale)
        draw.rectangle((x, y, x + rect_width - 1, y + rect_height - 1), fill=CONNECTOR_COLOR)

    resized_cards: dict[str, CardObject] = {}
    for person in scene.people:
        card_object = cards[person.person_id]
        card_image = card_object.image
        placement = scene.placements[person.person_id]
        if scale != 1.0:
            key = f"{person.person_id}|{scale:.4f}"
            resized = resized_cards.get(key)
            if resized is None:
                resized = card_image.resize(
                    (
                        max(1, round(card_image.width * scale)),
                        max(1, round(card_image.height * scale)),
                    ),
                    Image.Resampling.LANCZOS,
                )
                resized_cards[key] = resized
            card_to_paste = resized
        else:
            card_to_paste = card_image
        image.paste(
            card_to_paste,
            (round(placement.x * scale), round(placement.y * scale)),
            card_to_paste,
        )

    if scene.header_height > 0 and scene.title:
        title_size = max(20, round(title_font_size(scene.title, scene.width) * scale))
        title_font = load_font(title_size, bold=True)
        box = draw.textbbox((0, 0), scene.title, font=title_font)
        text_width = box[2] - box[0]
        text_height = box[3] - box[1]
        draw.text(
            ((width - text_width) // 2, (round(scene.header_height * scale) - text_height) // 2),
            scene.title,
            font=title_font,
            fill="white",
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_path, optimize=True)
    image.close()


def card_data_uri(card: CardObject) -> str:
    from io import BytesIO

    buffer = BytesIO()
    card.image.save(buffer, format="PNG", optimize=True)
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def render_svg(
    scene: Scene,
    output_path: Path,
    *,
    cards: dict[str, CardObject],
) -> None:
    lines: list[str] = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        (
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{scene.width}" '
            f'height="{scene.height}" viewBox="0 0 {scene.width} {scene.height}">'
        ),
        f'<rect x="0" y="0" width="{scene.width}" height="{scene.header_height}" fill="{BACKGROUND_HEADER_COLOR}"/>',
    ]

    for label, y, color, text_color in year_band_labels(scene):
        lines.append(
            f'<rect x="0" y="{y}" width="{scene.width}" height="{YEAR_STRIP_HEIGHT}" fill="{color}"/>'
        )
        lines.append(
            f'<text x="25" y="{y + YEAR_STRIP_HEIGHT / 2}" fill="{text_color}" '
            'font-family="Calibri, Arial, sans-serif" font-size="44" font-weight="700" '
            f'dominant-baseline="middle">{html.escape(label)}</text>'
        )

    if scene.connector_outline_rects and CONNECTOR_OUTLINE_COLOR:
        lines.append(f'<g fill="{CONNECTOR_OUTLINE_COLOR}" shape-rendering="crispEdges">')
        for x, y, width, height in scene.connector_outline_rects:
            lines.append(f'<rect x="{x}" y="{y}" width="{width}" height="{height}"/>')
        lines.append("</g>")
    lines.append(f'<g fill="{CONNECTOR_COLOR}" shape-rendering="crispEdges">')
    for x, y, width, height in scene.connector_rects:
        lines.append(f'<rect x="{x}" y="{y}" width="{width}" height="{height}"/>')
    lines.append("</g>")

    uri_cache: dict[str, str] = {}
    for person in scene.people:
        card = cards[person.person_id]
        uri = uri_cache.setdefault(person.person_id, card_data_uri(card))
        placement = scene.placements[person.person_id]
        lines.append(
            f'<image x="{placement.x}" y="{placement.y}" width="{card.width}" height="{card.height}" '
            f'href="{uri}"/>'
        )

    if scene.header_height > 0 and scene.title:
        title_size = title_font_size(scene.title, scene.width)
        lines.append(
            f'<text x="{scene.width / 2}" y="{scene.header_height / 2}" fill="#FFFFFF" '
            'font-family="Calibri, Arial, sans-serif" '
            f'font-size="{title_size}" font-weight="700" text-anchor="middle" '
            f'dominant-baseline="middle">{html.escape(scene.title)}</text>'
        )
    lines.append("</svg>")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")


def render_legend(output_dir: Path, *, png: bool, svg: bool) -> list[Path]:
    items = list(SECTION_COLORS.items())
    columns = 2
    row_height = 42
    column_width = 340
    margin = 30
    title_height = 70
    rows = math.ceil(len(items) / columns)
    width = columns * column_width + margin * 2
    height = title_height + rows * row_height + margin
    generated: list[Path] = []

    if png:
        image = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(image)
        title_font = load_font(30, bold=True)
        item_font = load_font(21)
        draw.text((margin, 20), "YJMB Full-Band Tree Color Legend", font=title_font, fill="black")
        for index, (label, color) in enumerate(items):
            column = index // rows
            row = index % rows
            x = margin + column * column_width
            y = title_height + row * row_height
            draw.rectangle((x, y + 4, x + 46, y + 32), fill=color, outline="#222222", width=2)
            draw.text((x + 60, y + 5), label, font=item_font, fill="black")
        path = output_dir / "YJMB_Full_Band_Tree_Color_Legend.png"
        image.save(path, optimize=True)
        image.close()
        generated.append(path)

    if svg:
        lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
            f'<rect width="{width}" height="{height}" fill="#FFFFFF"/>',
            f'<text x="{margin}" y="42" font-family="Calibri, Arial, sans-serif" font-size="30" font-weight="700">YJMB Full-Band Tree Color Legend</text>',
        ]
        for index, (label, color) in enumerate(items):
            column = index // rows
            row = index % rows
            x = margin + column * column_width
            y = title_height + row * row_height
            lines.append(
                f'<rect x="{x}" y="{y + 4}" width="46" height="28" fill="{color}" stroke="#222222" stroke-width="2"/>'
            )
            lines.append(
                f'<text x="{x + 60}" y="{y + 26}" font-family="Calibri, Arial, sans-serif" font-size="21">{html.escape(label)}</text>'
            )
        lines.append("</svg>")
        path = output_dir / "YJMB_Full_Band_Tree_Color_Legend.svg"
        path.write_text("\n".join(lines), encoding="utf-8")
        generated.append(path)
    return generated


def write_report(
    path: Path,
    *,
    workbook_path: Path,
    sheet_name: str,
    people: Sequence[Person],
    roots: Sequence[Person],
    edge_sources: dict[tuple[str, str], list[EdgeSource]],
    issues: IssueLog,
    generated: Sequence[Path],
) -> None:
    multi_instrument = [person for person in people if len(person.instruments) > 1]
    unknown = [person for person in people if "unknown" in person.instruments]
    lines = [
        "YJMB FULL-BAND TREE GENERATION REPORT",
        "=" * 43,
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"Workbook: {workbook_path}",
        f"Worksheet: {sheet_name}",
        f"People: {len(people)}",
        f"Family roots / disconnected trees: {len(roots)}",
        f"Selected parent-child edges: {sum(1 for person in people if person.parent_id)}",
        f"Relationship pairs found before conflict resolution: {len(edge_sources)}",
        f"People with multiple section colors: {len(multi_instrument)}",
        f"People colored as unknown: {len(unknown)}",
        "",
        "ROOTS",
        "-----",
    ]
    lines.extend(
        f"{index:02d}. {root.stable_name} ({root.year_label})"
        for index, root in enumerate(roots, start=1)
    )
    if multi_instrument:
        lines.extend(("", "MULTI-SECTION CARDS", "-------------------"))
        lines.extend(
            f"- {person.stable_name}: {', '.join(person.instruments)}"
            for person in multi_instrument
        )
    if unknown:
        lines.extend(("", "UNKNOWN SECTION CARDS", "---------------------"))
        lines.extend(
            f"- {person.stable_name}: {person.instrument_raw or '[blank]'}"
            for person in unknown
        )
    if issues.warnings:
        lines.extend(("", "WARNINGS", "--------", *[f"- {item}" for item in issues.warnings]))
    if issues.decisions:
        lines.extend(("", "INTERACTIVE DECISIONS", "---------------------", *[f"- {item}" for item in issues.decisions]))
    if issues.skipped_people:
        lines.extend(("", "SKIPPED PEOPLE", "--------------", *[f"- {item}" for item in issues.skipped_people]))
    if issues.skipped_ambiguities:
        lines.extend(("", "SKIPPED AMBIGUITIES", "-------------------", *[f"- {item}" for item in issues.skipped_ambiguities]))
    if issues.skipped_relations:
        lines.extend(("", "SKIPPED RELATIONSHIPS", "---------------------", *[f"- {item}" for item in issues.skipped_relations]))
    lines.extend(("", "GENERATED FILES", "---------------"))
    lines.extend(f"- {item}" for item in generated)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")



def connector_to_dict(connector: FamilyConnector) -> dict[str, object]:
    return {
        "parentId": connector.parent_id,
        "childIds": connector.child_ids,
        "parentStem": connector.parent_stem.as_dict(),
        "siblingBus": connector.sibling_bus.as_dict() if connector.sibling_bus else None,
        "childStems": [segment.as_dict() for segment in connector.child_stems],
    }


def scene_to_web_data(
    scene: Scene,
    *,
    cards: dict[str, CardObject],
    edge_sources: dict[tuple[str, str], list[EdgeSource]],
) -> dict[str, object]:
    """Serialize one full-band scene for both static rendering and interaction.

    Schema version 5 adds submitter-side relationship claims and reciprocity
    metadata so the browser can distinguish a person's own VET/RAT claim from
    a relationship that has also been confirmed on the other person's row.
    The browser can therefore filter whole trees, focus the component containing
    a searched person, and populate correction forms without re-reading Excel.
    """

    people_by_id = {person.person_id: person for person in scene.people}
    included = set(people_by_id)
    root_cache: dict[str, str] = {}

    def root_id_for(person: Person) -> str:
        cached = root_cache.get(person.person_id)
        if cached:
            return cached
        path: list[str] = []
        cursor = person
        seen: set[str] = set()
        while cursor.parent_id and cursor.parent_id in included and cursor.person_id not in seen:
            seen.add(cursor.person_id)
            path.append(cursor.person_id)
            cursor = people_by_id[cursor.parent_id]
        root_id = cursor.person_id
        root_cache[root_id] = root_id
        for person_id in path:
            root_cache[person_id] = root_id
        return root_id

    def source_field_value(person: Person, label: str) -> str:
        wanted = normalized_header(label)
        for field_item in person.source_fields:
            if normalized_header(field_item.get("label", "")) == wanted:
                return normalize_spaces(field_item.get("value", ""))
        return ""

    def source_types_for_edge(parent_id: str, child_id: str) -> set[int]:
        return {source.order[0] for source in edge_sources.get((parent_id, child_id), [])}

    def edge_reciprocated(parent_id: str, child_id: str) -> bool:
        # order[0] == 0 means the child listed its VET; order[0] == 1 means
        # the parent listed the child as a RAT.  Both directions means the
        # relationship is reciprocated in the two profiles.
        return {0, 1}.issubset(source_types_for_edge(parent_id, child_id))

    # Relationship claims are data, while selected_edges are layout.  Resolve a
    # claim independently so a safe visual-edge omission (conflicting parent,
    # chronology, cycle) does not make an otherwise identifiable person vanish
    # from the sidebar/Admin tools.
    claim_strict_index, claim_loose_index = build_alias_indexes(scene.people)

    def unique_claim_target(raw: str) -> Person | None:
        parsed = parse_relation(raw)
        if not parsed or not parsed.name:
            return None
        candidates = candidate_people_for_reference(
            parsed, scene.people, claim_strict_index, claim_loose_index
        )
        return candidates[0] if len(candidates) == 1 else None

    def own_vet_claim(person: Person) -> dict[str, object] | None:
        if not person.vet_raw:
            return None
        parsed = parse_relation(person.vet_raw)
        resolved_parent_id: str | None = None
        for (parent_id, child_id), sources in edge_sources.items():
            if child_id != person.person_id:
                continue
            if any(source.order[0] == 0 and source.order[1] == person.row for source in sources):
                resolved_parent_id = parent_id
                break
        if resolved_parent_id is None:
            fallback = unique_claim_target(person.vet_raw)
            resolved_parent_id = fallback.person_id if fallback else None
        related = people_by_id.get(resolved_parent_id) if resolved_parent_id else None
        reciprocal = bool(resolved_parent_id and edge_reciprocated(resolved_parent_id, person.person_id))
        stored_status = source_field_value(person, "VET Relationship Status")
        status = (stored_status if stored_status.casefold().startswith("reciprocated") else "Reciprocated — validated on both profiles") if reciprocal else (stored_status or "Unreciprocated — pending validation")
        return {
            "role": "VET",
            "raw": person.vet_raw,
            "id": resolved_parent_id,
            "name": related.stable_name if related else (parsed.name if parsed else person.vet_raw),
            "ratYear": related.year if related else ((parsed.year_candidates[0] if parsed and parsed.year_candidates else None)),
            "reciprocated": reciprocal,
            "status": status,
            "liveTreeEdge": person.parent_id == resolved_parent_id if resolved_parent_id else False,
            "tooltip": "" if reciprocal else "This relationship has not been reciprocated in this user's profile submission.",
        }

    def own_rat_claims(person: Person) -> list[dict[str, object]]:
        claims: list[dict[str, object]] = []
        for rat_index, (column_name, raw_rat) in enumerate(person.rat_raws, start=1):
            parsed = parse_relation(raw_rat)
            resolved_child_id: str | None = None
            for (parent_id, child_id), sources in edge_sources.items():
                if parent_id != person.person_id:
                    continue
                if any(
                    source.order[0] == 1
                    and source.order[1] == person.row
                    and source.order[2] == rat_index
                    for source in sources
                ):
                    resolved_child_id = child_id
                    break
            if resolved_child_id is None:
                fallback = unique_claim_target(raw_rat)
                resolved_child_id = fallback.person_id if fallback else None
            related = people_by_id.get(resolved_child_id) if resolved_child_id else None
            reciprocal = bool(resolved_child_id and edge_reciprocated(person.person_id, resolved_child_id))
            stored_status = source_field_value(person, f"RAT {rat_index} Relationship Status")
            status = (stored_status if stored_status.casefold().startswith("reciprocated") else "Reciprocated — validated on both profiles") if reciprocal else (stored_status or "Unreciprocated — pending validation")
            claims.append({
                "role": "RAT",
                "raw": raw_rat,
                "slot": column_name,
                "id": resolved_child_id,
                "name": related.stable_name if related else (parsed.name if parsed else raw_rat),
                "ratYear": related.year if related else ((parsed.year_candidates[0] if parsed and parsed.year_candidates else None)),
                "reciprocated": reciprocal,
                "status": status,
                "liveTreeEdge": bool(resolved_child_id and resolved_child_id in person.children_ids),
                "tooltip": "" if reciprocal else "This relationship has not been reciprocated in this user's profile submission.",
            })
        return claims

    members_by_root: dict[str, list[str]] = defaultdict(list)
    for person in scene.people:
        members_by_root[root_id_for(person)].append(person.person_id)

    tree_data: list[dict[str, object]] = []
    for root in scene.roots:
        member_ids = members_by_root.get(root.person_id, [root.person_id])
        member_people = [people_by_id[person_id] for person_id in member_ids]
        member_sections = {
            section
            for person in member_people
            for section in person.instruments
        }
        ordered_sections = [
            section for section in SECTION_COLORS
            if section in member_sections
        ]
        placements = [scene.placements[person_id] for person_id in member_ids]
        min_x = min(placement.x for placement in placements)
        min_y = min(placement.y for placement in placements)
        max_x = max(placement.x + placement.width for placement in placements)
        max_y = max(placement.y + placement.height for placement in placements)
        years = [person.year for person in member_people if person.year is not None]
        tree_data.append(
            {
                "rootId": root.person_id,
                "rootName": root.stable_name,
                "rootRatYear": root.year,
                "rootRatYearLabel": root.year_label,
                "memberIds": member_ids,
                "memberCount": len(member_ids),
                "sections": ordered_sections,
                "bounds": {
                    "minX": min_x,
                    "minY": min_y,
                    "maxX": max_x,
                    "maxY": max_y,
                },
                "ratYearRange": {
                    "min": min(years) if years else None,
                    "max": max(years) if years else None,
                },
            }
        )

    people_data: list[dict[str, object]] = []
    for person in scene.people:
        card = cards[person.person_id]
        placement = scene.placements[person.person_id]
        people_data.append(
            {
                "id": person.person_id,
                "rootId": root_id_for(person),
                "name": person.stable_name,
                "displayName": person.card_display_name,
                "currentName": person.current_name,
                "givenPreferredName": person.given,
                "nickname": person.nickname,
                "treeDisplayNamePreference": person.tree_name_preference or "Given/Preferred Name",
                "familyMaidenName": person.family,
                "marriedName": person.married,
                "personalNickname": person.nickname,
                "sectionNicknames": person.section_nicknames,
                "favoriteTechBandMemory": person.favorite_tech_band_memory,
                "marchingBandLeadershipHistory": person.marching_band_leadership_history,
                "bandClubLeadershipHistory": person.band_club_leadership_history,
                "currentlyRat": person.currently_rat,
                "bandClubLeadership": person.band_club_leadership,
                "ratYear": person.year,
                "ratYearLabel": person.year_label,
                "instrumentRaw": person.instrument_raw,
                "instruments": person.instruments,
                "uncategorizedInstrumentText": uncategorized_instrument_text(person.instrument_raw),
                "leadershipIcons": person.leadership_icons,
                "leadershipIconsEnabled": False,
                "parentId": person.parent_id,
                "childrenIds": person.children_ids,
                "sourceFields": person.source_fields,
                "relationshipClaims": {
                    "vet": own_vet_claim(person),
                    "rats": own_rat_claims(person),
                },
                "card": {
                    "x": placement.x,
                    "y": placement.y,
                    "width": card.width,
                    "height": card.height,
                    # v17 cards are rendered from structured data in the browser.
                    # No pre-rendered per-person image is required in the encrypted
                    # website payload, so site-only updates never depend on local
                    # card PNG generation.
                    "sectionColors": card.section_colors,
                    "localVetConnection": card.local_vet_connection.as_dict(),
                    "globalVetConnection": (
                        card.global_vet_connection.as_dict()
                        if card.global_vet_connection
                        else placement.vet_connection.as_dict()
                    ),
                    "localRatConnection": card.local_rat_connection.as_dict(),
                    "globalRatConnection": (
                        card.global_rat_connection.as_dict()
                        if card.global_rat_connection
                        else placement.rat_connection.as_dict()
                    ),
                },
            }
        )

    return {
        "schemaVersion": 8,
        "sceneId": scene.scene_id,
        "title": scene.title,
        "width": scene.width,
        "height": scene.height,
        "headerHeight": scene.header_height,
        "yearStripHeight": YEAR_STRIP_HEIGHT,
        "connectorColor": CONNECTOR_COLOR,
        "connectorWidth": CONNECTOR_WIDTH,
        "connectorOutlineColor": None,
        "connectorOutlineWidth": 0,
        "sectionColors": SECTION_COLORS,
        "yearBands": [
            {"label": label, "y": y, "color": color, "textColor": text_color}
            for label, y, color, text_color in year_band_labels(scene)
        ],
        "roots": [root.person_id for root in scene.roots],
        "trees": tree_data,
        "people": people_data,
        "connectors": [connector_to_dict(connector) for connector in scene.connectors],
    }


def load_tree_data_key(path: Path | None = None) -> bytes:
    """Load the 32-byte site data key from an environment variable or local secret file.

    The preferred CI path is ``TREE_DATA_KEY_B64``.  Local builds can keep the
    same value in gitignored ``access_secrets.json`` under ``treeDataKey``.
    No knowledge-answer strings or answer-derived verifiers are written to docs/.
    """
    raw_b64 = normalize_spaces(os.environ.get("TREE_DATA_KEY_B64", ""))
    secrets_path = (path or (SCRIPT_DIR / ACCESS_SECRETS_FILENAME)).expanduser().resolve()
    if not raw_b64 and secrets_path.exists():
        try:
            raw = json.loads(secrets_path.read_text(encoding="utf-8"))
            raw_b64 = normalize_spaces(raw.get("treeDataKey", ""))
        except Exception as exc:
            raise TreeDataError(f"Could not parse {secrets_path.name}: {exc}") from exc
    if not raw_b64:
        raise TreeDataError(
            "Missing TREE_DATA_KEY_B64/access_secrets.json treeDataKey. "
            "Run initialize_security.py once, then configure the same key as a Cloudflare Worker and GitHub Actions secret."
        )
    try:
        key = base64.b64decode(raw_b64, validate=True)
    except Exception as exc:
        raise TreeDataError("treeDataKey/TREE_DATA_KEY_B64 is not valid Base64.") from exc
    if len(key) != 32:
        raise TreeDataError("treeDataKey/TREE_DATA_KEY_B64 must decode to exactly 32 bytes (AES-256).")
    return key


def encrypt_web_payload(payload: dict[str, object], *, data_key: bytes) -> dict[str, object]:
    """Encrypt the complete browser payload with a server-delivered AES-256 key."""
    data_iv = os.urandom(12)
    plaintext = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    ciphertext = AESGCM(data_key).encrypt(data_iv, plaintext, None)
    return {
        "format": "yjmb-tree-encrypted-v3",
        "cipher": "AES-256-GCM",
        "keyDelivery": "authenticated-server-session",
        "dataIv": base64.b64encode(data_iv).decode("ascii"),
        "ciphertext": base64.b64encode(ciphertext).decode("ascii"),
    }


def export_github_pages_site(
    scene: Scene,
    site_dir: Path,
    *,
    cards: dict[str, CardObject],
    edge_sources: dict[tuple[str, str], list[EdgeSource]],
) -> list[Path]:
    """Write an encrypted, static GitHub Pages viewer for the full-band scene."""

    site_dir = site_dir.expanduser().resolve()
    data_dir = site_dir / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    # Remove privacy-sensitive artifacts produced by older versions before
    # writing the new public-site bundle.  This prevents a stale plaintext JSON
    # or name-bearing card PNG from accidentally being committed later.
    stale_plaintext = data_dir / "tree_data.json"
    if stale_plaintext.exists():
        stale_plaintext.unlink()
    stale_cards = site_dir / "assets" / "cards"
    if stale_cards.exists():
        shutil.rmtree(stale_cards)

    generated: list[Path] = []
    template_dir = SCRIPT_DIR / "web_template"
    template_names = (
        "index.html",
        "gate-2.html",
        "gate-3.html",
        "loading.html",
        "tree.html",
        "correction.html",
        "add-yourself.html",
        "admin.html",
        "styles.css",
        "gate.css",
        "gate.js",
        "secure-data.js",
        "developer-export.js",
        "app.js",
        "correction.js",
        "add-yourself.js",
        "admin.js",
        "admin-mail.js",
        "rat-cap-icon.png",
        "site_config.json",
    )
    missing_templates = [template_dir / name for name in template_names if not (template_dir / name).exists()]
    if missing_templates:
        missing = ", ".join(path.name for path in missing_templates)
        raise TreeDataError(
            f"GitHub Pages template files are missing from {template_dir}: {missing}. "
            "Copy the complete web_template folder beside fullBandTreeGenerator.py."
        )

    for name in template_names:
        source = template_dir / name
        destination = site_dir / name
        shutil.copy2(source, destination)
        generated.append(destination)

    nojekyll_path = site_dir / ".nojekyll"
    nojekyll_path.write_text("", encoding="utf-8")
    generated.append(nojekyll_path)

    # Remove the obsolete v12 public answer fingerprint file if present.
    access_config_path = site_dir / "access_config.js"
    if access_config_path.exists():
        access_config_path.unlink()

    payload = scene_to_web_data(scene, cards=cards, edge_sources=edge_sources)
    encrypted = encrypt_web_payload(payload, data_key=load_tree_data_key())
    encrypted_path = data_dir / ENCRYPTED_DATA_FILENAME
    encrypted_path.write_text(
        json.dumps(encrypted, separators=(",", ":"), ensure_ascii=False),
        encoding="utf-8",
    )
    generated.append(encrypted_path)

    return generated

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a color-coded full-band YJMB family tree from YJMB Trees.xlsx."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help="Input workbook. Defaults to YJMB Trees.xlsx beside this script, then sibling trumpettree.",
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Master worksheet name.")
    parser.add_argument(
        "--template",
        type=Path,
        default=None,
        help="Card template PNG. Defaults to blank_name_card.png beside this script or in sibling trumpettree.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=SCRIPT_DIR / "trees",
        help="Tree output directory. Default: .\\trees",
    )
    parser.add_argument(
        "--cards-dir",
        type=Path,
        default=SCRIPT_DIR / "cards" / "full_band",
        help="Generated card directory. Default: .\\cards\\full_band",
    )
    parser.add_argument(
        "--cache",
        type=Path,
        default=SCRIPT_DIR / ".full_band_tree_resolutions.json",
        help="Interactive resolution cache.",
    )
    parser.add_argument(
        "--output-mode",
        choices=("all", "giant", "families"),
        default="all",
        help="Generate everything, only the giant tree, or only per-family trees.",
    )
    parser.add_argument("--skip-png", action="store_true", help="Do not generate PNG files.")
    parser.add_argument("--skip-svg", action="store_true", help="Do not generate SVG files.")
    parser.add_argument("--no-legend", action="store_true", help="Do not generate the color legend.")
    parser.add_argument("--no-save-cards", action="store_true", help="Do not save individual card PNGs.")
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="Do not prompt. Existing non-interactive fallback behavior is used.",
    )
    parser.add_argument(
        "--skip-ambiguities",
        action="store_true",
        help=(
            "Automatically skip unresolved/ambiguous layout decisions instead of prompting. "
            "Workbook people and raw source fields are still preserved in the protected site/admin data; "
            "ambiguous visual edges may be omitted and unrecognized section fragments ignored."
        ),
    )
    parser.add_argument(
        "--png-scale",
        type=float,
        default=1.0,
        help="PNG rendering scale. SVG always retains full coordinates. Default: 1.0",
    )
    parser.add_argument(
        "--max-png-pixels",
        type=int,
        default=150_000_000,
        help="Prompt or auto-scale when a PNG exceeds this pixel count. Default: 150000000",
    )
    parser.add_argument(
        "--family-year-mode",
        choices=("compact", "global"),
        default="compact",
        help="Per-family year bands: compact to that family or aligned to the full-band range.",
    )
    parser.add_argument(
        "--family-limit",
        type=int,
        default=None,
        help="Generate only the first N family trees. Mainly useful for testing.",
    )
    parser.add_argument(
        "--site-dir",
        type=Path,
        default=SCRIPT_DIR / "docs",
        help="GitHub Pages output directory. Default: .\\docs",
    )
    parser.add_argument(
        "--no-site",
        action="store_true",
        help="Do not generate the static HTML/CSS/JS GitHub Pages viewer.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.skip_png and args.skip_svg and args.no_site:
        raise TreeDataError(
            "PNG, SVG, and GitHub Pages output are all disabled; there is nothing to generate."
        )
    if args.png_scale <= 0:
        raise TreeDataError("--png-scale must be greater than zero.")

    workbook_path = find_existing_file(MASTER_WORKBOOK, args.workbook)
    template_path = find_existing_file(DEFAULT_TEMPLATE, args.template)
    output_dir = args.output_dir.expanduser().resolve()
    cards_dir = args.cards_dir.expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    cards_dir.mkdir(parents=True, exist_ok=True)

    issues = IssueLog()
    prompts = PromptSession(
        args.cache.expanduser().resolve(),
        interactive=not args.non_interactive,
        issues=issues,
        skip_ambiguities=args.skip_ambiguities,
    )

    print(f"Workbook: {workbook_path}")
    print(f"Worksheet: {args.sheet}")
    print(f"Card template: {template_path}")
    print(f"Output directory: {output_dir}")

    workbook = None
    try:
        people, workbook = load_people(
            workbook_path,
            sheet_name=args.sheet,
            prompts=prompts,
            issues=issues,
        )
        people, edge_sources = build_relationships(
            people,
            prompts=prompts,
            issues=issues,
        )
    finally:
        if workbook is not None:
            workbook.close()
        prompts.save()

    people_by_id = {person.person_id: person for person in people}
    roots = roots_for_people(people)
    if not roots:
        raise TreeDataError("No tree roots were found after resolving relationships.")

    # Browser-only site builds do not render personalized card images locally.
    # Pillow cards are created only when a local PNG/SVG/card artifact is requested.
    render_local_card_art = not (args.no_save_cards and args.skip_png and args.skip_svg)
    template = None
    fill_mask = None
    fill_bounds = None
    geometry_size: tuple[int, int] | None = None
    if render_local_card_art:
        template = Image.open(template_path).convert("RGBA")
        fill_mask, fill_bounds = template_fill_mask(template)
    else:
        # Read only the template dimensions so historic/custom geometry stays intact;
        # no personalized card image is painted in the site-build path.
        with Image.open(template_path) as template_probe:
            geometry_size = template_probe.size

    card_size = template.size if template is not None else geometry_size
    assert card_size is not None
    card_width, card_height = card_size

    cards: dict[str, CardObject] = {}
    for index, person in enumerate(people, start=1):
        if render_local_card_art:
            assert template is not None and fill_mask is not None and fill_bounds is not None
            card = create_person_card(
                person,
                template=template,
                fill_mask=fill_mask,
                fill_bounds=fill_bounds,
            )
        else:
            assert geometry_size is not None
            card = create_card_geometry(person, geometry_size)
        cards[person.person_id] = card
        if not args.no_save_cards:
            # Include the row-backed ID so same-name/same-year people never
            # overwrite each other's generated card PNGs.  The ID is only in the
            # filename; the visible card text is unchanged.
            filename = safe_filename(
                f"{person.stable_name} [{person.year_label}] [{person.person_id}]"
            ) + ".png"
            card.image.save(cards_dir / filename, optimize=True)
        if index % 50 == 0 or index == len(people):
            print(f"Prepared {index}/{len(people)} cards.")

    generated: list[Path] = []
    all_years = [person.year for person in people if person.year is not None]
    global_years = (
        (min(all_years), max(all_years)) if all_years else (None, None)
    )
    latest_year_label = str(global_years[1]) if global_years[1] is not None else "Unknown"

    if args.output_mode in {"all", "giant"}:
        scene = make_scene(
            "full-band",
            "The YJMB Full Band Family Tree",
            people,
            roots,
            cards=cards,
            card_width=card_width,
            card_height=card_height,
            is_full_tree=True,
        )
        stem = f"YJMB_Full_Band_Family_Tree_{latest_year_label}"
        if not args.skip_svg:
            svg_path = output_dir / f"{stem}.svg"
            print(f"Writing {svg_path.name} ({scene.width:,} × {scene.height:,})...")
            render_svg(scene, svg_path, cards=cards)
            generated.append(svg_path)
        if not args.skip_png:
            scale = choose_png_scale(
                scene,
                args.png_scale,
                max_pixels=args.max_png_pixels,
                prompts=prompts,
            )
            if scale is not None:
                png_path = output_dir / f"{stem}.png"
                print(
                    f"Writing {png_path.name} at scale {scale:.2f} "
                    f"({round(scene.width * scale):,} × {round(scene.height * scale):,})..."
                )
                render_png(scene, png_path, cards=cards, scale=scale)
                generated.append(png_path)

        if not args.no_site:
            print(f"Writing GitHub Pages viewer to {args.site_dir.expanduser().resolve()}...")
            generated.extend(
                export_github_pages_site(
                    scene,
                    args.site_dir,
                    cards=cards,
                    edge_sources=edge_sources,
                )
            )

    if args.output_mode in {"all", "families"}:
        families_dir = output_dir / "families"
        families_dir.mkdir(parents=True, exist_ok=True)
        selected_roots = roots[: args.family_limit] if args.family_limit else roots
        for index, root in enumerate(selected_roots, start=1):
            family_people = descendants(root, people_by_id)
            family_global_years = global_years if args.family_year_mode == "global" else None
            scene = make_scene(
                f"family-{index:02d}",
                f"YJMB Family Tree — {root.stable_name}",
                family_people,
                [root],
                cards=cards,
                card_width=card_width,
                card_height=card_height,
                global_years=family_global_years,
                is_full_tree=False,
            )
            stem = f"{index:02d}_{safe_filename(root.stable_name)}_Family_Tree"
            print(
                f"Family {index}/{len(selected_roots)}: {root.stable_name} "
                f"({len(family_people)} people)"
            )
            if not args.skip_svg:
                path = families_dir / f"{stem}.svg"
                render_svg(scene, path, cards=cards)
                generated.append(path)
            if not args.skip_png:
                scale = choose_png_scale(
                    scene,
                    args.png_scale,
                    max_pixels=args.max_png_pixels,
                    prompts=prompts,
                )
                if scale is not None:
                    path = families_dir / f"{stem}.png"
                    render_png(scene, path, cards=cards, scale=scale)
                    generated.append(path)

    if not args.no_legend:
        generated.extend(
            render_legend(
                output_dir,
                png=not args.skip_png,
                svg=not args.skip_svg,
            )
        )

    report_path = output_dir / "YJMB_Full_Band_Tree_Report.txt"
    generated.append(report_path)
    write_report(
        report_path,
        workbook_path=workbook_path,
        sheet_name=args.sheet,
        people=people,
        roots=roots,
        edge_sources=edge_sources,
        issues=issues,
        generated=generated,
    )
    prompts.save()

    print("\nGeneration complete.")
    print(f"People: {len(people)}")
    print(f"Family trees: {len(roots)}")
    print(f"Generated files: {len(generated)}")
    print(f"Report: {report_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (TreeDataError, FileNotFoundError, PermissionError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
