#!/usr/bin/env python3
"""Scan/apply the v18 Saxophone label migration in a workbook.

Default mode is read-only. --apply creates a timestamped backup before replacing
the legacy combined sax label with Saxophone in non-formula string cells.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "YJMB Trees.xlsx"
LEGACY = re.compile(r"sax\s*/\s*saxophone", re.I)


def migrate(workbook: Path, *, apply_changes: bool, backup_dir: Path | None = None) -> dict[str, object]:
    workbook = workbook.expanduser().resolve()
    wb = load_workbook(workbook)
    hits: list[tuple[str, str]] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or value.startswith("=") or not LEGACY.search(value):
                        continue
                    hits.append((ws.title, cell.coordinate))
                    if apply_changes:
                        cell.value = LEGACY.sub("Saxophone", value)

        backup_path: Path | None = None
        if apply_changes and hits:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            destination = (backup_dir.expanduser().resolve() if backup_dir else workbook.parent / "backups" / "data_cleanup" / stamp)
            destination.mkdir(parents=True, exist_ok=True)
            backup_path = destination / workbook.name
            shutil.copy2(workbook, backup_path)
            tmp = workbook.with_suffix(".v18-saxophone.tmp.xlsx")
            wb.save(tmp)
            tmp.replace(workbook)

        return {
            "changed": bool(apply_changes and hits),
            "count": len(hits),
            "cells": [{"sheet": sheet, "cell": cell} for sheet, cell in hits],
            "backup": str(backup_path) if backup_path else None,
        }
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Scan/apply the v18 Saxophone label migration.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--apply", action="store_true", help="Write replacements after creating a backup.")
    parser.add_argument("--backup-dir", type=Path, default=None, help="Optional backup destination directory.")
    parser.add_argument("--result", type=Path, default=None, help="Optional JSON result file.")
    args = parser.parse_args()

    result = migrate(args.workbook, apply_changes=args.apply, backup_dir=args.backup_dir)
    mode = "Applied" if args.apply else "Found"
    print(f"{mode} {result['count']} legacy Saxophone-label occurrence(s).")
    for item in result["cells"]:
        print(f"  {item['sheet']}!{item['cell']}")
    if result.get("backup"):
        print(f"Backup: {result['backup']}")
    if args.result:
        args.result.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
