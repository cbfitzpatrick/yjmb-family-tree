#!/usr/bin/env python3
"""Normalize section-name variants in YJMB Trees.xlsx conservatively.

Default is read-only. Use --apply to write deterministic normalizations.
A timestamped backup is created before the first write.

The script updates:
  * Instrument / Section cells
  * VET and RAT relationship section text in the final (...) group
  * Section Nickname(s) and Specific Instrument(s) section labels before ':'

Broad section names are canonicalized while recognized subsection/instrument
information is retained, e.g. "Alto Sax" -> "Saxophone — Alto Saxophone" and
"Rifle" -> "Guard — Rifle". If a cell contains unexplained text, it is
reported and left unchanged instead of discarding the extra detail.
"""
from __future__ import annotations

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from yjmb_taxonomy import (
    canonical_formal_roles,
    canonical_section_text_with_details,
    formal_roles_in_text,
    key,
    norm,
    section_residual_words,
    strip_formal_role_phrases,
)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "YJMB Trees.xlsx"
DEFAULT_SHEET = "People on Tree"
RELATION_RE = re.compile(r"^(.*)\(([^()]*)\)\s*\(([^()]*)\)\s*$")


def header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", norm(value).casefold())


def find_headers(ws):
    for row in range(1, min(10, ws.max_row) + 1):
        headers = {c: header_key(ws.cell(row, c).value) for c in range(1, ws.max_column + 1)}
        if "givenpreferredname" in headers.values() and "familymaidenname" in headers.values():
            return row, headers
    raise RuntimeError("Could not find the People on Tree header row.")


def find_col(headers: dict[int, str], *names: str) -> int | None:
    wanted = {header_key(name) for name in names}
    for col, value in headers.items():
        if value in wanted:
            return col
    return None


def normalize_section_value(raw: object) -> tuple[str | None, str | None]:
    text = norm(raw)
    if not text:
        return None, None
    canonical = canonical_section_text_with_details(text)
    if not canonical:
        return None, f"no recognized section in {text!r}"
    residual = section_residual_words(text)
    if residual:
        return None, f"unexplained text {residual!r} in {text!r}"
    return canonical, None


def normalize_relation(raw: object) -> tuple[str | None, str | None]:
    text = norm(raw)
    if not text:
        return None, None
    match = RELATION_RE.match(text)
    if not match:
        return None, "relationship format was not Name (Year) (Section)"
    prefix = match.group(1).rstrip()
    year = norm(match.group(2))
    section = norm(match.group(3))
    canonical, error = normalize_section_value(section)
    if error:
        return None, error
    updated = f"{prefix} ({year}) ({canonical})"
    return updated, None


def normalize_labeled_pairs(raw: object) -> tuple[str | None, str | None]:
    """Normalize 'Trumpets: X; Mellos: Y' section labels without touching values."""
    text = norm(raw)
    if not text:
        return None, None
    parts = [part.strip() for part in text.split(";") if part.strip()]
    changed = False
    out = []
    for part in parts:
        if ":" not in part:
            return None, f"expected 'Section: value' but found {part!r}"
        label, value = part.split(":", 1)
        canonical, error = normalize_section_value(label)
        if error:
            return None, error
        out.append(f"{canonical}: {norm(value)}")
        changed |= canonical != norm(label)
    result = "; ".join(out)
    return result if changed else text, None


def backup(workbook: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    folder = workbook.parent / "backups" / "data_cleanup" / stamp
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / workbook.name
    shutil.copy2(workbook, target)
    return target


def main() -> int:
    ap = argparse.ArgumentParser(description="Normalize YJMB section name variants safely.")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--apply", action="store_true", help="Write deterministic changes. Default is scan-only.")
    args = ap.parse_args()

    path = args.workbook.expanduser().resolve()
    wb = load_workbook(path)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Worksheet not found: {args.sheet}")
    ws = wb[args.sheet]
    header_row, headers = find_headers(ws)

    instrument_cols = [c for c, h in headers.items() if h in {"instrument", "instruments", "section"}]
    relationship_cols = [c for c, h in headers.items() if h == "vet" or re.fullmatch(r"rat\d+", h)]
    pair_cols = [c for c, h in headers.items() if h in {"sectionnicknames", "specificinstruments"}]
    formal_col = find_col(headers, "Marching Band Leadership Role(s)", "Formal Leadership Position(s)", "Formal Leadership Role(s)")

    changes: list[tuple[int, int, str, str]] = []
    unresolved: list[str] = []

    for row in range(header_row + 1, ws.max_row + 1):
        for col in instrument_cols:
            old = norm(ws.cell(row, col).value)
            if not old:
                continue
            section_input = old
            embedded_roles = formal_roles_in_text(old)
            if embedded_roles:
                structured_roles = canonical_formal_roles(ws.cell(row, formal_col).value) if formal_col else []
                structured_keys = {key(role) for role in structured_roles}
                if all(key(role) in structured_keys for role in embedded_roles):
                    section_input = strip_formal_role_phrases(old)
                else:
                    unresolved.append(
                        f"row {row} {ws.cell(header_row,col).value}: leadership text {embedded_roles!r} is mixed into the section; "
                        "run classifyLeadershipPositions.py --apply first, then rerun this script"
                    )
                    continue
            new, error = normalize_section_value(section_input)
            if error:
                unresolved.append(f"row {row} {ws.cell(header_row,col).value}: {error}")
            elif new and new != old:
                changes.append((row, col, old, new))
        for col in relationship_cols:
            old = norm(ws.cell(row, col).value)
            if not old:
                continue
            new, error = normalize_relation(old)
            if error:
                unresolved.append(f"row {row} {ws.cell(header_row,col).value}: {error}")
            elif new and new != old:
                changes.append((row, col, old, new))
        for col in pair_cols:
            old = norm(ws.cell(row, col).value)
            if not old:
                continue
            new, error = normalize_labeled_pairs(old)
            if error:
                unresolved.append(f"row {row} {ws.cell(header_row,col).value}: {error}")
            elif new and new != old:
                changes.append((row, col, old, new))

    print(f"Workbook: {path}")
    print(f"Deterministic section normalizations: {len(changes)}")
    for row, col, old, new in changes[:100]:
        print(f"  row {row} {ws.cell(header_row,col).value}: {old!r} -> {new!r}")
    if len(changes) > 100:
        print(f"  ... {len(changes)-100} more")
    print(f"Needs manual review: {len(unresolved)}")
    for item in unresolved[:100]:
        print(f"  ! {item}")
    if len(unresolved) > 100:
        print(f"  ... {len(unresolved)-100} more")

    if not args.apply:
        print("\nScan only. Re-run with --apply to write the deterministic changes.")
        wb.close()
        return 0

    if not changes:
        print("\nNothing to change.")
        wb.close()
        return 0

    backup_path = backup(path)
    for row, col, _, new in changes:
        ws.cell(row, col).value = new
    wb.save(path)
    wb.close()
    print(f"\nApplied {len(changes)} change(s). Backup: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
