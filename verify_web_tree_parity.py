#!/usr/bin/env python3
"""Fail closed if the encrypted website dataset loses workbook rows/fields.

Designed for the protected GitHub Actions build.  It compares the generated
``docs/data/tree_data.enc`` against the just-decrypted authoritative workbook.
No member values are printed on failure; diagnostics identify only worksheet
row numbers and column labels so Actions logs do not disclose protected data.
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import re
import unicodedata
from pathlib import Path

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from openpyxl import load_workbook


def normalize_spaces(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u00a0", " ").replace("’", "'").replace("`", "'")
    return re.sub(r"\s+", " ", text).strip()


def normalized_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_spaces(value).casefold())


def load_key() -> bytes:
    raw = os.environ.get("TREE_DATA_KEY_B64", "").strip()
    if not raw:
        raise RuntimeError("TREE_DATA_KEY_B64 is required for protected parity validation.")
    key = base64.b64decode(raw, validate=True)
    if len(key) != 32:
        raise RuntimeError("TREE_DATA_KEY_B64 must decode to exactly 32 bytes.")
    return key


def decrypt_payload(path: Path) -> dict:
    envelope = json.loads(path.read_text(encoding="utf-8"))
    if envelope.get("format") != "yjmb-tree-encrypted-v3":
        raise RuntimeError("Generated tree payload is not yjmb-tree-encrypted-v3.")
    iv = base64.b64decode(envelope["dataIv"], validate=True)
    ciphertext = base64.b64decode(envelope["ciphertext"], validate=True)
    plaintext = AESGCM(load_key()).decrypt(iv, ciphertext, None)
    payload = json.loads(plaintext.decode("utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("people"), list):
        raise RuntimeError("Decrypted tree payload does not contain a people list.")
    return payload


def discover_header(ws) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row, 10) + 1):
        by_col = {col: normalized_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        values = set(by_col.values())
        if "givenpreferredname" in values and "familymaidenname" in values:
            name_cols: dict[str, int] = {}
            aliases = {
                "given": {"givenpreferredname", "givenname", "preferredname", "firstname"},
                "nickname": {"nickname", "nickname"},
                "family": {"familymaidenname", "familyname", "maidenname", "lastname", "surname"},
                "married": {"marriedname", "marriedsurname", "spousesurname", "currentlastname"},
            }
            for field, keys in aliases.items():
                for col, key in by_col.items():
                    if key in keys:
                        name_cols[field] = col
                        break
            missing = [field for field in ("given", "nickname", "family", "married") if field not in name_cols]
            if missing:
                raise RuntimeError("Master worksheet is missing required name columns.")
            return row, name_cols
    raise RuntimeError("Could not find the People on Tree header row.")


def expected_rows(workbook_path: Path, sheet_name: str) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Worksheet {sheet_name!r} not found.")
        ws = wb[sheet_name]
        header_row, name_cols = discover_header(ws)
        headers = [
            (col, normalize_spaces(ws.cell(header_row, col).value) or f"Column {col}")
            for col in range(1, ws.max_column + 1)
        ]
        result: dict[str, list[dict[str, str]]] = {}
        for row in range(header_row + 1, ws.max_row + 1):
            if not any(normalize_spaces(ws.cell(row, name_cols[field]).value) for field in ("given", "nickname", "family", "married")):
                continue
            fields: list[dict[str, str]] = []
            for col, label in headers:
                cell = ws.cell(row, col)
                value = normalize_spaces(cell.value)
                hyperlink = getattr(cell, "hyperlink", None)
                target = normalize_spaces(getattr(hyperlink, "target", "")) if hyperlink else ""
                if target:
                    value = f"{value} ({target})" if value and value != target else target
                fields.append({"label": label, "value": value})
            result[f"row-{row}"] = fields
        return result
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Verify encrypted website source-field parity with the protected workbook.")
    parser.add_argument("--workbook", type=Path, default=Path("YJMB Trees.xlsx"))
    parser.add_argument("--tree", type=Path, default=Path("docs/data/tree_data.enc"))
    parser.add_argument("--sheet", default="People on Tree")
    args = parser.parse_args()

    expected = expected_rows(args.workbook, args.sheet)
    payload = decrypt_payload(args.tree)
    actual_people = {str(person.get("id", "")): person for person in payload.get("people", [])}

    problems: list[str] = []
    expected_ids = set(expected)
    actual_ids = set(actual_people)
    for person_id in sorted(expected_ids - actual_ids, key=lambda x: int(x.split("-")[-1])):
        problems.append(f"worksheet row {person_id.split('-')[-1]} is missing from encrypted site data")
    for person_id in sorted(actual_ids - expected_ids):
        problems.append(f"encrypted site contains unexpected record {person_id}")

    for person_id in sorted(expected_ids & actual_ids, key=lambda x: int(x.split("-")[-1])):
        exp_fields = expected[person_id]
        got_fields = actual_people[person_id].get("sourceFields")
        if not isinstance(got_fields, list):
            problems.append(f"worksheet row {person_id.split('-')[-1]} has no sourceFields in encrypted site data")
            continue
        exp_map = {normalized_header(item["label"]): item for item in exp_fields}
        got_map = {normalized_header(item.get("label", "")): item for item in got_fields if isinstance(item, dict)}
        for key, exp in exp_map.items():
            got = got_map.get(key)
            if got is None:
                problems.append(f"worksheet row {person_id.split('-')[-1]} is missing field {exp['label']!r} in encrypted site data")
            elif normalize_spaces(got.get("value", "")) != exp["value"]:
                problems.append(f"worksheet row {person_id.split('-')[-1]} field {exp['label']!r} differs between workbook and encrypted site data")

    if problems:
        print("ERROR: generated protected site is not a lossless workbook projection.")
        for problem in problems[:30]:
            print(f"  - {problem}")
        if len(problems) > 30:
            print(f"  - ... and {len(problems) - 30} more parity problem(s)")
        return 1

    print(f"Protected tree parity OK: {len(expected)} workbook person row(s) and their source fields are present in encrypted site data.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
