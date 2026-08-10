"""Header-aware workbook helpers for the YJMB family-tree spreadsheets.

Supported input schemas:
- Original: ``Name``
- Earlier migration: ``Given/Preferred Name`` + ``Family/Maiden Name``
- Current: ``Given/Preferred Name`` + ``Nickname`` + ``Family/Maiden Name`` +
  ``Married Name``

Tree relationships continue to use the family/maiden surname as the stable
identity. The married surname is retained as a searchable alias and separate
spreadsheet field.
"""

from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
import shutil
import tempfile
from typing import Iterable, Iterator, Mapping

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from name_tools import canonical_key, normalize_spaces, normalized_header, split_instrument_terms


GIVEN_HEADER = "Given/Preferred Name"
NICKNAME_HEADER = "Nickname"
FAMILY_HEADER = "Family/Maiden Name"
MARRIED_HEADER = "Married Name"
OLD_NAME_HEADER = "Name"
NAME_FIELDS: tuple[str, ...] = ("given", "nickname", "family", "married")
CANONICAL_FIELDS: tuple[str, ...] = (
    "full_name",
    "current_name",
    "descriptive_name",
    "given",
    "nickname",
    "family",
    "married",
    "rat_year",
    "instrument",
    "position_and_year",
    "notes",
    "links",
    "vet",
    "rat_1",
    "rat_2",
    "rat_3",
    "rat_4",
    "rat_5",
    "rat_6",
    "rat_7",
)

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "given": ("givenpreferredname", "givenname", "preferredname", "firstname"),
    "nickname": ("nickname", "nick", "preferrednickname"),
    "family": ("familymaidenname", "familyname", "maidenname", "lastname", "surname"),
    "married": ("marriedname", "marriedsurname", "spousesurname", "currentlastname"),
    "old_name": ("name", "fullname"),
    "rat_year": ("ratyear", "year"),
    "instrument": ("instrument", "instruments"),
    "position_and_year": ("positionandyear", "positionyear"),
    "notes": ("notes", "pleaseincludeanyextrainformationaboutyourtreehere"),
    "links": ("links", "link"),
    "vet": ("vet", "vetsnameratyearandinstruments"),
    "rat_1": ("rat1", "rat1snameratyearandinstruments"),
    "rat_2": ("rat2", "rat2snameratyearandinstruments"),
    "rat_3": ("rat3", "rat3snameratyearandinstruments"),
    "rat_4": ("rat4", "rat4snameratyearandinstrumentssameformatting"),
    "rat_5": ("rat5", "rat5snameratyearandinstrumentssameformatting"),
    "rat_6": ("rat6", "rat6snameratyearandinstrumentssameformatting"),
    "rat_7": ("rat7",),
}

SECTION_RULES: dict[str, tuple[str, ...]] = {
    "FlutePiccolo": ("flute", "piccolo"),
    "Clarinet": ("clarinet",),
    "Saxophone": ("sax", "saxophone"),
    "Trumpet": ("trumpet",),
    "Mellophone": ("mellophone", "mello"),
    "Trombone": ("trombone",),
    "Baritone": ("baritone", "euphonium"),
    "Sousaphone": ("sousaphone", "tuba"),
    "Front Ensemble": ("front ensemble", "pit", "front emsemble"),
    "Battery": ("battery", "snare", "bass drum", "quads", "quad", "tenor", "cymbal", "drumline"),
    "Guard": ("guard", "colorguard", "color guard"),
    "Goldrush": ("goldrush", "gold rush", "dance"),
    "Golden Girl": ("golden girl",),
    "Drum Major": ("drum major",),
}


def project_root_from_script(script_file: str | Path) -> Path:
    return Path(script_file).resolve().parent


def full_name(given: object, family: object, married: object = "") -> str:
    """Return the stable tree identity, preferring family/maiden surname."""
    surname = normalize_spaces(family) or normalize_spaces(married)
    return normalize_spaces(f"{normalize_spaces(given)} {surname}")


def current_name(given: object, family: object, married: object = "") -> str:
    """Return the current surname when Married Name is populated."""
    surname = normalize_spaces(married) or normalize_spaces(family)
    return normalize_spaces(f"{normalize_spaces(given)} {surname}")


def descriptive_name(
    given: object,
    nickname: object,
    family: object,
    married: object,
) -> str:
    given_text = normalize_spaces(given)
    nickname_text = normalize_spaces(nickname)
    family_text = normalize_spaces(family)
    married_text = normalize_spaces(married)
    first = f'{given_text} "{nickname_text}"' if nickname_text else given_text
    base = normalize_spaces(f"{first} {family_text or married_text}")
    if married_text and canonical_key(married_text) != canonical_key(family_text):
        return f"{base} (married name: {married_text})"
    return base


def find_header_row(ws, *, max_rows: int = 10) -> int | None:
    for row_number in range(1, min(ws.max_row, max_rows) + 1):
        headers = {
            normalized_header(ws.cell(row_number, col).value)
            for col in range(1, ws.max_column + 1)
        }
        if "name" in headers or "givenpreferredname" in headers:
            return row_number
    return None


def header_map(ws, header_row: int | None = None) -> dict[str, int]:
    header_row = header_row or find_header_row(ws)
    if header_row is None:
        return {}
    raw = {
        normalized_header(ws.cell(header_row, column).value): column
        for column in range(1, ws.max_column + 1)
        if normalize_spaces(ws.cell(header_row, column).value)
    }
    result: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in raw:
                result[field] = raw[alias]
                break
    return result


def is_split_name_sheet(ws) -> bool:
    """Return True only for the current four-column name schema."""
    mapping = header_map(ws)
    return all(field in mapping for field in NAME_FIELDS)


def has_any_name_schema(ws) -> bool:
    mapping = header_map(ws)
    return "old_name" in mapping or "given" in mapping


def record_from_row(ws, row_number: int, mapping: Mapping[str, int] | None = None) -> dict[str, object]:
    mapping = dict(mapping or header_map(ws))

    def value(field: str) -> object:
        column = mapping.get(field)
        return ws.cell(row_number, column).value if column else None

    given = normalize_spaces(value("given"))
    nickname = normalize_spaces(value("nickname"))
    family = normalize_spaces(value("family"))
    married = normalize_spaces(value("married"))
    old_name = normalize_spaces(value("old_name"))
    stable = full_name(given, family, married) if given or family or married else old_name
    current = current_name(given, family, married) if given or family or married else old_name
    record: dict[str, object] = {
        "full_name": stable,
        "current_name": current,
        "descriptive_name": descriptive_name(given, nickname, family, married) if given else old_name,
        "given": given,
        "nickname": nickname,
        "family": family,
        "married": married,
        "rat_year": value("rat_year"),
        "instrument": value("instrument"),
        "position_and_year": value("position_and_year"),
        "notes": value("notes"),
        "links": value("links"),
        "vet": value("vet"),
    }
    for index in range(1, 8):
        record[f"rat_{index}"] = value(f"rat_{index}")
    return record


def record_name_aliases(record: Mapping[str, object]) -> list[str]:
    aliases = {
        normalize_spaces(record.get("full_name")),
        normalize_spaces(record.get("current_name")),
        full_name(record.get("given"), record.get("family"), record.get("married")),
        current_name(record.get("given"), record.get("family"), record.get("married")),
    }
    return sorted((name for name in aliases if name), key=str.casefold)


def iter_records(ws) -> Iterator[tuple[int, dict[str, object]]]:
    header_row = find_header_row(ws)
    if header_row is None:
        return
    mapping = header_map(ws, header_row)
    for row_number in range(header_row + 1, ws.max_row + 1):
        record = record_from_row(ws, row_number, mapping)
        if normalize_spaces(record["full_name"] or record["current_name"]):
            yield row_number, record


def load_people_data(path: str | Path, sheet_name: str | None = None, *, max_rats: int = 7) -> list[list[object]]:
    """Return the historical row positions expected by the tree generators.

    Output order: stable full name, RAT year, instrument, nickname,
    position/year, notes, links, VET, RAT 1..RAT 7.

    The stable full name uses Family/Maiden Name. Married Name remains available
    in the workbook and matching aliases but does not rewrite lineage references.
    """
    workbook = load_workbook(path, data_only=False)
    ws = workbook[sheet_name] if sheet_name else workbook.active
    rows: list[list[object]] = []
    for _, record in iter_records(ws):
        rows.append(
            [
                record["full_name"],
                record["rat_year"],
                record["instrument"],
                record["nickname"],
                record["position_and_year"],
                record["notes"],
                record["links"],
                record["vet"],
                *[record[f"rat_{index}"] for index in range(1, max_rats + 1)],
            ]
        )
    workbook.close()
    return rows


def collect_names_from_workbook(path: str | Path) -> list[str]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    names: set[str] = set()
    for ws in workbook.worksheets:
        for _, record in iter_records(ws):
            names.update(record_name_aliases(record))
    workbook.close()
    return sorted(names, key=str.casefold)


def row_index(ws) -> dict[tuple[str, str], int]:
    index: dict[tuple[str, str], int] = {}
    for row_number, record in iter_records(ws):
        year = normalize_spaces(record["rat_year"])[:4]
        for alias in record_name_aliases(record):
            index[(canonical_key(alias), year)] = row_number
    return index


def _copy_cell_style(source, destination) -> None:
    if source.has_style:
        destination._style = copy(source._style)
    if source.number_format:
        destination.number_format = source.number_format
    if source.font:
        destination.font = copy(source.font)
    if source.fill:
        destination.fill = copy(source.fill)
    if source.border:
        destination.border = copy(source.border)
    if source.alignment:
        destination.alignment = copy(source.alignment)
    if source.protection:
        destination.protection = copy(source.protection)


def copy_row_style(ws, source_row: int, destination_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        _copy_cell_style(ws.cell(source_row, column), ws.cell(destination_row, column))
    ws.row_dimensions[destination_row].height = ws.row_dimensions[source_row].height


def write_record(ws, row_number: int, record: Mapping[str, object]) -> None:
    mapping = header_map(ws)
    if not all(field in mapping for field in NAME_FIELDS):
        raise ValueError(
            f"Worksheet {ws.title!r} has not been migrated to the four name columns."
        )
    values = {
        "given": record.get("given"),
        "nickname": record.get("nickname"),
        "family": record.get("family"),
        "married": record.get("married"),
        "rat_year": record.get("rat_year"),
        "instrument": record.get("instrument"),
        "position_and_year": record.get("position_and_year"),
        "notes": record.get("notes"),
        "links": record.get("links"),
        "vet": record.get("vet"),
        **{f"rat_{index}": record.get(f"rat_{index}") for index in range(1, 8)},
    }
    for field, value in values.items():
        column = mapping.get(field)
        if column:
            ws.cell(row_number, column).value = value


def extend_tables_to_row(ws, row_number: int) -> None:
    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row <= row_number and row_number > max_row:
            table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{row_number}"
    if ws.auto_filter.ref:
        min_col, min_row, max_col, max_row = range_boundaries(ws.auto_filter.ref)
        if min_row <= row_number and row_number > max_row:
            ws.auto_filter.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{row_number}"


def append_record(ws, record: Mapping[str, object]) -> int:
    row_number = ws.max_row + 1
    header_row = find_header_row(ws) or 1
    if row_number > header_row + 1:
        copy_row_style(ws, row_number - 1, row_number)
    else:
        copy_row_style(ws, header_row, row_number)
    write_record(ws, row_number, record)
    extend_tables_to_row(ws, row_number)
    return row_number


def backup_file(path: str | Path, backup_root: str | Path) -> Path:
    source = Path(path)
    backup_root = Path(backup_root)
    backup_root.mkdir(parents=True, exist_ok=True)
    destination = backup_root / source.name
    counter = 1
    while destination.exists():
        destination = backup_root / f"{source.stem}_{counter}{source.suffix}"
        counter += 1
    shutil.copy2(source, destination)
    return destination


def save_workbook_atomic(workbook, destination: str | Path) -> None:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        suffix=destination.suffix,
        dir=destination.parent,
        delete=False,
    ) as handle:
        temp_path = Path(handle.name)
    try:
        workbook.save(temp_path)
        temp_path.replace(destination)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


def matching_section_sheets(instrument: object, available_sheet_names: Iterable[str]) -> list[str]:
    terms = split_instrument_terms(instrument)
    whole = normalize_spaces(instrument).casefold().replace("front emsemble", "front ensemble")
    available = set(available_sheet_names)
    matches: list[str] = []
    for sheet_name, keywords in SECTION_RULES.items():
        if sheet_name not in available:
            continue
        if any(keyword in whole for keyword in keywords) or any(
            any(keyword in term for keyword in keywords) for term in terms
        ):
            matches.append(sheet_name)
    return matches


def timestamped_backup_directory(project_root: str | Path, label: str) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(project_root) / "backups" / f"{label}_{stamp}"
