#!/usr/bin/env python3
"""Extract high-confidence married/current surnames from YJMB note fields.

Default is scan-only. Use --apply to populate blank Married Name cells.
A timestamped workbook backup is created before the first write.

The workbook keeps its established schema: Married Name stores the person's
changed/current surname, not a duplicated given name. The scanner prints the
full derived current name (Given/Preferred + Married Name) for every proposal.
The website generator can therefore search/display the full current name while
relationship identity remains Given/Preferred + Family/Maiden Name.

Only explicit patterns such as "married name: Smith", "married name is Jane
Smith", "current name: Jane Smith", and "now known as Jane Smith" are eligible
for automatic application. Phrases such as "married to John Smith" are never
auto-applied and are reported for review.
"""
from __future__ import annotations

import argparse
import re
import shutil
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "YJMB Trees.xlsx"
DEFAULT_SHEET = "People on Tree"


def norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).replace("\u00a0", " ").replace("’", "'")
    return re.sub(r"\s+", " ", text).strip()


def key(value: object) -> str:
    text = unicodedata.normalize("NFKD", norm(value).casefold())
    return "".join(ch for ch in text if ch.isalnum())


def hk(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", norm(value).casefold())


def discover(ws) -> tuple[int, dict[int, str]]:
    for row in range(1, min(10, ws.max_row) + 1):
        headers = {col: hk(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if "givenpreferredname" in headers.values() and "familymaidenname" in headers.values():
            return row, headers
    raise RuntimeError("Could not find People on Tree headers.")


def find_col(headers: dict[int, str], *names: str) -> int | None:
    wanted = {hk(name) for name in names}
    for col, value in headers.items():
        if value in wanted:
            return col
    return None


def note_columns(headers: dict[int, str], include_memory: bool) -> list[int]:
    columns: list[int] = []
    for col, header in headers.items():
        if any(token in header for token in (
            "note", "notes", "comment", "comments", "remark", "remarks",
            "additionalinfo", "additionalinformation",
        )):
            columns.append(col)
        elif include_memory and header == "favoritetechbandmemory":
            columns.append(col)
    return sorted(set(columns))


# Capture conservatively through punctuation/newline boundaries. The value is
# cleaned further before it is considered safe to apply.
VALUE = r"([^.;\n\r]{1,90})"
EXPLICIT_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("married name", re.compile(rf"\bmarried\s+(?:last\s+)?name\s*(?:is|was|:|=|-)?\s*{VALUE}", re.I)),
    ("married surname", re.compile(rf"\bmarried\s+surname\s*(?:is|was|:|=|-)?\s*{VALUE}", re.I)),
    ("current name", re.compile(rf"\bcurrent\s+(?:full\s+)?name\s*(?:is|:|=|-)?\s*{VALUE}", re.I)),
    ("now known as", re.compile(rf"\bnow\s+(?:known\s+as|goes\s+by)\s+{VALUE}", re.I)),
    ("name changed to", re.compile(rf"\bname\s+(?:was\s+)?changed\s+to\s+{VALUE}", re.I)),
)
AMBIGUOUS_MARRIAGE_RE = re.compile(
    r"\b(?:married\s+to|married\s+(?!name\b|surname\b)|spouse|husband|wife|wedding)\b",
    re.I,
)
TRAILING_NOISE_RE = re.compile(
    r"\s+(?:after|before|during|while|when|since|from|who|and\s+(?:is|was|played|marched)|in\s+the\s+band)\b.*$",
    re.I,
)
SURNAME_PREFIXES = {"de", "del", "de la", "van", "von", "la", "le", "st", "st.", "saint", "mc", "mac"}


@dataclass
class Proposal:
    row: int
    source_header: str
    source_text: str
    matched_kind: str
    captured: str
    surname: str
    full_current_name: str


def clean_candidate(value: str) -> str:
    value = norm(value)
    value = re.sub(r"^[\s:'\"()\[\]-]+|[\s:'\"()\[\]-]+$", "", value)
    value = TRAILING_NOISE_RE.sub("", value)
    # Remove common explanatory wrappers without touching apostrophes/hyphens in names.
    value = re.sub(r"\s*\((?:formerly|previously|nee|née).*$", "", value, flags=re.I)
    return norm(value)


def surname_from_candidate(candidate: str, given: str, nickname: str) -> tuple[str | None, str | None]:
    """Return (surname, reason_if_unsafe)."""
    candidate = clean_candidate(candidate)
    if not candidate:
        return None, "empty candidate"
    if re.search(r"\b(?:married\s+to|spouse|husband|wife)\b", candidate, re.I):
        return None, "looks like spouse information"
    tokens = candidate.split()
    if len(tokens) > 6:
        return None, "captured text is too long to be a name"

    aliases = [alias for alias in (norm(given), norm(nickname)) if alias]
    candidate_fold = candidate.casefold()
    for alias in aliases:
        alias_fold = alias.casefold()
        if candidate_fold == alias_fold:
            return None, "candidate contains only the given/nickname"
        if candidate_fold.startswith(alias_fold + " "):
            remainder = norm(candidate[len(alias):])
            # A 3+ token remainder can contain middle names. Do not guess unless
            # the tail clearly forms a surname prefix + surname.
            parts = remainder.split()
            if len(parts) <= 2:
                return remainder, None
            if len(parts) >= 2 and parts[-2].casefold().rstrip(".") in SURNAME_PREFIXES:
                return " ".join(parts[-2:]), None
            return None, "full-name capture may include middle names"

    # A one-token explicit married/current-name capture is safely a surname.
    if len(tokens) == 1:
        return candidate, None

    # Two-token candidates whose first token is not the person's given/nickname
    # could be a spouse or another person; report instead of applying.
    return None, "multi-word name does not begin with this person's given/nickname"


def backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    folder = path.parent / "backups" / "data_cleanup" / stamp
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / path.name
    shutil.copy2(path, target)
    return target


def main() -> int:
    ap = argparse.ArgumentParser(description="Extract married/current surnames from YJMB note fields safely.")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--include-memory", action="store_true", help="Also scan Favorite Tech Band Memory.")
    ap.add_argument("--apply", action="store_true", help="Populate only blank Married Name cells from high-confidence matches.")
    args = ap.parse_args()

    path = args.workbook.expanduser().resolve()
    wb = load_workbook(path)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Worksheet not found: {args.sheet}")
    ws = wb[args.sheet]
    header_row, headers = discover(ws)
    given_col = find_col(headers, "Given/Preferred Name", "Given Name", "First Name")
    nickname_col = find_col(headers, "Nickname")
    family_col = find_col(headers, "Family/Maiden Name", "Family Name", "Last Name", "Surname")
    married_col = find_col(headers, "Married Name", "Married Surname", "Current Last Name")
    notes = note_columns(headers, args.include_memory)
    if not all((given_col, family_col, married_col)):
        raise SystemExit("Required name columns were not found.")
    if not notes:
        print("No note/comment columns were found. Nothing was changed.")
        wb.close()
        return 0

    proposals: list[Proposal] = []
    conflicts: list[str] = []
    ambiguous: list[str] = []
    seen_rows: dict[int, Proposal] = {}

    for row in range(header_row + 1, ws.max_row + 1):
        given = norm(ws.cell(row, given_col).value)
        nickname = norm(ws.cell(row, nickname_col).value) if nickname_col else ""
        family = norm(ws.cell(row, family_col).value)
        existing = norm(ws.cell(row, married_col).value)
        if not given and not family:
            continue

        row_candidates: list[Proposal] = []
        row_ambiguous = False
        for col in notes:
            text = norm(ws.cell(row, col).value)
            if not text:
                continue
            header = norm(ws.cell(header_row, col).value) or f"Column {col}"
            matched_explicit = False
            for kind, pattern in EXPLICIT_PATTERNS:
                for match in pattern.finditer(text):
                    matched_explicit = True
                    captured = clean_candidate(match.group(1))
                    surname, reason = surname_from_candidate(captured, given, nickname)
                    if surname:
                        full = norm(f"{given} {surname}")
                        row_candidates.append(Proposal(row, header, text, kind, captured, surname, full))
                    else:
                        ambiguous.append(f"row {row} {given} {family}: {kind} -> {captured!r} ({reason})")
            if not matched_explicit and AMBIGUOUS_MARRIAGE_RE.search(text):
                row_ambiguous = True
        # Deduplicate identical surname proposals across note fields.
        by_surname: dict[str, Proposal] = {}
        for proposal in row_candidates:
            by_surname.setdefault(key(proposal.surname), proposal)
        if len(by_surname) > 1:
            conflicts.append(
                f"row {row} {given} {family}: notes suggest multiple married surnames: "
                + ", ".join(sorted(p.surname for p in by_surname.values()))
            )
            continue
        if not by_surname:
            if row_ambiguous:
                ambiguous.append(f"row {row} {given} {family}: marriage/spouse wording found without an explicit current-name pattern")
            continue
        proposal = next(iter(by_surname.values()))
        if existing:
            if key(existing) != key(proposal.surname):
                conflicts.append(
                    f"row {row} {given} {family}: Married Name is {existing!r}, but notes suggest {proposal.surname!r}"
                )
            continue
        proposals.append(proposal)
        seen_rows[row] = proposal

    print(f"Workbook: {path}")
    print("Note columns scanned: " + ", ".join(norm(ws.cell(header_row, col).value) for col in notes))
    print(f"High-confidence blank Married Name proposals: {len(proposals)}")
    for proposal in proposals[:150]:
        print(
            f"  row {proposal.row}: {proposal.full_current_name} "
            f"[Married Name={proposal.surname!r}; from {proposal.source_header}: {proposal.matched_kind}]"
        )
    if len(proposals) > 150:
        print(f"  ... {len(proposals) - 150} more")
    print(f"Conflicts requiring manual review: {len(conflicts)}")
    for item in conflicts[:100]:
        print(f"  ! {item}")
    print(f"Ambiguous marriage/name wording requiring manual review: {len(ambiguous)}")
    for item in ambiguous[:100]:
        print(f"  ? {item}")

    if not args.apply:
        print("\nScan only. Re-run with --apply to fill only the high-confidence blank Married Name cells.")
        wb.close()
        return 0
    if not proposals:
        print("\nNothing to apply.")
        wb.close()
        return 0

    backup_path = backup(path)
    for proposal in proposals:
        ws.cell(proposal.row, married_col).value = proposal.surname
    wb.save(path)
    wb.close()
    print(f"\nApplied {len(proposals)} Married Name update(s). Backup: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
