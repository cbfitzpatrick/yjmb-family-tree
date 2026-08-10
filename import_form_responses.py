"""Merge cleaned form responses into the YJMB master and section worksheets.

Run prepare_form_responses.py first. This importer reads the four-name-column ``Cleaned Responses``
sheet, merges exact matches, warns about likely typo/duplicate matches, appends new
people to ``People on Tree``, and mirrors each record to matching instrument sheets.
It can also synchronize the standalone trumpet and baritone workbooks.
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from difflib import get_close_matches
from pathlib import Path
import sys
from typing import Iterable, Mapping

from openpyxl import load_workbook

from name_tools import UserQuit, canonical_key, normalize_spaces, similarity
from tree_workbook import (
    append_record,
    backup_file,
    find_header_row,
    current_name,
    descriptive_name,
    full_name,
    header_map,
    is_split_name_sheet,
    iter_records,
    matching_section_sheets,
    record_from_row,
    record_name_aliases,
    row_index,
    save_workbook_atomic,
    timestamped_backup_directory,
    write_record,
)


CLEANED_SHEET = "Cleaned Responses"
MASTER_SHEET = "People on Tree"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--responses", type=Path, required=True, help="Workbook produced by prepare_form_responses.py.")
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="Directory containing the YJMB workbooks.",
    )
    parser.add_argument("--master", type=Path, help="Master workbook; default is <project-root>/YJMB Trees.xlsx.")
    parser.add_argument("--sync-legacy", action="store_true", help="Also update standalone trumpet/baritone workbooks.")
    parser.add_argument("--dry-run", action="store_true", help="Show changes without saving.")
    parser.add_argument("--yes", action="store_true", help="Save without the final confirmation prompt.")
    parser.add_argument("--non-interactive", action="store_true", help="Do not prompt about conflicts; preserve/merge values.")
    return parser.parse_args()


def ask(prompt: str, *, default: str = "") -> str:
    suffix = f" [{default}]" if default else ""
    try:
        value = input(f"{prompt}{suffix}: ").strip()
    except (EOFError, KeyboardInterrupt) as exc:
        raise UserQuit("Interactive input was cancelled.") from exc
    return value or default


def load_cleaned_records(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    if CLEANED_SHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError(
            f"{path.name} does not contain {CLEANED_SHEET!r}. Run prepare_form_responses.py first."
        )
    ws = workbook[CLEANED_SHEET]
    records = [record for _, record in iter_records(ws)]
    workbook.close()
    return records


def clean_value(value: object) -> str:
    return normalize_spaces(value)


def combine_distinct(existing: object, incoming: object, separator: str = " / ") -> str:
    old = clean_value(existing)
    new = clean_value(incoming)
    if not old:
        return new
    if not new:
        return old
    if canonical_key(old) == canonical_key(new) or canonical_key(new) in canonical_key(old):
        return old
    return f"{old}{separator}{new}"


def merge_records(
    existing: Mapping[str, object],
    incoming: Mapping[str, object],
    *,
    context: str,
    interactive: bool,
) -> tuple[dict[str, object], list[str]]:
    merged = dict(existing)
    changes: list[str] = []

    # The matched identity may use a maiden or married surname alias. Preserve
    # existing values unless a missing field can be filled or the user confirms
    # a replacement. Married Name is never combined with another surname.
    for field, label in (
        ("given", "Given/Preferred Name"),
        ("family", "Family/Maiden Name"),
        ("married", "Married Name"),
    ):
        old = clean_value(merged.get(field))
        new = clean_value(incoming.get(field))
        if not new or canonical_key(old) == canonical_key(new):
            continue
        if not old:
            merged[field] = new
            changes.append(f"filled {label}")
            continue
        if interactive:
            extra = " Confirm this is a spouse surname the person adopted." if field == "married" else ""
            answer = ask(
                f'{context}: {label} differs. Did you mean to use "{new}" instead of "{old}"?{extra} '
                "(y=replace, n=keep)",
                default="n",
            ).casefold()
            if answer in {"q", "quit", "exit"}:
                raise UserQuit("The user stopped the run.")
            if answer in {"y", "yes", "replace"}:
                merged[field] = new
                changes.append(f"replaced {label}")

    for field, label in (
        ("rat_year", "RAT year"),
        ("nickname", "nickname"),
        ("position_and_year", "position/year"),
        ("links", "links"),
    ):
        old = clean_value(merged.get(field))
        new = clean_value(incoming.get(field))
        if not new or canonical_key(old) == canonical_key(new):
            continue
        if not old:
            merged[field] = new
            changes.append(f"filled {label}")
            continue
        if interactive:
            answer = ask(
                f'{context}: {label} differs. Did you mean to use "{new}" instead of "{old}"? '
                "(y=replace, c=combine, n=keep)",
                default="n",
            ).casefold()
            if answer in {"q", "quit", "exit"}:
                raise UserQuit("The user stopped the run.")
            if answer in {"y", "yes", "replace"}:
                merged[field] = new
                changes.append(f"replaced {label}")
            elif answer in {"c", "combine"}:
                merged[field] = combine_distinct(old, new)
                changes.append(f"combined {label}")
        # Non-interactive mode preserves existing data.

    old_instrument = clean_value(merged.get("instrument"))
    new_instrument = clean_value(incoming.get("instrument"))
    combined_instrument = combine_distinct(old_instrument, new_instrument)
    if combined_instrument != old_instrument:
        merged["instrument"] = combined_instrument
        changes.append("combined instrument(s)")

    old_notes = clean_value(merged.get("notes"))
    new_notes = clean_value(incoming.get("notes"))
    if new_notes and canonical_key(new_notes) not in canonical_key(old_notes):
        merged["notes"] = f"{old_notes} | Form response: {new_notes}" if old_notes else new_notes
        changes.append("added notes")

    old_vet = clean_value(merged.get("vet"))
    new_vet = clean_value(incoming.get("vet"))
    if new_vet and canonical_key(new_vet) != canonical_key(old_vet):
        if not old_vet:
            merged["vet"] = new_vet
            changes.append("filled VET")
        elif interactive:
            answer = ask(
                f'{context}: VET differs. Did you mean "{new_vet}" instead of "{old_vet}"? '
                "(y=replace, n=keep, b=record both in notes)",
                default="n",
            ).casefold()
            if answer in {"q", "quit", "exit"}:
                raise UserQuit("The user stopped the run.")
            if answer in {"y", "yes", "replace"}:
                merged["vet"] = new_vet
                changes.append("replaced VET")
            elif answer in {"b", "both"}:
                note = f"Alternate VET from form: {new_vet}"
                merged["notes"] = combine_distinct(merged.get("notes"), note, separator=" | ")
                changes.append("recorded alternate VET in notes")
        else:
            note = f"Alternate VET from form: {new_vet}"
            merged["notes"] = combine_distinct(merged.get("notes"), note, separator=" | ")
            changes.append("recorded alternate VET in notes")

    existing_rats = [clean_value(merged.get(f"rat_{i}")) for i in range(1, 8)]
    incoming_rats = [clean_value(incoming.get(f"rat_{i}")) for i in range(1, 8)]
    unique = [value for value in existing_rats if value]
    keys = {canonical_key(value) for value in unique}
    for value in incoming_rats:
        if value and canonical_key(value) not in keys:
            unique.append(value)
            keys.add(canonical_key(value))
            changes.append("added RAT relationship")
    for i in range(1, 8):
        merged[f"rat_{i}"] = unique[i - 1] if i <= len(unique) else ""
    if len(unique) > 7:
        overflow = "; ".join(unique[7:])
        merged["notes"] = combine_distinct(
            merged.get("notes"),
            f"Additional RAT relationships: {overflow}",
            separator=" | ",
        )
        changes.append("recorded RAT overflow in notes")

    merged["full_name"] = full_name(merged.get("given"), merged.get("family"), merged.get("married"))
    merged["current_name"] = current_name(merged.get("given"), merged.get("family"), merged.get("married"))
    return merged, changes


def candidate_existing_row(ws, incoming: Mapping[str, object], *, interactive: bool) -> int | None:
    incoming_year = clean_value(incoming.get("rat_year"))[:4]
    incoming_aliases = {
        clean_value(incoming.get("full_name")),
        clean_value(incoming.get("current_name")),
        full_name(incoming.get("given"), incoming.get("family"), incoming.get("married")),
        current_name(incoming.get("given"), incoming.get("family"), incoming.get("married")),
    }
    incoming_aliases.discard("")
    index = row_index(ws)
    for alias in incoming_aliases:
        exact = index.get((canonical_key(alias), incoming_year))
        if exact:
            return exact

    candidates: list[tuple[float, int, str, str]] = []
    for row, record in iter_records(ws):
        existing_year = clean_value(record.get("rat_year"))[:4]
        existing_aliases = record_name_aliases(record)
        score = max(
            (similarity(incoming_alias, existing_alias) for incoming_alias in incoming_aliases for existing_alias in existing_aliases),
            default=0.0,
        )
        same_year = incoming_year and existing_year and incoming_year == existing_year
        if score >= (0.88 if same_year else 0.95):
            candidates.append((score, row, clean_value(record.get("descriptive_name") or record.get("full_name")), existing_year))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    score, row, name, year = candidates[0]
    if not interactive:
        return None
    incoming_label = descriptive_name(incoming.get("given"), incoming.get("nickname"), incoming.get("family"), incoming.get("married"))
    answer = ask(
        f'Did you mean existing entry "{name}" ({year}) instead of adding '
        f'"{incoming_label}" ({incoming_year})? Similarity {score:.1%} (y/n)',
        default="n",
    ).casefold()
    if answer in {"q", "quit", "exit"}:
        raise UserQuit("The user stopped the run.")
    return row if answer in {"y", "yes"} else None


def upsert_into_sheet(
    ws,
    incoming: Mapping[str, object],
    *,
    interactive: bool,
) -> tuple[str, int, list[str]]:
    row = candidate_existing_row(ws, incoming, interactive=interactive)
    context = f"{ws.title} / {incoming.get('descriptive_name') or incoming.get('full_name')} ({incoming.get('rat_year')})"
    if row:
        existing = record_from_row(ws, row)
        merged, changes = merge_records(existing, incoming, context=context, interactive=interactive)
        if changes:
            write_record(ws, row, merged)
            return "updated", row, changes
        return "unchanged", row, []
    new_row = append_record(ws, incoming)
    return "added", new_row, ["added new row"]


def choose_master_sheet(workbook):
    if MASTER_SHEET in workbook.sheetnames:
        return workbook[MASTER_SHEET]
    return workbook.active


def validate_target_workbook(path: Path, workbook) -> None:
    ws = choose_master_sheet(workbook)
    if not is_split_name_sheet(ws):
        raise ValueError(
            f"{path.name} does not use the four canonical name columns. Run migrate_name_columns.py before importing."
        )


def sync_record_to_workbook(
    workbook,
    incoming: Mapping[str, object],
    *,
    interactive: bool,
    include_sections: bool,
) -> list[tuple[str, str, int, list[str]]]:
    results: list[tuple[str, str, int, list[str]]] = []
    master_ws = choose_master_sheet(workbook)
    status, row, changes = upsert_into_sheet(master_ws, incoming, interactive=interactive)
    results.append((master_ws.title, status, row, changes))

    if include_sections:
        section_names = matching_section_sheets(incoming.get("instrument"), workbook.sheetnames)
        for sheet_name in section_names:
            if sheet_name == master_ws.title:
                continue
            ws = workbook[sheet_name]
            if not is_split_name_sheet(ws):
                continue
            status, row, changes = upsert_into_sheet(ws, incoming, interactive=interactive)
            results.append((ws.title, status, row, changes))
    return results


def should_sync_legacy(record: Mapping[str, object], kind: str) -> bool:
    instrument = clean_value(record.get("instrument")).casefold()
    if kind == "trumpet":
        return "trumpet" in instrument
    if kind == "baritone":
        return "baritone" in instrument or "euphonium" in instrument
    return False


def main() -> int:
    args = parse_args()
    responses = args.responses.expanduser().resolve()
    project_root = args.project_root.expanduser().resolve()
    master_path = args.master.expanduser().resolve() if args.master else project_root / "YJMB Trees.xlsx"
    if not responses.exists():
        print(f"Responses workbook does not exist: {responses}", file=sys.stderr)
        return 2
    if not master_path.exists():
        print(f"Master workbook does not exist: {master_path}", file=sys.stderr)
        return 2

    try:
        records = load_cleaned_records(responses)
    except ValueError as exc:
        print(exc, file=sys.stderr)
        return 2

    workbook_paths: dict[str, Path] = {"master": master_path}
    if args.sync_legacy:
        trumpet_path = project_root / "YJMB Trumpet Trees.xlsx"
        baritone_path = project_root / "YJMB Baritone Trees Updated.xlsx"
        if trumpet_path.exists():
            workbook_paths["trumpet"] = trumpet_path
        if baritone_path.exists():
            workbook_paths["baritone"] = baritone_path

    workbooks = {}
    try:
        for label, path in workbook_paths.items():
            wb = load_workbook(path)
            validate_target_workbook(path, wb)
            workbooks[label] = wb
    except Exception:
        for wb in workbooks.values():
            wb.close()
        raise

    report_rows: list[dict[str, object]] = []
    counts = {"added": 0, "updated": 0, "unchanged": 0}
    interactive = not args.non_interactive

    try:
        for index, incoming in enumerate(records, start=1):
            incoming = dict(incoming)
            incoming["full_name"] = full_name(incoming.get("given"), incoming.get("family"), incoming.get("married"))
            incoming["current_name"] = current_name(incoming.get("given"), incoming.get("family"), incoming.get("married"))
            incoming["descriptive_name"] = descriptive_name(incoming.get("given"), incoming.get("nickname"), incoming.get("family"), incoming.get("married"))
            if not incoming["full_name"]:
                continue
            targets: list[tuple[str, object, bool]] = [("master", workbooks["master"], True)]
            if "trumpet" in workbooks and should_sync_legacy(incoming, "trumpet"):
                targets.append(("trumpet", workbooks["trumpet"], False))
            if "baritone" in workbooks and should_sync_legacy(incoming, "baritone"):
                targets.append(("baritone", workbooks["baritone"], False))

            for label, workbook, include_sections in targets:
                results = sync_record_to_workbook(
                    workbook,
                    incoming,
                    interactive=interactive,
                    include_sections=include_sections,
                )
                for sheet_name, status, row, changes in results:
                    counts[status] += 1
                    report_rows.append(
                        {
                            "person": incoming["descriptive_name"],
                            "rat_year": incoming.get("rat_year"),
                            "workbook": workbook_paths[label].name,
                            "worksheet": sheet_name,
                            "status": status,
                            "row": row,
                            "changes": "; ".join(changes),
                        }
                    )
            if index % 25 == 0:
                print(f"Processed {index}/{len(records)} cleaned response rows...")
    except UserQuit as exc:
        print(f"Stopped: {exc}")
        for wb in workbooks.values():
            wb.close()
        return 130

    print("\nPlanned changes")
    print(f"  Added rows: {counts['added']}")
    print(f"  Updated rows: {counts['updated']}")
    print(f"  Unchanged matches: {counts['unchanged']}")

    save_changes = not args.dry_run
    if save_changes and not args.yes and interactive:
        answer = ask("Save these workbook changes? (y/n)", default="y").casefold()
        if answer not in {"y", "yes"}:
            save_changes = False

    backup_dir = timestamped_backup_directory(project_root, "before_form_import")
    if save_changes:
        for label, workbook in workbooks.items():
            path = workbook_paths[label]
            backup_file(path, backup_dir)
            save_workbook_atomic(workbook, path)
            print(f"Saved: {path}")
        print(f"Backups: {backup_dir}")
    else:
        print("No workbook files were written.")

    for workbook in workbooks.values():
        workbook.close()

    report_dir = project_root / "import_reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = report_dir / f"form_import_{stamp}.csv"
    with report_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["person", "rat_year", "workbook", "worksheet", "status", "row", "changes"],
        )
        writer.writeheader()
        writer.writerows(report_rows)
    print(f"Import report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
