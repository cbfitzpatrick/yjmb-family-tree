#!/usr/bin/env python3
"""Apply one decrypted protected YJMB update to the authoritative workbook.

v18 rules:
- Normal authenticated additions/corrections apply without an admin approval step.
- Add Yourself merges into a unique existing same-name/same-year profile instead
  of creating a duplicate. Newly referenced VETs/RATs are created as rows/cards
  and reciprocated when they do not already exist. Existing related profiles are
  never silently overwritten.
- When a person's canonical relationship identity changes, relationship cells
  that uniquely resolve to that person are rewritten to the new canonical text.
- Every caller receives a cell-level before/after change set so the queue
  processor can write an encrypted, revertible changelog.
- Missing additive compatibility columns are created automatically. Existing columns are
  never renamed and unrelated cells are never rewritten.
"""
from __future__ import annotations

import argparse
import copy
import json
import os
import re
import unicodedata
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from yjmb_schema import ensure_optional_columns, find_header_row, header_key
from yjmb_taxonomy import SECTION_DISPLAY, canonical_formal_roles, canonical_section_entry, informal_roles_from_text, recognized_sections


class ReviewRequired(Exception):
    pass


def norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def h(value: object) -> str:
    return header_key(value)


def namekey(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", unicodedata.normalize("NFKD", norm(value).casefold()))


def json_cell(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def safe_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Member/admin web forms are data entry surfaces, not formula editors.
    if text.startswith("="):
        raise ReviewRequired("Workbook formulas cannot be submitted through the website editor.")
    return text


def relation_section(value: object) -> str:
    raw = norm(value)
    direct = SECTION_DISPLAY.get(raw.casefold())
    if direct:
        return direct
    sections = recognized_sections(raw)
    if len(sections) == 1:
        return SECTION_DISPLAY.get(sections[0], raw)
    return raw


def relation(name: object, year: object, section: object) -> str:
    try:
        year_text = str(int(year))
    except (TypeError, ValueError):
        year_text = norm(year)
    return f"{norm(name)} ({year_text}) ({relation_section(section)})"


def section_entry_text(entry: dict[str, Any] | None) -> str:
    section = norm((entry or {}).get("section"))
    if not section:
        return ""
    broad = SECTION_DISPLAY.get(section.casefold(), section)
    detail = norm((entry or {}).get("specificInstrument"))
    if not detail:
        return broad
    return canonical_section_entry(section, detail) or f"{broad} — {detail}"


def parse_row_id(value: object) -> int | None:
    match = re.fullmatch(r"row-(\d+)", norm(value))
    return int(match.group(1)) if match else None


def discover(ws):
    header_row = find_header_row(ws)
    headers = {col: h(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)}
    aliases = {
        "given": ["givenpreferredname"],
        "nickname": ["nickname"],
        "family": ["familymaidenname"],
        "married": ["marriedname"],
        "year": ["ratyear", "year"],
        "instrument": ["instrument", "section"],
        "vet": ["vet"],
        "display": ["treedisplaynamepreference"],
        "lastDisplay": ["treedisplaylastnamepreference"],
        "sectionNick": ["sectionnicknames"],
        "specific": ["specificinstruments"],
        "memory": ["favoritetechbandmemory"],
        "otherFlag": ["participatedinothergtensembles"],
        "otherList": ["othergtensembles"],
        "otherInstFlag": ["playeddifferentinstrumentinothergtensembles"],
        "otherInst": ["othergtensembleinstruments"],
        "leadership": ["marchingbandleadershiproles"],
        "leadershipHistory": ["marchingbandleadershiphistory"],
        "informalFlag": ["servedininformalleadershipposition"],
        "informal": ["informalleadershippositions"],
        "bandClubFlag": ["servedinbandclubleadershipposition"],
        "bandClub": ["bandclubleadershippositions"],
        "bandClubHistory": ["bandclubleadershiphistory"],
        "leadershipClass": ["leadershippositionclassification"],
        "hasNick": ["hasnickname"],
        "changed": ["changedlastnamesinceband"],
        "multi": ["hasbeeninmultiplesections"],
        "currentRat": ["currentlyarat"],
        "pair": ["ratvetpairsystemapplied"],
    }
    mapping: dict[str, int] = {}
    for key, values in aliases.items():
        for col, value in headers.items():
            if value in values:
                mapping[key] = col
                break
    required = ["given", "nickname", "family", "married", "year", "instrument", "vet"]
    missing = [key for key in required if key not in mapping]
    if missing:
        raise RuntimeError("Missing required columns: " + ", ".join(missing))
    rats: list[tuple[int, int]] = []
    for col, value in headers.items():
        match = re.fullmatch(r"rat(\d+)", value)
        if match:
            rats.append((int(match.group(1)), col))
    rats.sort()
    if not rats:
        raise RuntimeError("No RAT columns found.")
    labels = {col: norm(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)}
    label_to_col = {h(label): col for col, label in labels.items() if label}
    return header_row, mapping, rats, labels, label_to_col


def row_name(ws, row: int, mapping: dict[str, int]) -> str:
    return norm(f"{ws.cell(row, mapping['given']).value or ''} {ws.cell(row, mapping['family']).value or ''}")


def row_year(ws, row: int, mapping: dict[str, int]) -> int | None:
    try:
        return int(ws.cell(row, mapping["year"]).value)
    except (TypeError, ValueError):
        return None


def append_style_row(ws, header_row: int) -> int:
    old_max_row = ws.max_row
    row = old_max_row + 1
    source = max(header_row + 1, old_max_row)
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source, col)
        dst = ws.cell(row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
        dst.number_format = src.number_format
    for table in ws.tables.values():
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        except ValueError:
            continue
        if min_row <= header_row <= max_row and max_row == old_max_row:
            table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{row}"
    if ws.auto_filter and ws.auto_filter.ref:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ws.auto_filter.ref)
            if max_row == old_max_row:
                ws.auto_filter.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{row}"
        except ValueError:
            pass
    return row


def change_cell(ws, row: int, col: int, value: Any, label: str, changes: list[dict[str, Any]]) -> None:
    value = safe_cell_value(value)
    cell = ws.cell(row, col)
    before = cell.value
    if before == value or (norm(before) == norm(value) and not isinstance(value, (int, float, bool))):
        return
    cell.value = value
    changes.append({"row": row, "label": label, "before": json_cell(before), "after": json_cell(value)})


def set_mapped(ws, row: int, mapping: dict[str, int], labels: dict[int, str], key: str, value: Any, changes: list[dict[str, Any]]) -> None:
    col = mapping.get(key)
    if col:
        change_cell(ws, row, col, value, labels[col], changes)


def note_column(label_to_col: dict[str, int]) -> int | None:
    for key in ("notes", "note", "comments", "comment"):
        if key in label_to_col:
            return label_to_col[key]
    return None


def append_note(ws, row: int, col: int | None, note: str, label: str, changes: list[dict[str, Any]]) -> None:
    note = norm(note)
    if not col or not note:
        return
    existing = norm(ws.cell(row, col).value)
    if existing and note.casefold() in existing.casefold():
        return
    value = note if not existing else f"{existing}\n{note}"
    change_cell(ws, row, col, value, label, changes)


def display_preference(value: object) -> str:
    key = norm(value).casefold()
    if key == "nickname":
        return "Nickname"
    if key == "both":
        return "Both"
    return "Given/Preferred Name"


def last_name_display_preference(value: object) -> str:
    key = h(value)
    if key in {"married", "marriedname", "current", "currentname"}:
        return "Married Name"
    if key in {"both", "maidenandmarried", "familyandmarried", "bothlastnames"}:
        return "Both"
    return "Maiden/Family Name"


RELATION_EDIT_RE = re.compile(r"^\s*(.*?)\s*\(((?:19|20)\d{2})\)\s*\(([^()]*)\)\s*$")


def parse_relation_edit(value: object) -> tuple[str, int, str] | None:
    raw = norm(value)
    if not raw:
        return None
    match = RELATION_EDIT_RE.fullmatch(raw)
    if not match:
        return None
    name = norm(match.group(1))
    year = int(match.group(2))
    section = norm(match.group(3))
    if not name or not section:
        return None
    return name, year, section


def split_person_name(value: object) -> tuple[str, str]:
    """Conservatively split an externally referenced person's full name."""
    parts = norm(value).split()
    if len(parts) < 2:
        raise ReviewRequired(
            "A new RAT reference needs at least a first/preferred name and family name "
            "before a person row can be created automatically."
        )
    prefixes = {"de", "del", "van", "von", "la", "le", "st.", "st", "saint"}
    if len(parts) >= 3 and parts[-2].casefold() in prefixes:
        return " ".join(parts[:-2]), " ".join(parts[-2:])
    return " ".join(parts[:-1]), parts[-1]


def rows_matching_person(ws, header_row: int, mapping: dict[str, int], name: str, year: int) -> list[int]:
    target = namekey(name)
    return [
        row
        for row in range(header_row + 1, ws.max_row + 1)
        if namekey(row_name(ws, row, mapping)) == target and row_year(ws, row, mapping) == year
    ]


def relationship_columns(mapping: dict[str, int], rat_cols: list[tuple[int, int]]) -> list[int]:
    return [mapping["vet"], *(col for _, col in rat_cols)]


def relationship_locations_resolving_to(
    ws, header_row: int, target_row: int, mapping: dict[str, int], rat_cols: list[tuple[int, int]]
) -> list[tuple[int, int]]:
    """Return relationship cells that uniquely identify ``target_row``.

    Name + RAT year is normally sufficient. If duplicate same-name/same-year
    rows exist, the relationship's section is used as a conservative
    disambiguator. Ambiguous cells are intentionally left untouched.
    """
    locations: list[tuple[int, int]] = []
    for source_row in range(header_row + 1, ws.max_row + 1):
        for col in relationship_columns(mapping, rat_cols):
            parsed = parse_relation_edit(ws.cell(source_row, col).value)
            if not parsed:
                continue
            ref_name, ref_year, ref_section = parsed
            candidates = rows_matching_person(ws, header_row, mapping, ref_name, ref_year)
            if len(candidates) > 1:
                ref_sections = set(recognized_sections(ref_section))
                if ref_sections:
                    by_section = [
                        row for row in candidates
                        if ref_sections & set(recognized_sections(ws.cell(row, mapping["instrument"]).value))
                    ]
                    if by_section:
                        candidates = by_section
            if candidates == [target_row]:
                locations.append((source_row, col))
    return locations


def propagate_person_relationship_identity(
    ws, header_row: int, target_row: int, mapping: dict[str, int], rat_cols: list[tuple[int, int]],
    labels: dict[int, str], locations: list[tuple[int, int]], old_relation: str, changes: list[dict[str, Any]]
) -> int:
    """Rewrite incoming VET/RAT references after a person's identity changes."""
    new_relation = person_relation(ws, target_row, mapping)
    if norm(new_relation) == norm(old_relation):
        return 0
    updated = 0
    for source_row, col in locations:
        before_count = len(changes)
        change_cell(ws, source_row, col, new_relation, labels[col], changes)
        if len(changes) > before_count:
            updated += 1
    return updated


def split_list_text(value: object) -> list[str]:
    return [norm(part) for part in re.split(r"\s*;\s*", norm(value)) if norm(part)]


def merge_list_text(existing: object, additions: list[str]) -> str | None:
    values = split_list_text(existing)
    seen = {norm(value).casefold() for value in values}
    for value in additions:
        clean = norm(value)
        if clean and clean.casefold() not in seen:
            values.append(clean)
            seen.add(clean.casefold())
    return "; ".join(values) or None


def merge_free_text(existing: object, addition: object) -> str | None:
    current = norm(existing)
    extra = norm(addition)
    if not extra:
        return current or None
    if not current:
        return extra
    if extra.casefold() in current.casefold():
        return current
    return f"{current}\n\n{extra}"


def merge_yes(existing: object, submitted_yes: bool) -> str:
    return "Yes" if submitted_yes or norm(existing).casefold() in {"yes", "y", "true", "1"} else "No"


def merge_section_entries(existing: object, additions: list[str]) -> str | None:
    """Merge submitted section memberships without throwing away old subsections."""
    result = split_list_text(existing)
    for addition in additions:
        clean = norm(addition)
        if not clean:
            continue
        new_sections = recognized_sections(clean)
        new_key = new_sections[0] if len(new_sections) == 1 else None
        replaced = False
        if new_key:
            for index, current in enumerate(result):
                current_sections = recognized_sections(current)
                if len(current_sections) == 1 and current_sections[0] == new_key:
                    # Prefer the more informative representation for the same broad section.
                    if len(clean) > len(current) or "—" in clean:
                        result[index] = clean
                    replaced = True
                    break
        if not replaced and clean.casefold() not in {value.casefold() for value in result}:
            result.append(clean)
    return "; ".join(result) or None


def relation_identity(value: object) -> tuple[str, int] | None:
    parsed = parse_relation_edit(value)
    if not parsed:
        return None
    name, year, _ = parsed
    return namekey(name), year


def append_rat_reference(ws, row: int, rat_cols, labels, relationship: str, changes) -> None:
    wanted = relation_identity(relationship)
    for _, col in rat_cols:
        current = norm(ws.cell(row, col).value)
        if current and wanted and relation_identity(current) == wanted:
            change_cell(ws, row, col, relationship, labels[col], changes)
            return
    for _, col in rat_cols:
        if not norm(ws.cell(row, col).value):
            change_cell(ws, row, col, relationship, labels[col], changes)
            return
    raise ReviewRequired(f"Row {row} has no blank RAT slot for reciprocal relationship creation.")


def insert_rat_reference_first(ws, row: int, rat_cols, labels, relationship: str, changes) -> None:
    """Put a reciprocal RAT in RAT 1 and shift the existing ordered RATs down.

    If the relationship already exists in a later RAT slot it is moved to RAT 1
    rather than duplicated.  No existing RAT is discarded; a completely full
    row is held for review before insertion.
    """
    wanted = relation_identity(relationship)
    existing: list[str] = []
    for _, col in rat_cols:
        current = norm(ws.cell(row, col).value)
        if not current:
            continue
        if wanted and relation_identity(current) == wanted:
            continue
        existing.append(current)

    if len(existing) >= len(rat_cols):
        raise ReviewRequired(
            f"Row {row} has no RAT slot available to insert the newly reciprocated person as RAT 1."
        )

    ordered = [relationship, *existing]
    for index, (_, col) in enumerate(rat_cols):
        value = ordered[index] if index < len(ordered) else None
        change_cell(ws, row, col, value, labels[col], changes)


def ensure_relationship_reciprocal(
    ws, source_row: int, target_row: int, role: str, mapping, rat_cols, labels, changes
) -> None:
    """Reciprocate a submitted relationship against an existing or new row.

    A submitted VET means ``target_row`` is the source person's VET, so the
    source must appear as RAT 1 on that target (existing RATs shift down).
    A submitted RAT means ``target_row`` is the source person's RAT, so the
    target's VET is filled when blank or canonicalized when already equivalent.
    """
    source_relationship = person_relation(ws, source_row, mapping)
    if role.upper() == "VET":
        insert_rat_reference_first(ws, target_row, rat_cols, labels, source_relationship, changes)
        return
    if role.upper() == "RAT":
        vet_col = mapping["vet"]
        current = norm(ws.cell(target_row, vet_col).value)
        if not current or relation_identity(current) == relation_identity(source_relationship):
            change_cell(ws, target_row, vet_col, source_relationship, labels[vet_col], changes)
            return
        raise ReviewRequired(
            f"Row {target_row} already lists a different VET; the RAT relationship was not overwritten automatically."
        )
    raise ReviewRequired(f"Unsupported relationship role for reciprocity: {role}")


def create_person_from_relation(
    ws, header_row, mapping, rat_cols, labels, *, source_row: int, role: str,
    relationship: object, changes
) -> int | None:
    """Create an externally referenced person only when no name/year row exists.

    New people are reciprocated immediately because the newly created row cannot
    contain conflicting relationship data. Existing people are never duplicated
    or silently overwritten.
    """
    parsed = parse_relation_edit(relationship)
    if not parsed:
        raise ReviewRequired(
            f"{role} must use Name (RAT Year) (Section) so a missing person can be created safely."
        )
    name, year, section = parsed
    matches = rows_matching_person(ws, header_row, mapping, name, year)
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise ReviewRequired(
            f"{role} {name} ({year}) matches multiple existing rows. "
            "No duplicate person was created; resolve the existing identity first."
        )

    given, family = split_person_name(name)
    new_row = append_style_row(ws, header_row)
    set_mapped(ws, new_row, mapping, labels, "given", given, changes)
    set_mapped(ws, new_row, mapping, labels, "family", family, changes)
    set_mapped(ws, new_row, mapping, labels, "year", year, changes)
    instrument = canonical_section_entry(section) or relation_section(section)
    set_mapped(ws, new_row, mapping, labels, "instrument", instrument, changes)
    set_mapped(ws, new_row, mapping, labels, "display", "Given/Preferred Name", changes)
    set_mapped(ws, new_row, mapping, labels, "lastDisplay", "Maiden/Family Name", changes)
    set_mapped(ws, new_row, mapping, labels, "hasNick", "No", changes)
    set_mapped(ws, new_row, mapping, labels, "changed", "No", changes)
    set_mapped(ws, new_row, mapping, labels, "pair", "Yes", changes)

    source_relationship = person_relation(ws, source_row, mapping)
    if role.upper() == "RAT":
        change_cell(ws, new_row, mapping["vet"], source_relationship, labels[mapping["vet"]], changes)
    elif role.upper() == "VET":
        append_rat_reference(ws, new_row, rat_cols, labels, source_relationship, changes)
    else:
        raise ReviewRequired(f"Unsupported relationship role for new person: {role}")
    return new_row


def create_missing_people_from_field_changes(
    ws, header_row, source_row, mapping, rat_cols, labels, change_payload, changes
) -> list[int]:
    """Ensure every changed correction/admin VET or RAT resolves to a real row.

    If no person with that name/RAT-year identity exists, create the person and
    reciprocate the new relationship immediately.  Ambiguous existing identities
    are held for review rather than leaving a dangling text-only relationship.
    """
    created: list[int] = []
    if not isinstance(change_payload, list):
        return created
    for item in change_payload:
        if not isinstance(item, dict):
            continue
        label = norm(item.get("label"))
        role = "VET" if label.casefold() == "vet" else ("RAT" if re.fullmatch(r"RAT\s+\d+", label, re.I) else "")
        if not role:
            continue
        after = norm(item.get("after"))
        if not after:
            continue
        before_rows = ws.max_row
        target = create_person_from_relation(
            ws, header_row, mapping, rat_cols, labels,
            source_row=source_row, role=role, relationship=after, changes=changes,
        )
        if target:
            ensure_relationship_reciprocal(
                ws, source_row, target, role, mapping, rat_cols, labels, changes
            )
        if target and target > before_rows:
            created.append(target)
    return created


def merge_relation_claims(ws, row, mapping, rat_cols, labels, payload, changes) -> list[tuple[str, str, int | None]]:
    """Merge submitted VET/RAT claims into an existing profile without erasing old claims."""
    touched: list[tuple[str, str, int | None]] = []
    vet = payload.get("vet") or None
    if isinstance(vet, dict) and norm(vet.get("name")):
        wanted = relation(vet.get("name"), vet.get("year"), vet.get("section"))
        current = norm(ws.cell(row, mapping["vet"]).value)
        if current and relation_identity(current) != relation_identity(wanted):
            raise ReviewRequired(
                "The existing profile already has a different VET. The Add Yourself merge was held instead of overwriting it."
            )
        change_cell(ws, row, mapping["vet"], wanted, labels[mapping["vet"]], changes)
        touched.append(("VET", wanted, parse_row_id(vet.get("matchedId"))))

    rats = payload.get("rats") or []
    if not isinstance(rats, list):
        rats = []
    for item in rats:
        if not isinstance(item, dict) or not norm(item.get("name")):
            continue
        wanted = relation(item.get("name"), item.get("year"), item.get("section"))
        identity = relation_identity(wanted)
        target_col = None
        for _, col in rat_cols:
            current = norm(ws.cell(row, col).value)
            if current and identity and relation_identity(current) == identity:
                target_col = col
                break
        if target_col is None:
            for _, col in rat_cols:
                if not norm(ws.cell(row, col).value):
                    target_col = col
                    break
        if target_col is None:
            raise ReviewRequired("The existing profile has no blank RAT slot for all submitted RATs.")
        change_cell(ws, row, target_col, wanted, labels[target_col], changes)
        touched.append(("RAT", wanted, parse_row_id(item.get("matchedId"))))
    return touched


def create_missing_people_for_claims(
    ws, header_row, source_row, mapping, rat_cols, labels,
    claims: list[tuple[str, str, int | None]], changes
) -> dict[str, int]:
    created: dict[str, int] = {}
    for role, raw, matched_row in claims:
        target = None
        if matched_row and header_row < matched_row <= ws.max_row:
            # Do not trust a stale browser row id blindly.  It must still identify
            # the same name/year relationship in the authoritative workbook.
            if relation_identity(person_relation(ws, matched_row, mapping)) == relation_identity(raw):
                target = matched_row
        before = ws.max_row
        if target is None:
            target = create_person_from_relation(
                ws, header_row, mapping, rat_cols, labels,
                source_row=source_row, role=role, relationship=raw, changes=changes,
            )
        if target:
            ensure_relationship_reciprocal(
                ws, source_row, target, role, mapping, rat_cols, labels, changes
            )
        if target and target > before:
            created[f"{role}:{relation_identity(raw)}"] = target
    return created

def leadership_values(self_data: dict[str, Any]) -> tuple[list[str], str, str, str, str, str]:
    history = self_data.get("leadershipHistory") or []
    if not isinstance(history, list):
        history = []
    if not history:
        # Compatibility with the v15/v16 questionnaire payload.
        for role in self_data.get("marchingBandLeadershipRoles") or []:
            history.append({"type": "formal", "role": role, "years": ""})
        if self_data.get("informalLeadership") and norm(self_data.get("informalLeadershipDescription")):
            history.append({"type": "informal", "role": norm(self_data.get("informalLeadershipDescription")), "years": ""})

    formal_entries: list[dict[str, str]] = []
    informal_entries: list[dict[str, str]] = []
    for item in history:
        if not isinstance(item, dict):
            continue
        role = norm(item.get("role"))
        years = norm(item.get("years"))
        if not role:
            continue
        entry = {"role": role, "years": years}
        if norm(item.get("type")).casefold() == "informal":
            informal_entries.append(entry)
        else:
            formal_entries.append(entry)

    formal_roles = canonical_formal_roles(", ".join(item["role"] for item in formal_entries))
    # Preserve Other formal answers that taxonomy intentionally cannot canonicalize.
    for item in formal_entries:
        if item["role"] not in formal_roles and not canonical_formal_roles(item["role"]):
            formal_roles.append(item["role"])

    def fmt(item: dict[str, str]) -> str:
        return f"{item['role']} ({item['years']})" if item["years"] else item["role"]

    history_text = "; ".join(
        [*(f"Formal: {fmt(item)}" for item in formal_entries), *(f"Informal: {fmt(item)}" for item in informal_entries)]
    )
    informal_text = "; ".join(fmt(item) for item in informal_entries)
    informal_classified = []
    for item in informal_entries:
        values = informal_roles_from_text(item["role"])
        informal_classified.extend(values or [item["role"]])
    classification = "; ".join(
        [*(f"Formal: {role}" for role in formal_roles), *(f"Informal: {role}" for role in dict.fromkeys(informal_classified))]
    )

    club_history = self_data.get("bandClubLeadershipHistory") or []
    if not isinstance(club_history, list):
        club_history = []
    club_items = []
    for item in club_history:
        if not isinstance(item, dict):
            continue
        position = norm(item.get("position"))
        years = norm(item.get("years"))
        if position:
            club_items.append({"position": position, "years": years})
    club_text = "; ".join(f"{x['position']} ({x['years']})" if x["years"] else x["position"] for x in club_items)
    club_history_text = "; ".join(f"{x['position']} ({x['years']})" if x["years"] else x["position"] for x in club_items)
    return formal_roles, history_text, informal_text, classification, club_text, club_history_text


def apply_addition(ws, header_row, mapping, rat_cols, labels, label_to_col, payload, changes):
    self_data = payload.get("self") or {}
    given = norm(self_data.get("givenPreferredName"))
    family = norm(self_data.get("familyMaidenName"))
    try:
        year = int(self_data.get("ratYear"))
    except (TypeError, ValueError):
        raise ReviewRequired("Submitter RAT year is missing or invalid.")
    if not given or not family:
        raise ReviewRequired("Submitter name is incomplete.")

    matches = rows_matching_person(ws, header_row, mapping, f"{given} {family}", year)
    requested_existing = parse_row_id(payload.get("existingPersonId"))
    if requested_existing and requested_existing not in matches:
        requested_existing = None
    if len(matches) > 1 and requested_existing is None:
        raise ReviewRequired(
            "More than one same-name/same-year profile exists. An administrator must choose which row receives this Add Yourself merge."
        )
    existing_row = requested_existing or (matches[0] if len(matches) == 1 else None)
    merging = existing_row is not None
    row = existing_row if existing_row is not None else append_style_row(ws, header_row)

    old_relation = person_relation(ws, row, mapping) if merging else ""
    incoming_locations = relationship_locations_resolving_to(ws, header_row, row, mapping, rat_cols) if merging else []

    sections = self_data.get("sections") or []
    section_names = [section_entry_text(item) for item in sections if isinstance(item, dict) and norm(item.get("section"))]
    submitted_instrument = "; ".join(filter(None, section_names)) or relation_section(self_data.get("section"))
    instrument = merge_section_entries(ws.cell(row, mapping["instrument"]).value, section_names or [submitted_instrument]) if merging else submitted_instrument

    # Identity is safe to refresh because the same-name/same-year check selected this row.
    set_mapped(ws, row, mapping, labels, "given", given, changes)
    set_mapped(ws, row, mapping, labels, "family", family, changes)
    set_mapped(ws, row, mapping, labels, "year", year, changes)
    set_mapped(ws, row, mapping, labels, "instrument", instrument, changes)

    nickname = norm(self_data.get("nickname"))
    married = norm(self_data.get("marriedName"))
    if nickname or not merging:
        set_mapped(ws, row, mapping, labels, "nickname", nickname or None, changes)
    if married or not merging:
        set_mapped(ws, row, mapping, labels, "married", married or None, changes)

    if self_data.get("hasNickname") or not merging or not norm(ws.cell(row, mapping.get("display", mapping["given"])).value):
        set_mapped(ws, row, mapping, labels, "display", display_preference(self_data.get("treeNamePreference")), changes)
    if self_data.get("changedLastName") or not merging:
        set_mapped(ws, row, mapping, labels, "lastDisplay", last_name_display_preference(self_data.get("lastNamePreference")), changes)

    submitted_section_nicks = [
        f"{relation_section(item.get('section'))}: {norm(item.get('sectionNickname'))}"
        for item in sections if isinstance(item, dict) and norm(item.get("sectionNickname"))
    ]
    submitted_specific = [
        f"{relation_section(item.get('section'))}: {norm(item.get('specificInstrument'))}"
        for item in sections if isinstance(item, dict) and norm(item.get("specificInstrument"))
    ]
    if merging:
        if mapping.get("sectionNick") and submitted_section_nicks:
            set_mapped(ws, row, mapping, labels, "sectionNick", merge_list_text(ws.cell(row, mapping["sectionNick"]).value, submitted_section_nicks), changes)
        if mapping.get("specific") and submitted_specific:
            set_mapped(ws, row, mapping, labels, "specific", merge_list_text(ws.cell(row, mapping["specific"]).value, submitted_specific), changes)
    else:
        set_mapped(ws, row, mapping, labels, "sectionNick", "; ".join(submitted_section_nicks) or None, changes)
        set_mapped(ws, row, mapping, labels, "specific", "; ".join(submitted_specific) or None, changes)

    memory = norm(payload.get("favoriteTechBandMemory"))
    if merging and memory and mapping.get("memory"):
        set_mapped(ws, row, mapping, labels, "memory", merge_free_text(ws.cell(row, mapping["memory"]).value, memory), changes)
    elif not merging:
        set_mapped(ws, row, mapping, labels, "memory", memory or None, changes)

    # Free-text/list fields are additive during an existing-profile merge.
    other_list = norm(self_data.get("otherGtEnsemblesList"))
    other_inst = norm(self_data.get("otherGtInstruments"))
    if merging:
        if other_list and mapping.get("otherList"):
            set_mapped(ws, row, mapping, labels, "otherList", merge_list_text(ws.cell(row, mapping["otherList"]).value, [other_list]), changes)
        if other_inst and mapping.get("otherInst"):
            set_mapped(ws, row, mapping, labels, "otherInst", merge_list_text(ws.cell(row, mapping["otherInst"]).value, [other_inst]), changes)
    else:
        set_mapped(ws, row, mapping, labels, "otherList", other_list or None, changes)
        set_mapped(ws, row, mapping, labels, "otherInst", other_inst or None, changes)

    formal_roles, leadership_history, informal_text, classification, club_text, club_history = leadership_values(self_data)
    if merging:
        if mapping.get("leadership") and formal_roles:
            existing_roles = canonical_formal_roles(ws.cell(row, mapping["leadership"]).value)
            set_mapped(ws, row, mapping, labels, "leadership", ", ".join(dict.fromkeys(existing_roles + formal_roles)), changes)
        for key_name, submitted in (
            ("leadershipHistory", leadership_history), ("informal", informal_text),
            ("leadershipClass", classification), ("bandClub", club_text), ("bandClubHistory", club_history),
        ):
            col = mapping.get(key_name)
            if col and submitted:
                set_mapped(ws, row, mapping, labels, key_name, merge_list_text(ws.cell(row, col).value, split_list_text(submitted)), changes)
    else:
        set_mapped(ws, row, mapping, labels, "leadership", ", ".join(formal_roles) or None, changes)
        set_mapped(ws, row, mapping, labels, "leadershipHistory", leadership_history or None, changes)
        set_mapped(ws, row, mapping, labels, "informal", informal_text or None, changes)
        set_mapped(ws, row, mapping, labels, "leadershipClass", classification or None, changes)
        set_mapped(ws, row, mapping, labels, "bandClub", club_text or None, changes)
        set_mapped(ws, row, mapping, labels, "bandClubHistory", club_history or None, changes)

    bool_specs = (
        ("otherFlag", bool(self_data.get("otherGtEnsembles"))),
        ("otherInstFlag", bool(self_data.get("playedDifferentGtInstrument"))),
        ("informalFlag", bool(informal_text)),
        ("bandClubFlag", bool(club_text or self_data.get("bandClubLeadership"))),
        ("hasNick", bool(self_data.get("hasNickname"))),
        ("changed", bool(self_data.get("changedLastName"))),
        ("multi", bool(self_data.get("multipleSections"))),
        ("currentRat", bool(self_data.get("currentlyRat"))),
        ("pair", bool((payload.get("pairSystem") or {}).get("applies"))),
    )
    for key_name, submitted_yes in bool_specs:
        col = mapping.get(key_name)
        if not col:
            continue
        value = merge_yes(ws.cell(row, col).value, submitted_yes) if merging else ("Yes" if submitted_yes else "No")
        set_mapped(ws, row, mapping, labels, key_name, value, changes)

    claims: list[tuple[str, str, int | None]] = []
    if merging:
        claims = merge_relation_claims(ws, row, mapping, rat_cols, labels, payload, changes)
    else:
        vet = payload.get("vet") or None
        if isinstance(vet, dict) and norm(vet.get("name")):
            raw = relation(vet.get("name"), vet.get("year"), vet.get("section"))
            change_cell(ws, row, mapping["vet"], raw, labels[mapping["vet"]], changes)
            claims.append(("VET", raw, parse_row_id(vet.get("matchedId"))))
        rats = payload.get("rats") or []
        if len(rats) > len(rat_cols):
            raise ReviewRequired("Submission has more RATs than the workbook has RAT columns.")
        for index, item in enumerate(rats):
            if not isinstance(item, dict) or not norm(item.get("name")):
                continue
            col = rat_cols[index][1]
            raw = relation(item.get("name"), item.get("year"), item.get("section"))
            change_cell(ws, row, col, raw, labels[col], changes)
            claims.append(("RAT", raw, parse_row_id(item.get("matchedId"))))

    created_by_claim = create_missing_people_for_claims(
        ws, header_row, row, mapping, rat_cols, labels, claims, changes
    )

    notes = payload.get("notes") or {}
    ncol = note_column(label_to_col)
    if isinstance(notes, dict):
        append_note(ws, row, ncol, notes.get("self", ""), labels.get(ncol, "Notes") if ncol else "Notes", changes)
        related_by_key: dict[str, int] = {}
        vet = payload.get("vet") or None
        if isinstance(vet, dict):
            vet_row = parse_row_id(vet.get("matchedId"))
            if not vet_row and norm(vet.get("name")):
                ident = relation_identity(relation(vet.get("name"), vet.get("year"), vet.get("section")))
                matches2 = rows_matching_person(ws, header_row, mapping, vet.get("name"), int(vet.get("year"))) if ident else []
                vet_row = matches2[0] if len(matches2) == 1 else None
            if vet_row:
                related_by_key["vet"] = vet_row
        for item in payload.get("rats") or []:
            if not isinstance(item, dict):
                continue
            rid = parse_row_id(item.get("matchedId"))
            if not rid and norm(item.get("name")):
                try:
                    matches2 = rows_matching_person(ws, header_row, mapping, item.get("name"), int(item.get("year")))
                except (TypeError, ValueError):
                    matches2 = []
                rid = matches2[0] if len(matches2) == 1 else None
            if rid and item.get("key"):
                related_by_key[str(item["key"])] = rid
        for key_name, note in notes.items():
            target_row = related_by_key.get(str(key_name))
            if target_row and header_row < target_row <= ws.max_row:
                append_note(ws, target_row, ncol, note, labels.get(ncol, "Notes") if ncol else "Notes", changes)

    propagated = 0
    if merging:
        propagated = propagate_person_relationship_identity(
            ws, header_row, row, mapping, rat_cols, labels, incoming_locations, old_relation, changes
        )
    created_count = len(created_by_claim)
    if merging:
        suffix = []
        if created_count:
            suffix.append(f"created/reciprocated {created_count} missing related person row(s)")
        if propagated:
            suffix.append(f"updated {propagated} incoming relationship reference(s)")
        return row, f"Merged Add Yourself submission into {given} {family} ({year})" + (" and " + "; ".join(suffix) if suffix else "")
    return row, f"Added {given} {family} ({year})" + (f" and created/reciprocated {created_count} missing related person row(s)" if created_count else "")

def row_for_id(ws, header_row: int, person_id: object) -> int:
    row = parse_row_id(person_id)
    if not row or not (header_row < row <= ws.max_row):
        raise ReviewRequired(f"Person record {norm(person_id) or '[missing]'} no longer exists.")
    return row


def apply_field_changes(ws, row, changes_payload, labels, label_to_col, changes, *, stale_check: bool = True):
    if not isinstance(changes_payload, list) or not changes_payload:
        raise ReviewRequired("No field changes were supplied.")
    for item in changes_payload:
        if not isinstance(item, dict):
            continue
        label = norm(item.get("label"))
        col = label_to_col.get(h(label))
        if not label or not col:
            raise ReviewRequired(f"Workbook field {label!r} does not exist.")
        current = ws.cell(row, col).value
        if stale_check and "before" in item and norm(current) != norm(item.get("before")):
            raise ReviewRequired(f"{label} changed after this edit form was loaded; reload before overwriting it.")
        change_cell(ws, row, col, item.get("after"), labels[col], changes)


def apply_correction(ws, header_row, mapping, rat_cols, labels, label_to_col, payload, changes):
    row = row_for_id(ws, header_row, payload.get("personId"))
    old_relation = person_relation(ws, row, mapping)
    incoming_locations = relationship_locations_resolving_to(ws, header_row, row, mapping, rat_cols)
    change_payload = payload.get("changes")
    apply_field_changes(ws, row, change_payload, labels, label_to_col, changes, stale_check=False)
    propagated = propagate_person_relationship_identity(
        ws, header_row, row, mapping, rat_cols, labels, incoming_locations, old_relation, changes
    )
    created = create_missing_people_from_field_changes(
        ws, header_row, row, mapping, rat_cols, labels, change_payload, changes
    )
    suffixes = []
    if created:
        suffixes.append(f"created/reciprocated {len(created)} missing related person row(s)")
    if propagated:
        suffixes.append(f"updated {propagated} incoming relationship reference(s)")
    suffix = " and " + "; ".join(suffixes) if suffixes else ""
    return row, f"Updated {norm(payload.get('personId'))}{suffix}"

def apply_admin_patch(ws, header_row, mapping, rat_cols, labels, label_to_col, payload, changes):
    row = row_for_id(ws, header_row, payload.get("personId"))
    old_relation = person_relation(ws, row, mapping)
    incoming_locations = relationship_locations_resolving_to(ws, header_row, row, mapping, rat_cols)
    change_payload = payload.get("changes")
    apply_field_changes(ws, row, change_payload, labels, label_to_col, changes, stale_check=False)
    propagated = propagate_person_relationship_identity(
        ws, header_row, row, mapping, rat_cols, labels, incoming_locations, old_relation, changes
    )
    created = create_missing_people_from_field_changes(
        ws, header_row, row, mapping, rat_cols, labels, change_payload, changes
    )
    suffixes = []
    if created:
        suffixes.append(f"created/reciprocated {len(created)} missing related person row(s)")
    if propagated:
        suffixes.append(f"updated {propagated} incoming relationship reference(s)")
    suffix = " and " + "; ".join(suffixes) if suffixes else ""
    return row, f"Admin edited {norm(payload.get('personId'))}{suffix}"

def apply_admin_add(ws, header_row, mapping, rat_cols, labels, label_to_col, payload, changes):
    fields = payload.get("fields") or {}
    if not isinstance(fields, dict):
        raise ReviewRequired("Admin add requires a field mapping.")
    row = append_style_row(ws, header_row)
    for label, value in fields.items():
        col = label_to_col.get(h(label))
        if col:
            change_cell(ws, row, col, value, labels[col], changes)
    if not norm(ws.cell(row, mapping["given"]).value) or not norm(ws.cell(row, mapping["family"]).value):
        raise ReviewRequired("Admin-added rows require Given/Preferred Name and Family/Maiden Name.")
    year = row_year(ws, row, mapping)
    if year is None or not (1908 <= year <= datetime.now().year + 1):
        raise ReviewRequired("Admin-added rows require a valid four-digit RAT Year.")
    relation_changes = [
        {"label": label, "after": value}
        for label, value in fields.items()
        if h(label) == "vet" or re.fullmatch(r"rat\d+", h(label))
    ]
    created = create_missing_people_from_field_changes(
        ws, header_row, row, mapping, rat_cols, labels, relation_changes, changes
    )
    suffix = f" and created/reciprocated {len(created)} missing related person row(s)" if created else ""
    return row, f"Admin added {row_name(ws, row, mapping)}{suffix}"

def apply_admin_delete(ws, header_row, labels, payload, changes):
    row = row_for_id(ws, header_row, payload.get("personId"))
    for col in range(1, ws.max_column + 1):
        if ws.cell(row, col).value is not None:
            change_cell(ws, row, col, None, labels.get(col, f"Column {col}"), changes)
    return row, f"Admin removed row {row}"


def person_relation(ws, row, mapping) -> str:
    return relation(row_name(ws, row, mapping), row_year(ws, row, mapping) or "", ws.cell(row, mapping["instrument"]).value)


def apply_admin_reciprocate(ws, header_row, mapping, rat_cols, labels, payload, changes):
    source_row = row_for_id(ws, header_row, payload.get("sourceId"))
    target_row = row_for_id(ws, header_row, payload.get("targetId"))
    role = norm(payload.get("role")).upper()
    if role == "VET":
        # Source says target is their VET; first verify that claim still exists,
        # then add source as a RAT on target. This prevents a stale admin page
        # from reciprocating a relationship that was edited in the meantime.
        source_claim = norm(ws.cell(source_row, mapping["vet"]).value)
        if not source_claim or namekey(source_claim.split(" (", 1)[0]) != namekey(row_name(ws, target_row, mapping)):
            raise ReviewRequired("The source profile no longer lists the selected target as its VET. Reload Admin Mode before validating.")
        source_rel = person_relation(ws, source_row, mapping)
        source_key = namekey(row_name(ws, source_row, mapping))
        for _, col in rat_cols:
            raw = norm(ws.cell(target_row, col).value)
            if raw and namekey(raw.split(" (", 1)[0]) == source_key:
                change_cell(ws, target_row, col, source_rel, labels[col], changes)
                return target_row, f"Validated reciprocal VET/RAT relationship for rows {source_row} and {target_row}"
        for _, col in rat_cols:
            if not norm(ws.cell(target_row, col).value):
                change_cell(ws, target_row, col, source_rel, labels[col], changes)
                return target_row, f"Validated reciprocal VET/RAT relationship for rows {source_row} and {target_row}"
        raise ReviewRequired(f"Target VET row {target_row} has no blank RAT slot.")
    if role == "RAT":
        # Source says target is their RAT; verify that one of source's RAT slots
        # still names this target before changing the target's VET field.
        target_key = namekey(row_name(ws, target_row, mapping))
        if not any(
            norm(ws.cell(source_row, rat_col).value)
            and namekey(norm(ws.cell(source_row, rat_col).value).split(" (", 1)[0]) == target_key
            for _, rat_col in rat_cols
        ):
            raise ReviewRequired("The source profile no longer lists the selected target as a RAT. Reload Admin Mode before validating.")
        source_rel = person_relation(ws, source_row, mapping)
        col = mapping["vet"]
        existing = norm(ws.cell(target_row, col).value)
        if existing and namekey(existing.split(" (", 1)[0]) != namekey(row_name(ws, source_row, mapping)):
            raise ReviewRequired(f"Target RAT row {target_row} already has a different VET.")
        change_cell(ws, target_row, col, source_rel, labels[col], changes)
        return target_row, f"Validated reciprocal RAT/VET relationship for rows {source_row} and {target_row}"
    raise ReviewRequired("Reciprocal validation role must be VET or RAT.")


def values_equivalent(left: Any, right: Any) -> bool:
    if left == right:
        return True
    return norm(left) == norm(right)


def apply_admin_revert(ws, header_row, labels, label_to_col, payload, changes, changelog_dir: Path):
    change_id = re.sub(r"[^A-Za-z0-9_.-]", "", norm(payload.get("changeId")))
    if not change_id:
        raise ReviewRequired("A changelog ID is required for revert.")
    path = changelog_dir / f"{change_id}.enc.json"
    if not path.exists():
        raise ReviewRequired(f"Changelog entry {change_id} was not found.")
    from secure_submission import decrypt_file
    entry = decrypt_file(path)
    original = entry.get("changes") or []
    if not original:
        raise ReviewRequired("The selected changelog entry has no cell changes to revert.")
    # Fail safely if a later update touched any of these cells.
    for item in original:
        row = int(item.get("row"))
        if not (header_row < row <= ws.max_row):
            raise ReviewRequired(f"Row {row} from the changelog no longer exists.")
        col = label_to_col.get(h(item.get("label")))
        if not col:
            raise ReviewRequired(f"Field {item.get('label')!r} from the changelog no longer exists.")
        if not values_equivalent(ws.cell(row, col).value, item.get("after")):
            raise ReviewRequired(f"Cannot safely revert {item.get('label')} on row {row}: it was changed again later.")
    for item in reversed(original):
        row = int(item.get("row"))
        col = label_to_col[h(item.get("label"))]
        change_cell(ws, row, col, item.get("before"), labels[col], changes)
    return int(original[0].get("row")), f"Reverted changelog entry {change_id}"


def apply(workbook: Path, submission_file: Path, *, changelog_dir: Path = Path("secure/changelog")) -> dict[str, Any]:
    wrapper = json.loads(submission_file.read_text(encoding="utf-8"))
    payload = wrapper.get("payload", wrapper)
    if not isinstance(payload, dict):
        raise ReviewRequired("Submission payload is invalid.")
    kind = norm(payload.get("kind") or wrapper.get("kind") or "addition").casefold()

    wb = load_workbook(workbook)
    try:
        if "People on Tree" not in wb.sheetnames:
            raise RuntimeError("Worksheet People on Tree not found.")
        ws = wb["People on Tree"]
        ensure_optional_columns(ws, header_row=find_header_row(ws))
        header_row, mapping, rat_cols, labels, label_to_col = discover(ws)
        changes: list[dict[str, Any]] = []

        if kind in {"addition", "add", "member-add"}:
            row, summary = apply_addition(ws, header_row, mapping, rat_cols, labels, label_to_col, payload, changes)
        elif kind == "correction":
            row, summary = apply_correction(ws, header_row, mapping, rat_cols, labels, label_to_col, payload, changes)
        elif kind == "admin-patch":
            row, summary = apply_admin_patch(ws, header_row, mapping, rat_cols, labels, label_to_col, payload, changes)
        elif kind == "admin-add":
            row, summary = apply_admin_add(ws, header_row, mapping, rat_cols, labels, label_to_col, payload, changes)
        elif kind == "admin-delete":
            row, summary = apply_admin_delete(ws, header_row, labels, payload, changes)
        elif kind == "admin-reciprocate":
            row, summary = apply_admin_reciprocate(ws, header_row, mapping, rat_cols, labels, payload, changes)
        elif kind == "admin-revert":
            row, summary = apply_admin_revert(ws, header_row, labels, label_to_col, payload, changes, changelog_dir)
        else:
            raise ReviewRequired(f"Unsupported update kind: {kind}")

        resulting_name = row_name(ws, row, mapping) if kind != "admin-delete" else ""
        if not changes:
            # Idempotent submissions are successful, not conflicts. This can happen
            # when the authoritative workbook already contains an edit that the
            # currently deployed encrypted tree has not caught up with yet. Marking
            # the request as a no-op lets the queue processor remove it cleanly and
            # ask the workflow to rebuild/redeploy the public tree from the current
            # authoritative workbook.
            return {
                "row": row,
                "kind": kind,
                "summary": summary,
                "changes": [],
                "name": resulting_name,
                "noop": True,
            }
        tmp = workbook.with_suffix(".secure-update.tmp.xlsx")
        wb.save(tmp)
        wb.close()
        os.replace(tmp, workbook)
        return {
            "row": row,
            "kind": kind,
            "summary": summary,
            "changes": changes,
            "name": resulting_name,
            "noop": False,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", type=Path, required=True)
    ap.add_argument("--submission", type=Path, required=True)
    ap.add_argument("--result", type=Path)
    ap.add_argument("--changelog-dir", type=Path, default=Path("secure/changelog"))
    args = ap.parse_args()
    try:
        result = apply(args.workbook, args.submission, changelog_dir=args.changelog_dir)
        output = {"status": "applied", **result}
        code = 0
    except ReviewRequired as exc:
        output = {"status": "review", "reason": str(exc)}
        code = 20
    if args.result:
        args.result.write_text(json.dumps(output, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    else:
        print(json.dumps(output, indent=2, ensure_ascii=False))
    raise SystemExit(code)


if __name__ == "__main__":
    main()
