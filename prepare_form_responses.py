"""Create a cleaned four-name-column worksheet from a form-response export.

The source worksheet is preserved. ``Cleaned Responses`` uses:
Given/Preferred Name, Nickname, Family/Maiden Name, Married Name.
Ambiguous names, capitalization, spacing, likely typos, and possible married-name
annotations are confirmed with ``Did you mean ...?`` prompts.
"""
from __future__ import annotations
import argparse
from pathlib import Path
import shutil
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from name_tools import InteractiveResolver, SkipRecord, UserQuit, has_ambiguous_annotation, normalize_spaces, normalized_header
from tree_workbook import collect_names_from_workbook, full_name, save_workbook_atomic

CLEANED_SHEET = "Cleaned Responses"
OUTPUT_HEADERS = [
    "Given/Preferred Name", "Nickname", "Family/Maiden Name", "Married Name",
    "RAT Year", "Instrument", "Position and Year", "Notes", "Links", "VET",
    "RAT 1", "RAT 2", "RAT 3", "RAT 4", "RAT 5", "RAT 6", "RAT 7",
    "Timestamp", "Source Row",
]

def find_source_sheet(workbook):
    for ws in workbook.worksheets:
        headers = {normalized_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
        has_name = "name" in headers or "fullname" in headers or "givenpreferredname" in headers
        if "timestamp" in headers and has_name and "ratyear" in headers:
            return ws
    raise ValueError("Could not find a response sheet with Timestamp, a name field, and RAT Year.")

def source_columns(ws):
    normalized = {normalized_header(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if normalize_spaces(ws.cell(1, c).value)}
    def first(prefixes):
        for header, col in normalized.items():
            if any(header == p or header.startswith(p) for p in prefixes): return col
        return None
    cols = {
        "timestamp": first(("timestamp",)), "name": first(("name", "fullname")),
        "given": first(("givenpreferredname", "preferredname", "givenname")),
        "nickname": first(("nickname",)), "family": first(("familymaidenname", "familyname", "maidenname")),
        "married": first(("marriedname", "marriedsurname")), "rat_year": first(("ratyear",)),
        "instrument": first(("instruments", "instrument")), "vet": first(("vetsnameratyearandinstruments", "vet")),
        **{f"rat_{i}": first((f"rat{i}snameratyearandinstruments", f"rat{i}")) for i in range(1, 7)},
        "notes": first(("pleaseincludeanyextrainformationaboutyourtreehere", "notes")),
    }
    if not cols["rat_year"] or not cols["instrument"] or not (cols["name"] or cols["given"]):
        raise ValueError("Missing a required name, RAT Year, or Instrument response column.")
    return {k: v for k, v in cols.items() if v is not None}

def cell_value(ws, row, columns, field):
    col = columns.get(field); return ws.cell(row, col).value if col else None

def normalize_year(raw, *, context, interactive, accept_auto=False):
    text = normalize_spaces(raw)
    for token in text.replace("/", " ").replace("-", " ").replace("'", " ").split():
        digits = "".join(ch for ch in token if ch.isdigit())
        if len(digits) >= 4 and 1900 <= int(digits[:4]) <= 2100:
            return digits[:4]
        if len(digits) == 4 and 1000 <= int(digits) <= 1099:
            proposal = str(int(digits) + 1000)
            if not interactive and accept_auto:
                return proposal
            if interactive:
                answer = input(f'[{context}] Did you mean RAT year "{proposal}" instead of "{text}"? [Y/n]: ').strip().casefold()
                if answer in {"", "y", "yes"}:
                    return proposal
                if answer in {"q", "quit", "exit"}:
                    raise UserQuit("The user stopped the run.")
        if len(digits) == 2:
            proposal = str((1900 if int(digits) >= 70 else 2000) + int(digits))
            if not interactive and accept_auto:
                return proposal
            if interactive:
                answer = input(f'[{context}] Did you mean RAT year "{proposal}" instead of "{text}"? [Y/n]: ').strip().casefold()
                if answer in {"", "y", "yes"}:
                    return proposal
                if answer in {"q", "quit", "exit"}:
                    raise UserQuit("The user stopped the run.")
    if not interactive:
        raise ValueError(f"{context}: invalid RAT year {text!r}")
    while True:
        answer = input(f"[{context}] RAT year {text!r} is unusual. Did you mean which four-digit year? ").strip()
        if answer.lower() in {"q","quit","exit"}: raise UserQuit("The user stopped the run.")
        if len(answer)==4 and answer.isdigit() and 1900 <= int(answer) <= 2100: return answer
        print("Enter a four-digit year from 1900 through 2100.")

def build_known_names(ws, columns, master):
    names=set()
    for row in range(2, ws.max_row+1):
        raw=normalize_spaces(cell_value(ws,row,columns,"name"))
        if not raw:
            raw=full_name(cell_value(ws,row,columns,"given"), cell_value(ws,row,columns,"family"), cell_value(ws,row,columns,"married"))
        if raw: names.add(raw)
    if master and master.exists(): names.update(collect_names_from_workbook(master))
    return sorted(names,key=str.casefold)

def style_cleaned_sheet(ws):
    fill=PatternFill("solid",fgColor="B3A369"); font=Font(color="FFFFFF",bold=True)
    for cell in ws[1]:
        cell.fill=fill; cell.font=font; cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.freeze_panes="A2"
    widths=[23,18,23,23,11,30,24,42,28,42,42,42,42,42,42,42,42,22,12]
    for col,width in enumerate(widths,1): ws.column_dimensions[get_column_letter(col)].width=width
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)

def parse_args():
    p=argparse.ArgumentParser(description=__doc__); p.add_argument("--input",type=Path,required=True); p.add_argument("--output",type=Path)
    p.add_argument("--project-root",type=Path,default=Path(__file__).resolve().parent); p.add_argument("--master",type=Path)
    p.add_argument("--non-interactive",action="store_true"); p.add_argument("--accept-auto",action="store_true"); return p.parse_args()

def main():
    args=parse_args(); input_path=args.input.expanduser().resolve()
    if not input_path.exists(): print(f"Input workbook does not exist: {input_path}",file=sys.stderr); return 2
    project_root=args.project_root.expanduser().resolve(); master=args.master.expanduser().resolve() if args.master else project_root/"YJMB Trees.xlsx"
    output_path=args.output.expanduser().resolve() if args.output else input_path.with_name(f"{input_path.stem} - Four Name Columns.xlsx")
    if output_path==input_path: print("Output must differ from input.",file=sys.stderr); return 2
    project_root.mkdir(parents=True,exist_ok=True)
    resolver=InteractiveResolver(project_root/".name_resolution_cache.json",interactive=not args.non_interactive,accept_auto=args.accept_auto)
    shutil.copy2(input_path,output_path); workbook=load_workbook(output_path); source=find_source_sheet(workbook); columns=source_columns(source)
    known=build_known_names(source,columns,master)
    if CLEANED_SHEET in workbook.sheetnames: del workbook[CLEANED_SHEET]
    cleaned=workbook.create_sheet(CLEANED_SHEET,0); cleaned.append(OUTPUT_HEADERS); processed=skipped=0
    try:
        for row in range(2,source.max_row+1):
            raw=normalize_spaces(cell_value(source,row,columns,"name")); given=normalize_spaces(cell_value(source,row,columns,"given"))
            family=normalize_spaces(cell_value(source,row,columns,"family")); married=normalize_spaces(cell_value(source,row,columns,"married")); nickname=normalize_spaces(cell_value(source,row,columns,"nickname"))
            if not raw: raw=full_name(given,family,married)
            if not raw: continue
            context=f"{source.title} row {row}"
            try:
                parts=resolver.resolve_name(raw,context=context,existing_names=known,provided_given=given,provided_nickname=nickname,
                    provided_family=family,provided_married=married,
                    force_confirm=(len(raw.split())>=3 and not raw.casefold().endswith(" el akkad")) or has_ambiguous_annotation(raw) or bool(married))
                year=normalize_year(cell_value(source,row,columns,"rat_year"),context=context,interactive=not args.non_interactive,accept_auto=args.accept_auto)
                rel={}
                for field,label in [("vet","VET")]+[(f"rat_{i}",f"RAT {i}") for i in range(1,7)]:
                    rel[field]=resolver.resolve_relationship(cell_value(source,row,columns,field),context=f"{context} / {label}",known_names=known)
            except SkipRecord as exc: print(exc); skipped+=1; continue
            cleaned.append([parts.given,parts.nickname,parts.family,parts.married,year,normalize_spaces(cell_value(source,row,columns,"instrument")),"",
                normalize_spaces(cell_value(source,row,columns,"notes")),"",rel.get("vet",""),*[rel.get(f"rat_{i}","") for i in range(1,7)],"",
                cell_value(source,row,columns,"timestamp"),row]); processed+=1
    except UserQuit as exc:
        print(f"Stopped: {exc}"); workbook.close(); output_path.unlink(missing_ok=True); return 130
    finally: resolver.save()
    style_cleaned_sheet(cleaned)
    if cleaned.max_row>=2:
        table=Table(displayName="CleanedResponsesTable",ref=f"A1:S{cleaned.max_row}")
        table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False); cleaned.add_table(table)
    save_workbook_atomic(workbook,output_path); workbook.close()
    print(f"Created: {output_path}\nPrepared rows: {processed}\nSkipped rows: {skipped}\nDecision cache: {resolver.cache_path}"); return 0
if __name__=="__main__": raise SystemExit(main())
