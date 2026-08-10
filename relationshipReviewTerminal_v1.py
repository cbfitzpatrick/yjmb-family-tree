#!/usr/bin/env python3
"""
relationshipReviewTerminal.py

Terminal-only YJMB RAT/VET relationship review tool.

Scans ``YJMB Trees.xlsx`` (default worksheet: ``People on Tree``) for VET/RAT
references that do not resolve to exactly one unique person row.  For each
unresolved reference it can:

* show fuzzy/similar existing-person suggestions;
* let you choose an existing person or create a new person;
* let you correct name, RAT year, and section/instrument values;
* optionally make the reciprocal VET/RAT field consistent when safe;
* print a localized one-hop tree containing the reviewed person, their VET(s),
  and their RAT(s);
* ask ``Does this look correct?`` before saving;
* skip any unresolved item without changing it.

No fuzzy match is ever accepted automatically.  A timestamped workbook backup
is created before the first write.  The actual workbook is checked for a
Windows write lock; stale Excel ``~$...xlsx`` owner files are ignored.

Typical use::

    python .\relationshipReviewTerminal.py

Read-only scan::

    python .\relationshipReviewTerminal.py --scan-only
"""
from __future__ import annotations

import argparse
import copy
import difflib
import os
import re
import shutil
import sys
import tempfile
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries



SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "YJMB Trees.xlsx"
DEFAULT_SHEET = "People on Tree"

# The full-band project's existing section palette.
SECTION_COLORS: dict[str, str] = {
    "flute/piccolo": "#f5c1ce",
    "clarinet": "#ECD5E3",
    "sax/saxophone": "#F3B0C3",
    "trumpet": "#fa9078",
    "mellophone": "#fcb17e",
    "trombone": "#7cf7b7",
    "baritone": "#8bf0e6",
    "sousaphone": "#f797d2",
    "front ensemble": "#C6DBDA",
    "battery": "#D4F0F0",
    "guard": "#FF968A",
    "goldrush": "#F6EAC2",
    "golden girl": "#FFFFB5",
    "unknown": "#D9D9D9",
}

SECTION_ALIASES: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("flute/piccolo", re.compile(r"\b(?:flutes?|piccolos?)\b", re.I)),
    ("clarinet", re.compile(r"\bclarinets?\b", re.I)),
    ("sax/saxophone", re.compile(r"\b(?:sax(?:ophone)?s?|alto\s+sax(?:ophone)?s?|tenor\s+sax(?:ophone)?s?|baritone\s+sax(?:ophone)?s?|bari\s+sax(?:ophone)?s?)\b", re.I)),
    ("trumpet", re.compile(r"\btrumpets?\b", re.I)),
    ("mellophone", re.compile(r"\b(?:mellos?|mellophones?)\b", re.I)),
    ("trombone", re.compile(r"\b(?:trombones?|bones?)\b", re.I)),
    ("baritone", re.compile(r"\b(?:baritones?|euphoniums?)\b", re.I)),
    ("sousaphone", re.compile(r"\b(?:sousaphones?|tubas?)\b", re.I)),
    ("front ensemble", re.compile(r"\b(?:front\s+ensemble|pit)\b", re.I)),
    ("battery", re.compile(r"\b(?:battery|drum\s*line|drumline|snare|tenors?|quads?|bass\s+drums?|cymbals?)\b", re.I)),
    ("guard", re.compile(r"\b(?:color\s+guard|guard)\b", re.I)),
    ("goldrush", re.compile(r"\bgold\s*rush\b|\bgoldrush\b", re.I)),
    ("golden girl", re.compile(r"\bgolden\s+girls?\b", re.I)),
)

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "given": ("givenpreferredname", "givenname", "preferredname", "firstname"),
    "nickname": ("nickname", "nick name"),
    "family": ("familymaidenname", "familyname", "maidenname", "lastname", "surname"),
    "married": ("marriedname", "marriedsurname", "spousesurname", "currentlastname"),
    "year": ("ratyear", "year"),
    "instrument": ("instrument", "instruments", "section"),
    "vet": ("vet", "vetsnameratyearandinstruments"),
}

RELATION_RE = re.compile(r"^\s*(.*?)\s*\((.*?)\)\s*\(([^()]*)\)\s*$")
YEAR_RE = re.compile(r"(?<!\d)((?:19|20)\d{2})(?!\d)")

CARD_WIDTH = 150
CARD_HEIGHT = 80
CARD_GAP = 30
TREE_MARGIN_X = 105
TREE_MARGIN_Y = 70
YEAR_BAND_HEIGHT = 105
CONNECTOR_COLOR = "#777777"
CONNECTOR_WIDTH = 4


# ---------------------------------------------------------------------------
# Normalization and data objects
# ---------------------------------------------------------------------------


def normalize_spaces(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u00a0", " ").replace("’", "'").replace("`", "'")
    return re.sub(r"\s+", " ", text).strip()


def normalized_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_spaces(value).casefold())


def loose_name_key(value: object) -> str:
    text = unicodedata.normalize("NFKD", normalize_spaces(value).casefold())
    return "".join(char for char in text if char.isalnum())


def stable_name(given: str, family: str) -> str:
    return normalize_spaces(f"{given} {family}") or given or family


def parse_year(value: object) -> int | None:
    text = normalize_spaces(value)
    match = YEAR_RE.search(text)
    if not match:
        return None
    return int(match.group(1))


def normalize_sections(raw: object) -> list[str]:
    text = normalize_spaces(raw)
    if not text:
        return ["unknown"]
    found: list[tuple[int, str]] = []
    for section, pattern in SECTION_ALIASES:
        for match in pattern.finditer(text):
            found.append((match.start(), section))
    found.sort(key=lambda item: item[0])
    result: list[str] = []
    for _, section in found:
        if section not in result:
            result.append(section)
    return result or ["unknown"]


def section_overlap(left: object, right: object) -> bool:
    return bool(set(normalize_sections(left)) & set(normalize_sections(right)) - {"unknown"})


@dataclass(frozen=True)
class RelationReference:
    raw: str
    name: str
    year: int | None
    instrument: str


@dataclass
class PersonRecord:
    row: int
    given: str
    nickname: str
    family: str
    married: str
    year: int | None
    instrument: str
    vet_raw: str
    rat_raws: list[tuple[str, str]]

    @property
    def name(self) -> str:
        return stable_name(self.given, self.family)

    @property
    def aliases(self) -> set[str]:
        aliases = {self.name}
        if self.married:
            aliases.add(stable_name(self.given, self.married))
        if self.nickname:
            aliases.add(stable_name(self.nickname, self.family))
            if self.married:
                aliases.add(stable_name(self.nickname, self.married))
        return {alias for alias in aliases if alias}

    @property
    def canonical_relation(self) -> str:
        return format_relation(self.name, self.year, self.instrument)


@dataclass
class CandidateMatch:
    person: PersonRecord
    score: float
    explanation: str


@dataclass
class ReviewItem:
    source_row: int
    source_name: str
    column: int
    column_label: str
    role: str  # "VET" or "RAT"
    raw: str
    reference: RelationReference
    exact_candidate_rows: list[int] = field(default_factory=list)

    @property
    def key(self) -> tuple[int, int, str]:
        return (self.source_row, self.column, self.raw)


@dataclass
class ResolutionDraft:
    action: str  # existing | new
    target_row: int | None
    given: str
    family: str
    year: int | None
    instrument: str
    ensure_reciprocal: bool
    replace_conflicting_vet: bool
    update_all_references: bool

    @property
    def name(self) -> str:
        return stable_name(self.given, self.family)

    @property
    def canonical_relation(self) -> str:
        return format_relation(self.name, self.year, self.instrument)


@dataclass
class PreviewPerson:
    key: str
    name: str
    year: int | None
    instrument: str
    unresolved: bool = False
    proposed: bool = False


@dataclass
class LocalPreview:
    center: PreviewPerson
    vets: list[PreviewPerson]
    rats: list[PreviewPerson]
    warnings: list[str]


# ---------------------------------------------------------------------------
# Workbook model
# ---------------------------------------------------------------------------


def parse_relation(raw_value: object) -> RelationReference | None:
    raw = normalize_spaces(raw_value)
    if not raw:
        return None
    match = RELATION_RE.match(raw)
    if match:
        name = normalize_spaces(match.group(1))
        year = parse_year(match.group(2))
        instrument = normalize_spaces(match.group(3))
    else:
        name = normalize_spaces(raw.split(" (", 1)[0])
        year = parse_year(raw)
        instrument = ""
    if not name:
        return None
    return RelationReference(raw=raw, name=name, year=year, instrument=instrument)


def format_relation(name: str, year: int | None, instrument: str) -> str:
    return f"{normalize_spaces(name)} ({year or ''}) ({normalize_spaces(instrument)})"


class WorkbookModel:
    def __init__(self, path: Path, sheet_name: str = DEFAULT_SHEET) -> None:
        self.path = path.resolve()
        self.sheet_name = sheet_name
        self.workbook = None
        self.ws = None
        self.header_row = 1
        self.mapping: dict[str, int] = {}
        self.rat_columns: list[tuple[str, int]] = []
        self.header_labels: dict[int, str] = {}
        self.people: list[PersonRecord] = []
        self.people_by_row: dict[int, PersonRecord] = {}
        self._backup_path: Path | None = None
        self.load()

    def load(self) -> None:
        if self.workbook is not None:
            try:
                self.workbook.close()
            except Exception:
                pass
        self.workbook = load_workbook(self.path, data_only=False, read_only=False)
        if self.sheet_name not in self.workbook.sheetnames:
            possible = []
            for candidate in self.workbook.sheetnames:
                try:
                    self._discover_headers(self.workbook[candidate])
                except ValueError:
                    continue
                possible.append(candidate)
            if not possible:
                raise ValueError(
                    f"No worksheet contains the expected name/year/VET/RAT headers. "
                    f"Available sheets: {', '.join(self.workbook.sheetnames)}"
                )
            self.sheet_name = possible[0]
        self.ws = self.workbook[self.sheet_name]
        self.header_row, self.mapping, self.rat_columns = self._discover_headers(self.ws)
        self.header_labels = {
            col: normalize_spaces(self.ws.cell(self.header_row, col).value) or f"Column {col}"
            for col in range(1, self.ws.max_column + 1)
        }
        self.refresh_people()

    @staticmethod
    def _discover_headers(ws) -> tuple[int, dict[str, int], list[tuple[str, int]]]:
        header_row = None
        headers: dict[int, str] = {}
        for row in range(1, min(ws.max_row, 10) + 1):
            candidate = {
                col: normalized_header(ws.cell(row, col).value)
                for col in range(1, ws.max_column + 1)
            }
            values = set(candidate.values())
            if "givenpreferredname" in values and "familymaidenname" in values:
                header_row = row
                headers = candidate
                break
        if header_row is None:
            raise ValueError("Four-column name headers were not found.")

        mapping: dict[str, int] = {}
        for field_name, aliases in HEADER_ALIASES.items():
            keys = {normalized_header(alias) for alias in aliases}
            for col, header in headers.items():
                if header in keys:
                    mapping[field_name] = col
                    break
        required = ("given", "nickname", "family", "married", "year", "instrument", "vet")
        missing = [field_name for field_name in required if field_name not in mapping]
        if missing:
            raise ValueError("Missing required columns: " + ", ".join(missing))

        rat_columns: list[tuple[str, int]] = []
        for col, header in headers.items():
            match = re.fullmatch(r"rat(\d+)", header)
            if match:
                rat_columns.append((f"RAT {int(match.group(1))}", col))
        rat_columns.sort(key=lambda item: int(item[0].split()[1]))
        if not rat_columns:
            raise ValueError("No RAT relationship columns were found.")
        return header_row, mapping, rat_columns

    def refresh_people(self) -> None:
        assert self.ws is not None
        people: list[PersonRecord] = []
        for row in range(self.header_row + 1, self.ws.max_row + 1):
            given = normalize_spaces(self.ws.cell(row, self.mapping["given"]).value)
            nickname = normalize_spaces(self.ws.cell(row, self.mapping["nickname"]).value)
            family = normalize_spaces(self.ws.cell(row, self.mapping["family"]).value)
            married = normalize_spaces(self.ws.cell(row, self.mapping["married"]).value)
            if not any((given, nickname, family, married)):
                continue
            person = PersonRecord(
                row=row,
                given=given,
                nickname=nickname,
                family=family,
                married=married,
                year=parse_year(self.ws.cell(row, self.mapping["year"]).value),
                instrument=normalize_spaces(self.ws.cell(row, self.mapping["instrument"]).value),
                vet_raw=normalize_spaces(self.ws.cell(row, self.mapping["vet"]).value),
                rat_raws=[
                    (label, normalize_spaces(self.ws.cell(row, col).value))
                    for label, col in self.rat_columns
                    if normalize_spaces(self.ws.cell(row, col).value)
                ],
            )
            people.append(person)
        self.people = people
        self.people_by_row = {person.row: person for person in people}

    def unique_candidates(self, reference: RelationReference) -> list[PersonRecord]:
        """Return the strongest exact-name candidates for a relationship reference.

        A relationship is considered uniquely resolved when exactly one candidate
        survives name + (when available) RAT-year filtering. Instrument is used as
        an additional disambiguator only when it reduces a multi-candidate set.
        """
        key = loose_name_key(reference.name)
        candidates = [
            person
            for person in self.people
            if any(loose_name_key(alias) == key for alias in person.aliases)
        ]
        if reference.year is not None:
            by_year = [person for person in candidates if person.year == reference.year]
            if by_year:
                candidates = by_year
        if len(candidates) > 1 and reference.instrument:
            by_section = [
                person for person in candidates if section_overlap(reference.instrument, person.instrument)
            ]
            if by_section:
                candidates = by_section
        return candidates

    def resolve_unique(self, reference: RelationReference) -> PersonRecord | None:
        candidates = self.unique_candidates(reference)
        return candidates[0] if len(candidates) == 1 else None

    def scan_review_items(self, skipped: set[tuple[int, int, str]] | None = None) -> list[ReviewItem]:
        skipped = skipped or set()
        items: list[ReviewItem] = []
        for person in self.people:
            relation_cells = [("VET", self.mapping["vet"], person.vet_raw)]
            relation_cells.extend(
                (label, col, normalize_spaces(self.ws.cell(person.row, col).value))
                for label, col in self.rat_columns
                if normalize_spaces(self.ws.cell(person.row, col).value)
            )
            for label, col, raw in relation_cells:
                if not raw:
                    continue
                key = (person.row, col, raw)
                if key in skipped:
                    continue
                reference = parse_relation(raw)
                if not reference:
                    items.append(
                        ReviewItem(
                            source_row=person.row,
                            source_name=person.name,
                            column=col,
                            column_label=label,
                            role="VET" if label == "VET" else "RAT",
                            raw=raw,
                            reference=RelationReference(raw=raw, name=raw, year=None, instrument=""),
                            exact_candidate_rows=[],
                        )
                    )
                    continue
                candidates = self.unique_candidates(reference)
                if len(candidates) != 1:
                    items.append(
                        ReviewItem(
                            source_row=person.row,
                            source_name=person.name,
                            column=col,
                            column_label=label,
                            role="VET" if label == "VET" else "RAT",
                            raw=raw,
                            reference=reference,
                            exact_candidate_rows=[candidate.row for candidate in candidates],
                        )
                    )
        items.sort(key=lambda item: (item.source_row, item.column))
        return items

    def fuzzy_candidates(self, reference: RelationReference, limit: int = 10) -> list[CandidateMatch]:
        ref_name = loose_name_key(reference.name)
        scored: list[CandidateMatch] = []
        for person in self.people:
            best_name_ratio = 0.0
            for alias in person.aliases:
                ratio = difflib.SequenceMatcher(None, ref_name, loose_name_key(alias)).ratio()
                best_name_ratio = max(best_name_ratio, ratio)
            score = best_name_ratio * 100.0
            explanations = [f"name {best_name_ratio * 100:.0f}%"]
            if reference.year is not None and person.year is not None:
                delta = abs(reference.year - person.year)
                if delta == 0:
                    score += 18
                    explanations.append("same RAT year")
                elif delta == 1:
                    score += 8
                    explanations.append("RAT year ±1")
                elif delta <= 3:
                    score += 3
            if reference.instrument and person.instrument and section_overlap(reference.instrument, person.instrument):
                score += 9
                explanations.append("same section")
            # Exact candidate rows (for duplicate-name ambiguity) should always be visible.
            if person.row in self.unique_candidates(reference):
                score += 25
                explanations.append("exact-name candidate")
            scored.append(CandidateMatch(person=person, score=score, explanation=", ".join(explanations)))
        scored.sort(key=lambda match: (-match.score, match.person.year or 9999, match.person.family.casefold(), match.person.given.casefold()))
        return scored[:limit]

    def relationship_cells_resolving_to(self, target_row: int) -> list[tuple[int, int]]:
        locations: list[tuple[int, int]] = []
        for person in self.people:
            columns = [self.mapping["vet"], *(col for _, col in self.rat_columns)]
            for col in columns:
                raw = normalize_spaces(self.ws.cell(person.row, col).value)
                reference = parse_relation(raw)
                if not reference:
                    continue
                resolved = self.resolve_unique(reference)
                if resolved and resolved.row == target_row:
                    locations.append((person.row, col))
        return locations

    def build_preview(self, item: ReviewItem, draft: ResolutionDraft) -> LocalPreview:
        source = self.people_by_row[item.source_row]
        if draft.action == "existing" and draft.target_row is not None:
            original = self.people_by_row[draft.target_row]
            center = PreviewPerson(
                key=f"row-{original.row}",
                name=draft.name,
                year=draft.year,
                instrument=draft.instrument,
                proposed=True,
            )
            existing_vet_raw = original.vet_raw
            existing_rat_raws = list(original.rat_raws)
        else:
            center = PreviewPerson(
                key="new-person",
                name=draft.name,
                year=draft.year,
                instrument=draft.instrument,
                proposed=True,
            )
            existing_vet_raw = ""
            existing_rat_raws = []

        warnings: list[str] = []
        vets: list[PreviewPerson] = []
        rats: list[PreviewPerson] = []

        # Existing one-hop relationships for the target person.
        if existing_vet_raw:
            ref = parse_relation(existing_vet_raw)
            if ref:
                resolved = self.resolve_unique(ref)
                vets.append(self._preview_from_ref(ref, resolved, suffix="existing-vet"))

        for index, (_, raw) in enumerate(existing_rat_raws, start=1):
            ref = parse_relation(raw)
            if not ref:
                continue
            resolved = self.resolve_unique(ref)
            rats.append(self._preview_from_ref(ref, resolved, suffix=f"existing-rat-{index}"))

        source_preview = PreviewPerson(
            key=f"source-{source.row}",
            name=source.name,
            year=source.year,
            instrument=source.instrument,
            proposed=True,
        )

        if item.role == "VET":
            # The reviewed target is the source person's VET, so source is a RAT of center.
            if not any(self._same_preview_person(person, source_preview) for person in rats):
                rats.append(source_preview)
        else:
            # The reviewed target is a RAT of source, so source is center's VET.
            existing_different = [
                person for person in vets if not self._same_preview_person(person, source_preview)
            ]
            if existing_different:
                warnings.append(
                    "The selected/new RAT already has a different VET in its own row. "
                    "The proposed VET is shown too. Reciprocal replacement will occur only if you enable it."
                )
            if not any(self._same_preview_person(person, source_preview) for person in vets):
                vets.append(source_preview)

        vets.sort(key=lambda person: (person.year or 9999, person.name.casefold()))
        rats.sort(key=lambda person: (person.year or 9999, person.name.casefold()))
        return LocalPreview(center=center, vets=vets, rats=rats, warnings=warnings)

    def _preview_from_ref(
        self, reference: RelationReference, resolved: PersonRecord | None, *, suffix: str
    ) -> PreviewPerson:
        if resolved:
            return PreviewPerson(
                key=f"row-{resolved.row}-{suffix}",
                name=resolved.name,
                year=resolved.year,
                instrument=resolved.instrument,
                unresolved=False,
            )
        return PreviewPerson(
            key=f"unresolved-{suffix}-{loose_name_key(reference.name)}",
            name=reference.name,
            year=reference.year,
            instrument=reference.instrument,
            unresolved=True,
        )

    @staticmethod
    def _same_preview_person(left: PreviewPerson, right: PreviewPerson) -> bool:
        return (
            loose_name_key(left.name) == loose_name_key(right.name)
            and (left.year is None or right.year is None or left.year == right.year)
        )

    def validate_draft(self, item: ReviewItem, draft: ResolutionDraft) -> list[str]:
        errors: list[str] = []
        if not draft.given.strip():
            errors.append("Given/Preferred Name is required.")
        if not draft.family.strip():
            errors.append("Family/Maiden Name is required.")
        if draft.year is None:
            errors.append("A four-digit RAT year is required for a unique person entry.")
        else:
            current_year = datetime.now().year
            if not (1908 <= draft.year <= current_year + 1):
                errors.append(f"RAT year must be between 1908 and {current_year + 1}.")
        if draft.action == "existing":
            if draft.target_row is None or draft.target_row not in self.people_by_row:
                errors.append("Select an existing person from the suggestion list.")
        elif draft.action == "new":
            conflicts = [
                person
                for person in self.people
                if loose_name_key(person.name) == loose_name_key(draft.name)
                and person.year == draft.year
            ]
            if conflicts:
                errors.append(
                    "A person with the same name and RAT year already exists. "
                    "Choose that existing row or correct the new-person details."
                )
        else:
            errors.append("Choose whether this is an existing person or a new person.")

        if item.role == "RAT" and draft.action == "existing" and draft.target_row:
            target = self.people_by_row[draft.target_row]
            if target.vet_raw:
                ref = parse_relation(target.vet_raw)
                resolved = self.resolve_unique(ref) if ref else None
                if resolved and resolved.row != item.source_row and draft.ensure_reciprocal and not draft.replace_conflicting_vet:
                    # Not a hard error: the source relationship can still be corrected while
                    # leaving the target's existing VET untouched. Warn in GUI instead.
                    pass
        return errors

    def commit_resolution(self, item: ReviewItem, draft: ResolutionDraft) -> tuple[int, list[str]]:
        """Apply a confirmed resolution and save atomically.

        Returns ``(target_row, messages)``.

        The workbook is checked for an *actual* Windows write lock before any
        in-memory edits are made.  Excel's ``~$...xlsx`` owner file is only a
        hint and can be left behind after a crash, so its mere presence must
        not prevent a save.
        """
        assert self.ws is not None
        self._assert_workbook_writable()
        messages: list[str] = []
        source = self.people_by_row[item.source_row]

        # Collect relationship locations before editing an existing person's identity.
        related_locations: list[tuple[int, int]] = []
        if draft.action == "existing" and draft.target_row is not None and draft.update_all_references:
            related_locations = self.relationship_cells_resolving_to(draft.target_row)

        if draft.action == "new":
            target_row = self._append_new_person(draft)
            messages.append(f"Created new person row {target_row}: {draft.name}.")
        else:
            assert draft.target_row is not None
            target_row = draft.target_row
            self._update_person_identity(target_row, draft)
            messages.append(f"Updated existing person row {target_row}: {draft.name}.")

        # Rewrite the currently reviewed source cell to the target's confirmed identity.
        canonical = draft.canonical_relation
        self.ws.cell(item.source_row, item.column).value = canonical
        messages.append(
            f"Updated {item.source_name} {item.column_label} to {canonical}."
        )

        # If an existing person's identity/year/section changed, keep all references
        # that uniquely pointed to that row consistent with the new canonical text.
        if draft.action == "existing" and draft.update_all_references:
            for row, col in related_locations:
                self.ws.cell(row, col).value = canonical
            if related_locations:
                messages.append(
                    f"Updated {len(related_locations)} other relationship reference(s) to the confirmed identity."
                )

        reciprocal_written = False
        if draft.ensure_reciprocal:
            if item.role == "VET":
                reciprocal_written = self._ensure_rat_reference(
                    parent_row=target_row,
                    child_row=item.source_row,
                    messages=messages,
                )
            else:
                reciprocal_written = self._ensure_vet_reference(
                    child_row=target_row,
                    parent_row=item.source_row,
                    replace_conflict=draft.replace_conflicting_vet,
                    messages=messages,
                )

        self._mark_status_if_present(item, reciprocal_written)
        self._backup_once()
        self._save_atomic()
        self.refresh_people()
        return target_row, messages

    def _update_person_identity(self, row: int, draft: ResolutionDraft) -> None:
        self.ws.cell(row, self.mapping["given"]).value = draft.given
        self.ws.cell(row, self.mapping["family"]).value = draft.family
        self.ws.cell(row, self.mapping["year"]).value = draft.year
        self.ws.cell(row, self.mapping["instrument"]).value = draft.instrument

    def _append_new_person(self, draft: ResolutionDraft) -> int:
        assert self.ws is not None
        old_max_row = self.ws.max_row
        row = old_max_row + 1

        # Copy style from the previous populated person row for visual consistency.
        source_style_row = max(self.header_row + 1, old_max_row)
        for col in range(1, self.ws.max_column + 1):
            src = self.ws.cell(source_style_row, col)
            dst = self.ws.cell(row, col)
            if src.has_style:
                dst._style = copy.copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)

        self.ws.cell(row, self.mapping["given"]).value = draft.given
        self.ws.cell(row, self.mapping["nickname"]).value = None
        self.ws.cell(row, self.mapping["family"]).value = draft.family
        self.ws.cell(row, self.mapping["married"]).value = None
        self.ws.cell(row, self.mapping["year"]).value = draft.year
        self.ws.cell(row, self.mapping["instrument"]).value = draft.instrument
        self.ws.cell(row, self.mapping["vet"]).value = None
        for _, col in self.rat_columns:
            self.ws.cell(row, col).value = None

        # Extend Excel tables that cover the master data area.
        for table in self.ws.tables.values():
            try:
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            except ValueError:
                continue
            if min_row <= self.header_row <= max_row and max_row == old_max_row:
                table.ref = (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(max_col)}{row}"
                )
        if self.ws.auto_filter and self.ws.auto_filter.ref:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(self.ws.auto_filter.ref)
                if max_row == old_max_row:
                    self.ws.auto_filter.ref = (
                        f"{get_column_letter(min_col)}{min_row}:"
                        f"{get_column_letter(max_col)}{row}"
                    )
            except ValueError:
                pass
        return row

    def _person_relation_for_row(self, row: int) -> str:
        given = normalize_spaces(self.ws.cell(row, self.mapping["given"]).value)
        family = normalize_spaces(self.ws.cell(row, self.mapping["family"]).value)
        year = parse_year(self.ws.cell(row, self.mapping["year"]).value)
        instrument = normalize_spaces(self.ws.cell(row, self.mapping["instrument"]).value)
        return format_relation(stable_name(given, family), year, instrument)

    def _ensure_rat_reference(self, parent_row: int, child_row: int, messages: list[str]) -> bool:
        child_relation = self._person_relation_for_row(child_row)
        child_ref = parse_relation(child_relation)
        # Existing equivalent reference?
        for label, col in self.rat_columns:
            raw = normalize_spaces(self.ws.cell(parent_row, col).value)
            if not raw:
                continue
            ref = parse_relation(raw)
            if ref and child_ref and loose_name_key(ref.name) == loose_name_key(child_ref.name) and (
                ref.year is None or child_ref.year is None or ref.year == child_ref.year
            ):
                self.ws.cell(parent_row, col).value = child_relation
                messages.append(f"Confirmed reciprocal {label} on row {parent_row}.")
                return True
        # First blank RAT slot.
        for label, col in self.rat_columns:
            if not normalize_spaces(self.ws.cell(parent_row, col).value):
                self.ws.cell(parent_row, col).value = child_relation
                messages.append(f"Added reciprocal {label} on row {parent_row}.")
                return True
        messages.append(
            f"Could not add reciprocal RAT to row {parent_row}: all existing RAT columns are occupied."
        )
        return False

    def _ensure_vet_reference(
        self,
        child_row: int,
        parent_row: int,
        *,
        replace_conflict: bool,
        messages: list[str],
    ) -> bool:
        parent_relation = self._person_relation_for_row(parent_row)
        vet_cell = self.ws.cell(child_row, self.mapping["vet"])
        existing = normalize_spaces(vet_cell.value)
        if not existing:
            vet_cell.value = parent_relation
            messages.append(f"Added reciprocal VET on row {child_row}.")
            return True
        existing_ref = parse_relation(existing)
        parent_ref = parse_relation(parent_relation)
        if existing_ref and parent_ref and loose_name_key(existing_ref.name) == loose_name_key(parent_ref.name) and (
            existing_ref.year is None or parent_ref.year is None or existing_ref.year == parent_ref.year
        ):
            vet_cell.value = parent_relation
            messages.append(f"Confirmed reciprocal VET on row {child_row}.")
            return True
        if replace_conflict:
            vet_cell.value = parent_relation
            messages.append(
                f"Replaced conflicting VET on row {child_row}: {existing!r} -> {parent_relation!r}."
            )
            return True
        messages.append(
            f"Left row {child_row}'s existing VET unchanged because it conflicts with the proposed relationship: {existing!r}."
        )
        return False

    def _mark_status_if_present(self, item: ReviewItem, reciprocal_written: bool) -> None:
        """If v9-style per-slot status columns already exist, update the source status."""
        status_label = (
            "VET Relationship Status"
            if item.role == "VET"
            else f"{item.column_label} Relationship Status"
        )
        status_key = normalized_header(status_label)
        status_col = None
        for col in range(1, self.ws.max_column + 1):
            if normalized_header(self.ws.cell(self.header_row, col).value) == status_key:
                status_col = col
                break
        if status_col:
            self.ws.cell(item.source_row, status_col).value = (
                "Reciprocated — validated on both profiles"
                if reciprocal_written
                else "Unreciprocated — pending validation from related profile"
            )

    def _backup_once(self) -> None:
        if self._backup_path is not None:
            return
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = self.path.parent / "backups" / "relationship_review" / stamp
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / self.path.name
        shutil.copy2(self.path, backup_path)
        self._backup_path = backup_path

    def _assert_workbook_writable(self) -> None:
        """Raise PermissionError only when the workbook itself is really locked.

        Excel normally creates an owner file named ``~$<workbook>.xlsx`` while
        a workbook is open.  Those owner files occasionally remain after Excel
        closes or crashes, so checking only for that filename creates false
        positives.  Instead, try to open the actual workbook for read/write
        access.  On Windows, also attempt a non-blocking one-byte lock.
        """
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")

        handle = None
        byte_locked = False
        try:
            # On Windows this open normally fails immediately when Excel has
            # opened the workbook with write sharing denied.
            handle = open(self.path, "r+b")

            if os.name == "nt":
                try:
                    import msvcrt

                    handle.seek(0)
                    msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
                    byte_locked = True
                except (OSError, PermissionError) as exc:
                    raise PermissionError(
                        f"{self.path.name} is currently locked by another program. "
                        "Close the workbook in Excel (and any File Explorer preview pane "
                        "or other program using it), then try again."
                    ) from exc
                finally:
                    if byte_locked:
                        try:
                            handle.seek(0)
                            msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
                        except OSError:
                            pass
        except PermissionError as exc:
            # Preserve our clearer message above when one was already created.
            if str(exc).startswith(self.path.name):
                raise
            raise PermissionError(
                f"{self.path.name} is currently locked by another program. "
                "Close the workbook in Excel (and any File Explorer preview pane "
                "or other program using it), then try again."
            ) from exc
        finally:
            if handle is not None:
                try:
                    handle.close()
                except OSError:
                    pass

    def _save_atomic(self) -> None:
        # Re-check immediately before replacement in case Excel was opened after
        # the user entered the confirmation screen.  Do NOT block merely because
        # a stale ``~$...xlsx`` owner file exists.
        self._assert_workbook_writable()

        tmp_fd, tmp_name = tempfile.mkstemp(
            prefix=f".{self.path.stem}_relationship_review_",
            suffix=self.path.suffix,
            dir=self.path.parent,
        )
        os.close(tmp_fd)
        tmp_path = Path(tmp_name)
        try:
            self.workbook.save(tmp_path)
            try:
                os.replace(tmp_path, self.path)
            except PermissionError as exc:
                raise PermissionError(
                    f"Windows would not replace {self.path.name}. Another program may "
                    "have opened the workbook while this review was in progress. "
                    "Close it and try the save again."
                ) from exc
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except OSError:
                    pass


# ---------------------------------------------------------------------------
# Terminal workflow
# ---------------------------------------------------------------------------


def split_name_guess(name: str) -> tuple[str, str]:
    """Conservative prefill only; the user always gets a chance to correct it."""
    parts = normalize_spaces(name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    lower = [part.casefold() for part in parts]
    surname_prefixes = {"el", "de", "del", "van", "von", "la", "le", "st.", "saint"}
    if len(parts) >= 3 and lower[-2] in surname_prefixes:
        return " ".join(parts[:-2]), " ".join(parts[-2:])
    return " ".join(parts[:-1]), parts[-1]


def hr(char: str = "─", width: int = 88) -> str:
    return char * width


def prompt_choice(prompt: str, valid: set[str], *, default: str | None = None) -> str:
    valid_lower = {v.casefold() for v in valid}
    while True:
        try:
            raw = input(prompt).strip()
        except (EOFError, KeyboardInterrupt):
            print("\nQuit requested.")
            return "q" if "q" in valid_lower else next(iter(valid_lower))
        if not raw and default is not None:
            return default.casefold()
        choice = raw.casefold()
        if choice in valid_lower:
            return choice
        print(f"Please enter one of: {', '.join(sorted(valid))}.")


def prompt_yes_no(prompt: str, *, default: bool = True) -> bool:
    marker = "[Y/n]" if default else "[y/N]"
    while True:
        try:
            raw = input(f"{prompt} {marker} ").strip().casefold()
        except (EOFError, KeyboardInterrupt):
            print()
            return default
        if not raw:
            return default
        if raw in {"y", "yes"}:
            return True
        if raw in {"n", "no"}:
            return False
        print("Enter y or n.")


def prompt_text(label: str, current: str = "", *, required: bool = False) -> str:
    while True:
        suffix = f" [{current}]" if current else ""
        try:
            raw = input(f"{label}{suffix}: ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            return current
        value = normalize_spaces(raw) if raw else current
        if required and not value:
            print(f"{label} cannot be blank.")
            continue
        return value


def prompt_year(current: int | None) -> int | None:
    current_text = str(current) if current is not None else ""
    while True:
        value = prompt_text("RAT Year", current_text)
        if not value:
            return None
        parsed = parse_year(value)
        if parsed is None:
            print("Enter a four-digit year, e.g. 2018.")
            continue
        current_year = datetime.now().year
        if not (1908 <= parsed <= current_year + 1):
            print(f"Year must be between 1908 and {current_year + 1}.")
            continue
        return parsed


def format_person_line(person: PreviewPerson | PersonRecord, prefix: str = "") -> str:
    year = person.year if person.year is not None else "?"
    instrument = normalize_spaces(person.instrument) or "Unknown section"
    unresolved = " [UNRESOLVED]" if getattr(person, "unresolved", False) else ""
    proposed = " [REVIEWED]" if getattr(person, "proposed", False) else ""
    return f"{prefix}[{year}] {person.name} — {instrument}{unresolved}{proposed}"


def print_local_tree(preview: LocalPreview) -> None:
    """Print a readable one-hop tree; this is a confirmation preview, not layout output."""
    print("\n" + hr("═"))
    print("LOCALIZED TREE PREVIEW")
    print(hr("═"))

    if preview.vets:
        print("VET(S):")
        for vet in preview.vets:
            print(format_person_line(vet, "  ┌─ "))
        print("  │")
    else:
        print("VET(S): none")
        print("  │")

    print(format_person_line(preview.center, "  ●  "))

    if preview.rats:
        print("  │")
        print("RAT(S):")
        for rat in preview.rats:
            print(format_person_line(rat, "  └─ "))
    else:
        print("  │")
        print("RAT(S): none")

    if preview.warnings:
        print("\nWarnings:")
        for warning in preview.warnings:
            print(f"  ! {warning}")
    print(hr("═"))


def print_item(model: WorkbookModel, item: ReviewItem, index: int, total: int) -> list[CandidateMatch]:
    print("\n" + hr("═"))
    print(f"RELATIONSHIP REVIEW  {index}/{total}")
    print(hr("═"))
    print(f"Person whose cell needs review : {item.source_name} (worksheet row {item.source_row})")
    print(f"Relationship field             : {item.column_label}")
    print(f"Current spreadsheet value      : {item.raw}")
    print(f"Parsed referenced name         : {item.reference.name or '(unknown)'}")
    print(f"Parsed RAT year                : {item.reference.year or '(unknown)'}")
    print(f"Parsed section/instrument      : {item.reference.instrument or '(unknown)'}")

    candidates = model.fuzzy_candidates(item.reference, limit=10)
    print("\nPossible existing people:")
    if not candidates:
        print("  (No candidate rows found.)")
    else:
        print(f"  {'#':>2}  {'Score':>5}  {'Row':>4}  {'RAT':>4}  {'Name':<31}  {'Section / instrument'}")
        print("  " + hr("-", 80))
        for number, match in enumerate(candidates, 1):
            p = match.person
            print(
                f"  {number:>2}  {match.score:>5.0f}  {p.row:>4}  "
                f"{str(p.year or '?'):>4}  {p.name[:31]:<31}  {p.instrument or '?'}"
            )
            print(f"      match basis: {match.explanation}")
    return candidates


def draft_from_existing(model: WorkbookModel, candidate: CandidateMatch, item: ReviewItem) -> ResolutionDraft:
    p = candidate.person
    print(f"\nSelected existing row {p.row}: {p.name} ({p.year or '?'}) ({p.instrument or '?'})")
    edit = prompt_yes_no("Does anything about this person's name/year/section need to be corrected?", default=False)
    given, family, year, instrument = p.given, p.family, p.year, p.instrument
    if edit:
        given = prompt_text("First / Preferred Name", given, required=True)
        family = prompt_text("Last / Family Name", family, required=True)
        year = prompt_year(year)
        instrument = prompt_text("Section / Instrument", instrument)

    ensure = prompt_yes_no("Ensure the reciprocal VET/RAT field is present when safe?", default=True)
    replace_conflict = False
    if item.role == "RAT" and ensure:
        target = model.people_by_row[p.row]
        if target.vet_raw:
            ref = parse_relation(target.vet_raw)
            resolved = model.resolve_unique(ref) if ref else None
            if resolved and resolved.row != item.source_row:
                print(f"\nWARNING: {p.name} already has a different VET: {target.vet_raw}")
                replace_conflict = prompt_yes_no(
                    "If this relationship is confirmed, replace that conflicting existing VET?",
                    default=False,
                )

    update_all = True
    if edit:
        update_all = prompt_yes_no(
            "Update other relationship references that already uniquely point to this existing person?",
            default=True,
        )

    return ResolutionDraft(
        action="existing",
        target_row=p.row,
        given=given,
        family=family,
        year=year,
        instrument=instrument,
        ensure_reciprocal=ensure,
        replace_conflicting_vet=replace_conflict,
        update_all_references=update_all,
    )


def draft_new_person(item: ReviewItem) -> ResolutionDraft:
    guessed_given, guessed_family = split_name_guess(item.reference.name)
    print("\nCreate a new unique person row. Correct any prefilled values before previewing.")
    given = prompt_text("First / Preferred Name", guessed_given, required=True)
    family = prompt_text("Last / Family Name", guessed_family, required=True)
    year = prompt_year(item.reference.year)
    instrument = prompt_text("Section / Instrument", item.reference.instrument)
    ensure = prompt_yes_no("Ensure the reciprocal VET/RAT field is present when safe?", default=True)
    return ResolutionDraft(
        action="new",
        target_row=None,
        given=given,
        family=family,
        year=year,
        instrument=instrument,
        ensure_reciprocal=ensure,
        replace_conflicting_vet=False,
        update_all_references=False,
    )


def edit_draft(draft: ResolutionDraft) -> ResolutionDraft:
    print("\nRevise proposed person details:")
    draft.given = prompt_text("First / Preferred Name", draft.given, required=True)
    draft.family = prompt_text("Last / Family Name", draft.family, required=True)
    draft.year = prompt_year(draft.year)
    draft.instrument = prompt_text("Section / Instrument", draft.instrument)
    return draft


def review_one(model: WorkbookModel, item: ReviewItem, index: int, total: int) -> str:
    """Return saved | skipped | quit."""
    while True:
        candidates = print_item(model, item, index, total)
        print("\nChoose what this reference means:")
        print("  1-10  Use one of the possible existing people above")
        print("  N     This is a new person who needs their own unique row")
        print("  S     Skip this reference; make no changes")
        print("  Q     Quit the reviewer")
        try:
            raw = input("Choice: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nQuit requested.")
            return "quit"

        low = raw.casefold()
        if low in {"q", "quit"}:
            return "quit"
        if low in {"s", "skip"}:
            return "skipped"
        if low in {"n", "new"}:
            draft = draft_new_person(item)
        else:
            try:
                candidate_index = int(raw)
            except ValueError:
                print("Invalid choice. Enter a candidate number, N, S, or Q.")
                continue
            if not (1 <= candidate_index <= len(candidates)):
                print(f"Choose a candidate between 1 and {len(candidates)}, or N/S/Q.")
                continue
            draft = draft_from_existing(model, candidates[candidate_index - 1], item)

        while True:
            errors = model.validate_draft(item, draft)
            if errors:
                print("\nCannot preview/save this proposal yet:")
                for error in errors:
                    print(f"  - {error}")
                action = prompt_choice("[R]evise details, [B]ack to match selection, [S]kip, [Q]uit: ", {"r", "b", "s", "q"})
                if action == "r":
                    draft = edit_draft(draft)
                    continue
                if action == "b":
                    break
                if action == "s":
                    return "skipped"
                return "quit"

            preview = model.build_preview(item, draft)
            print_local_tree(preview)
            print("Does this look correct?")
            print("  Y  Yes — save this resolution and continue")
            print("  N  No — revise the proposed details")
            print("  B  Back — choose a different existing/new person")
            print("  S  Skip — leave this spreadsheet reference unchanged")
            print("  Q  Quit")
            confirm = prompt_choice("Choice [Y/n/b/s/q]: ", {"y", "n", "b", "s", "q"}, default="y")
            if confirm == "n":
                draft = edit_draft(draft)
                continue
            if confirm == "b":
                break
            if confirm == "s":
                return "skipped"
            if confirm == "q":
                return "quit"

            try:
                _, messages = model.commit_resolution(item, draft)
            except PermissionError as exc:
                print(f"\nSAVE BLOCKED: {exc}")
                retry = prompt_choice("[R]etry save, [B]ack, [S]kip, [Q]uit: ", {"r", "b", "s", "q"})
                if retry == "r":
                    continue
                if retry == "b":
                    break
                if retry == "s":
                    return "skipped"
                return "quit"
            except Exception as exc:
                print(f"\nERROR while saving: {exc}")
                return "quit"

            print("\nSaved successfully:")
            for message in messages:
                print(f"  - {message}")
            if model._backup_path:
                print(f"  Backup: {model._backup_path}")
            return "saved"

        # Inner loop broke to choose an existing/new person again.
        continue


def print_scan(model: WorkbookModel, items: list[ReviewItem]) -> None:
    print(f"Workbook: {model.path}")
    print(f"Sheet: {model.sheet_name}")
    print(f"People rows: {len(model.people)}")
    print(f"Unresolved/non-unique references: {len(items)}")
    for item in items:
        candidates = model.fuzzy_candidates(item.reference, limit=3)
        suggestion = "; ".join(
            f"{match.person.name} ({match.person.year or '?'}) score={match.score:.0f}"
            for match in candidates
        )
        print(
            f"- row {item.source_row} {item.source_name} {item.column_label}: "
            f"{item.raw!r} | suggestions: {suggestion or 'none'}"
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Terminal reviewer for missing/ambiguous YJMB VET/RAT references."
    )
    parser.add_argument("--workbook", type=Path, default=None, help="Path to YJMB Trees.xlsx")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Master worksheet (default: {DEFAULT_SHEET!r})")
    parser.add_argument("--scan-only", action="store_true", help="Print unresolved references and exit without modifying the workbook")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve() if args.workbook else DEFAULT_WORKBOOK
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        print("Use --workbook \"C:\\path\\to\\YJMB Trees.xlsx\" to select a different file.", file=sys.stderr)
        return 2

    try:
        model = WorkbookModel(workbook_path, args.sheet)
    except Exception as exc:
        print(f"ERROR opening workbook: {exc}", file=sys.stderr)
        return 2

    items = model.scan_review_items()
    if args.scan_only:
        print_scan(model, items)
        return 0

    print("YJMB VET/RAT TERMINAL RELATIONSHIP REVIEWER")
    print(hr("═"))
    print(f"Workbook : {model.path}")
    print(f"Worksheet: {model.sheet_name}")
    print(f"People   : {len(model.people)}")
    print(f"Items needing review: {len(items)}")
    print("\nNo fuzzy match is automatic. A change is saved only after you answer Yes to")
    print("'Does this look correct?'. Skip makes no change to that reference.")

    if not items:
        print("\nNo unresolved or non-unique VET/RAT references were found.")
        return 0

    skipped_keys: set[tuple[int, int, str]] = set()
    saved = 0
    skipped = 0

    while True:
        # Rescan after each write because one correction can make several other
        # relationship references resolve automatically.
        current_items = model.scan_review_items(skipped_keys)
        if not current_items:
            break
        item = current_items[0]
        total = saved + skipped + len(current_items)
        result = review_one(model, item, saved + skipped + 1, total)
        if result == "quit":
            print("\nReview stopped by user.")
            break
        if result == "skipped":
            skipped_keys.add(item.key)
            skipped += 1
            print("Skipped. The spreadsheet was not changed for this reference.")
        elif result == "saved":
            saved += 1

    remaining_all = model.scan_review_items()
    print("\n" + hr("═"))
    print("REVIEW SUMMARY")
    print(hr("═"))
    print(f"Saved resolutions this run : {saved}")
    print(f"Skipped this run           : {skipped}")
    print(f"Still unresolved in workbook: {len(remaining_all)}")
    if model._backup_path:
        print(f"Safety backup              : {model._backup_path}")
    if remaining_all:
        print("\nUse --scan-only to list what remains, or rerun the reviewer to revisit skipped items.")
    else:
        print("\nAll VET/RAT references now resolve to unique person rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
