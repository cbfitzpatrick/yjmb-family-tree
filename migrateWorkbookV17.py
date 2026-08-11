#!/usr/bin/env python3
"""Prepare an existing YJMB workbook for v17 without overwriting member data.

Default: scan-only. Use --apply to append missing compatibility columns. A
backup is created before the first write. Existing columns and cell values are
never renamed or replaced.
"""
from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from yjmb_schema import V17_OPTIONAL_COLUMNS, ensure_optional_columns, find_header_row, missing_optional_columns

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "YJMB Trees.xlsx"
DEFAULT_SHEET = "People on Tree"


def backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    folder = path.parent / "backups" / "data_cleanup" / stamp
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / path.name
    shutil.copy2(path, target)
    return target


def main() -> int:
    ap = argparse.ArgumentParser(description="Scan/apply the additive v17 workbook schema migration.")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--apply", action="store_true", help="Append missing v17 compatibility columns after making a backup.")
    args = ap.parse_args()

    path = args.workbook.expanduser().resolve()
    wb = load_workbook(path)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Worksheet not found: {args.sheet}")
    ws = wb[args.sheet]
    header_row = find_header_row(ws)
    missing = missing_optional_columns(ws, header_row)

    print(f"Workbook: {path}")
    print(f"Worksheet: {args.sheet}")
    print(f"v17 compatibility fields present: {len(V17_OPTIONAL_COLUMNS) - len(missing)}/{len(V17_OPTIONAL_COLUMNS)}")
    if missing:
        print("Missing fields that v17 can append without touching existing values:")
        for name in missing:
            print(f"  + {name}")
    else:
        print("No v17 compatibility columns are missing.")

    if not args.apply:
        print("\nScan only. Re-run with --apply to append only the missing columns.")
        wb.close()
        return 0
    if not missing:
        wb.close()
        return 0

    backup_path = backup(path)
    _, added = ensure_optional_columns(ws, header_row=header_row)
    wb.save(path)
    wb.close()
    print(f"\nAdded {len(added)} column(s). Existing cell values were not changed.")
    print(f"Backup: {backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
